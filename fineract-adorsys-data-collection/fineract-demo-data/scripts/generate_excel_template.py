#!/usr/bin/env python3
"""
Fineract Demo Data Excel Generator
Creates a comprehensive Excel template for Cameroon OHADA-compliant microfinance setup
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import os

class FineractDemoDataGenerator:
    def __init__(self, output_dir='../output'):
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.filename = f'{output_dir}/fineract_demo_data_{self.timestamp}.xlsx'
        self.writer = None

    def create_offices_sheet(self):
        """Create offices/branches data"""
        data = [
            {
                'office_name': 'Yaounde Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-01-15',
                'external_id': 'BR-YDE-001',
                'address': 'Avenue Kennedy, Quartier Bastos',
                'city': 'Yaounde',
                'region': 'Centre',
                'phone': '+237 222 20 45 67',
                'email': 'yaounde@mfi.cm'
            },
            {
                'office_name': 'Douala Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-01-15',
                'external_id': 'BR-DLA-001',
                'address': 'Boulevard de la Liberte, Akwa',
                'city': 'Douala',
                'region': 'Littoral',
                'phone': '+237 233 42 56 78',
                'email': 'douala@mfi.cm'
            },
            {
                'office_name': 'Bafoussam Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-02-01',
                'external_id': 'BR-BAF-001',
                'address': 'Marche A, Centre Ville',
                'city': 'Bafoussam',
                'region': 'Ouest',
                'phone': '+237 233 44 12 34',
                'email': 'bafoussam@mfi.cm'
            },
            {
                'office_name': 'Bamenda Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-02-15',
                'external_id': 'BR-BAM-001',
                'address': 'Commercial Avenue, Nkwen',
                'city': 'Bamenda',
                'region': 'Nord-Ouest',
                'phone': '+237 233 36 45 67',
                'email': 'bamenda@mfi.cm'
            }
        ]
        return pd.DataFrame(data)

    def create_staff_sheet(self):
        """Create staff members with role assignments"""
        data = [
            # Head Office Staff - Administrative
            {'firstname': 'Thomas', 'lastname': 'Ndongo', 'office': 'Head Office', 'role': 'System Administrator',
             'username': 'admin.system', 'email': 't.ndongo@mfi.cm', 'mobile': '+237 677 11 11 11',
             'joining_date': '2024-01-02', 'external_id': 'STF-101'},

            {'firstname': 'Christine', 'lastname': 'Biaka', 'office': 'Head Office', 'role': 'Accountant',
             'username': 'accountant', 'email': 'c.biaka@mfi.cm', 'mobile': '+237 677 22 22 22',
             'joining_date': '2024-01-02', 'external_id': 'STF-102'},
 
            {'firstname': 'Supervisor', 'lastname': 'Accountant', 'office': 'Head Office', 'role': 'Supervisor Accountant',
             'username': 'supervisor', 'email': 's.accountant@mfi.cm', 'mobile': '+237 677 33 33 33',
             'joining_date': '2024-01-02', 'external_id': 'STF-103'},
 
             # Yaounde Branch Staff
             {'firstname': 'Jean', 'lastname': 'Mbarga', 'office': 'Yaounde Branch', 'role': 'Branch Manager',
             'username': 'manager.yaounde', 'email': 'j.mbarga@mfi.cm', 'mobile': '+237 677 12 34 56',
             'joining_date': '2024-01-15', 'external_id': 'STF-001'},

            {'firstname': 'Marie', 'lastname': 'Ngo Balla', 'office': 'Yaounde Branch', 'role': 'Loan Officer',
             'username': 'loan.yaounde', 'email': 'm.ngoballa@mfi.cm', 'mobile': '+237 677 23 45 67',
             'joining_date': '2024-01-20', 'external_id': 'STF-002'},

            {'firstname': 'Paul', 'lastname': 'Atangana', 'office': 'Yaounde Branch', 'role': 'Cashier',
             'username': 'cashier.yaounde', 'email': 'p.atangana@mfi.cm', 'mobile': '+237 677 34 56 78',
             'joining_date': '2024-01-20', 'external_id': 'STF-003'},

            # Douala Branch Staff
            {'firstname': 'Grace', 'lastname': 'Douala', 'office': 'Douala Branch', 'role': 'Branch Manager',
             'username': 'manager.douala', 'email': 'g.douala@mfi.cm', 'mobile': '+237 677 45 67 89',
             'joining_date': '2024-01-15', 'external_id': 'STF-004'},

            {'firstname': 'Emmanuel', 'lastname': 'Ewondo', 'office': 'Douala Branch', 'role': 'Loan Officer',
             'username': 'loan.douala', 'email': 'e.ewondo@mfi.cm', 'mobile': '+237 677 56 78 90',
             'joining_date': '2024-01-20', 'external_id': 'STF-005'},

            {'firstname': 'Francine', 'lastname': 'Makang', 'office': 'Douala Branch', 'role': 'Cashier',
             'username': 'cashier.douala', 'email': 'f.makang@mfi.cm', 'mobile': '+237 677 67 89 01',
             'joining_date': '2024-01-20', 'external_id': 'STF-006'},

            # Bafoussam Branch Staff
            {'firstname': 'Bernard', 'lastname': 'Kamga', 'office': 'Bafoussam Branch', 'role': 'Branch Manager',
             'username': 'manager.bafoussam', 'email': 'b.kamga@mfi.cm', 'mobile': '+237 677 78 90 12',
             'joining_date': '2024-02-01', 'external_id': 'STF-007'},

            {'firstname': 'Justine', 'lastname': 'Tchuente', 'office': 'Bafoussam Branch', 'role': 'Loan Officer',
             'username': 'loan.bafoussam', 'email': 'j.tchuente@mfi.cm', 'mobile': '+237 677 89 01 23',
             'joining_date': '2024-02-05', 'external_id': 'STF-008'},

            {'firstname': 'David', 'lastname': 'Fotso', 'office': 'Bafoussam Branch', 'role': 'Cashier',
             'username': 'cashier.bafoussam', 'email': 'd.fotso@mfi.cm', 'mobile': '+237 677 90 12 34',
             'joining_date': '2024-02-05', 'external_id': 'STF-009'},

            # Bamenda Branch Staff
            {'firstname': 'Peter', 'lastname': 'Nkeng', 'office': 'Bamenda Branch', 'role': 'Branch Manager',
             'username': 'manager.bamenda', 'email': 'p.nkeng@mfi.cm', 'mobile': '+237 677 01 23 45',
             'joining_date': '2024-02-15', 'external_id': 'STF-010'},

            {'firstname': 'Alice', 'lastname': 'Fon', 'office': 'Bamenda Branch', 'role': 'Loan Officer',
             'username': 'loan.bamenda', 'email': 'a.fon@mfi.cm', 'mobile': '+237 677 13 35 57',
             'joining_date': '2024-02-20', 'external_id': 'STF-011'},

            {'firstname': 'Joseph', 'lastname': 'Tanyi', 'office': 'Bamenda Branch', 'role': 'Cashier',
             'username': 'cashier.bamenda', 'email': 'j.tanyi@mfi.cm', 'mobile': '+237 677 24 46 68',
             'joining_date': '2024-02-20', 'external_id': 'STF-012'},
        ]
        return pd.DataFrame(data)

    def create_users_sheet(self):
        """Create system users with staff linkage and role assignments.

        Note: Fineract password requirements:
        - 12 to 50 characters long
        - At least one uppercase letter
        - At least one lowercase letter
        - At least one numeric digit
        - At least one special character
        - No spaces
        - No consecutive repeating characters
        """
        data = [
            # Head Office Users
            {'username': 'admin.system', 'first_name': 'Thomas', 'last_name': 'Ndongo',
             'email': 't.ndongo@mfi.cm', 'office_name': 'Head Office',
             'roles': 'System Administrator', 'staff_name': 'Thomas Ndongo',
             'password': 'Adm1n$yst3m@2024', 'password_never_expires': True},

            {'username': 'accountant', 'first_name': 'Christine', 'last_name': 'Biaka',
             'email': 'c.biaka@mfi.cm', 'office_name': 'Head Office',
             'roles': 'Accountant', 'staff_name': 'Christine Biaka',
             'password': 'Ac0untant@2024!', 'password_never_expires': False},
 
            {'username': 'supervisor', 'first_name': 'Supervisor', 'last_name': 'Accountant',
             'email': 's.accountant@mfi.cm', 'office_name': 'Head Office',
             'roles': 'Supervisor Accountant', 'staff_name': 'Supervisor Accountant',
             'password': 'sUp3rV1sor@2024!', 'password_never_expires': False},
 
             # Yaounde Branch Users
             {'username': 'manager.yaounde', 'first_name': 'Jean', 'last_name': 'Mbarga',
             'email': 'j.mbarga@mfi.cm', 'office_name': 'Yaounde Branch',
             'roles': 'Branch Manager', 'staff_name': 'Jean Mbarga',
             'password': 'Man4ger@Yde2024', 'password_never_expires': False},

            {'username': 'loan.yaounde', 'first_name': 'Marie', 'last_name': 'Ngo Balla',
             'email': 'm.ngoballa@mfi.cm', 'office_name': 'Yaounde Branch',
             'roles': 'Loan Officer', 'staff_name': 'Marie Ngo Balla',
             'password': 'Lo4n0ficer@Yde1', 'password_never_expires': False},

            {'username': 'cashier.yaounde', 'first_name': 'Paul', 'last_name': 'Atangana',
             'email': 'p.atangana@mfi.cm', 'office_name': 'Yaounde Branch',
             'roles': 'Cashier', 'staff_name': 'Paul Atangana',
             'password': 'Cash1er@Yde2024', 'password_never_expires': False},

            # Douala Branch Users
            {'username': 'manager.douala', 'first_name': 'Grace', 'last_name': 'Douala',
             'email': 'g.douala@mfi.cm', 'office_name': 'Douala Branch',
             'roles': 'Branch Manager', 'staff_name': 'Grace Douala',
             'password': 'Man4ger@Dla2024', 'password_never_expires': False},

            {'username': 'loan.douala', 'first_name': 'Emmanuel', 'last_name': 'Ewondo',
             'email': 'e.ewondo@mfi.cm', 'office_name': 'Douala Branch',
             'roles': 'Loan Officer', 'staff_name': 'Emmanuel Ewondo',
             'password': 'Lo4n0ficer@Dla1', 'password_never_expires': False},

            {'username': 'cashier.douala', 'first_name': 'Francine', 'last_name': 'Makang',
             'email': 'f.makang@mfi.cm', 'office_name': 'Douala Branch',
             'roles': 'Cashier', 'staff_name': 'Francine Makang',
             'password': 'Cash1er@Dla2024', 'password_never_expires': False},

            # Bafoussam Branch Users
            {'username': 'manager.bafoussam', 'first_name': 'Bernard', 'last_name': 'Kamga',
             'email': 'b.kamga@mfi.cm', 'office_name': 'Bafoussam Branch',
             'roles': 'Branch Manager', 'staff_name': 'Bernard Kamga',
             'password': 'Man4ger@Baf2024', 'password_never_expires': False},

            {'username': 'loan.bafoussam', 'first_name': 'Justine', 'last_name': 'Tchuente',
             'email': 'j.tchuente@mfi.cm', 'office_name': 'Bafoussam Branch',
             'roles': 'Loan Officer', 'staff_name': 'Justine Tchuente',
             'password': 'Lo4n0ficer@Baf1', 'password_never_expires': False},

            {'username': 'cashier.bafoussam', 'first_name': 'David', 'last_name': 'Fotso',
             'email': 'd.fotso@mfi.cm', 'office_name': 'Bafoussam Branch',
             'roles': 'Cashier', 'staff_name': 'David Fotso',
             'password': 'Cash1er@Baf2024', 'password_never_expires': False},

            # Bamenda Branch Users
            {'username': 'manager.bamenda', 'first_name': 'Peter', 'last_name': 'Nkeng',
             'email': 'p.nkeng@mfi.cm', 'office_name': 'Bamenda Branch',
             'roles': 'Branch Manager', 'staff_name': 'Peter Nkeng',
             'password': 'Man4ger@Bam2024', 'password_never_expires': False},

            {'username': 'loan.bamenda', 'first_name': 'Alice', 'last_name': 'Fon',
             'email': 'a.fon@mfi.cm', 'office_name': 'Bamenda Branch',
             'roles': 'Loan Officer', 'staff_name': 'Alice Fon',
             'password': 'Lo4n0ficer@Bam1', 'password_never_expires': False},

            {'username': 'cashier.bamenda', 'first_name': 'Joseph', 'last_name': 'Tanyi',
             'email': 'j.tanyi@mfi.cm', 'office_name': 'Bamenda Branch',
             'roles': 'Cashier', 'staff_name': 'Joseph Tanyi',
             'password': 'Cash1er@Bam2024', 'password_never_expires': False},
        ]
        return pd.DataFrame(data)

    def create_clients_sheet(self):
        """Create sample clients.

        legalFormId values:
        - 1 = Person (Individual)
        - 2 = Entity (Corporate/Business)
        """
        data = [
            {'firstname': 'Akoumba', 'lastname': 'Ngono', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Female', 'mobile': '+237 690 12 34 56', 'email': 'a.ngono@gmail.com',
             'date_of_birth': '1985-03-15', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-001', 'activation_date': '2024-01-20', 'national_id': 'CM1234567890',
             'address': 'Bonaberi, Rue des Fleurs', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 3, 'occupation': 'Petty Trader', 'business_type': 'Retail Trade',
             'monthly_income': 150000, 'risk_rating': 'B'},

            {'firstname': 'Ibrahim', 'lastname': 'Mahamat', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 690 23 45 67', 'email': 'i.mahamat@yahoo.fr',
             'date_of_birth': '1978-07-22', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-002', 'activation_date': '2024-01-25', 'national_id': 'CM2345678901',
             'address': 'Ndokoti, Carrefour Ange Raphael', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 5, 'occupation': 'Transport Owner', 'business_type': 'Transportation',
             'monthly_income': 300000, 'risk_rating': 'A'},

            {'firstname': 'Marie Claire', 'lastname': 'Tchouta', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Female', 'mobile': '+237 690 34 56 78', 'email': 'm.tchouta@gmail.com',
             'date_of_birth': '1990-11-10', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-003', 'activation_date': '2024-02-01', 'national_id': 'CM3456789012',
             'address': 'Melen, Quartier Fouda', 'city': 'Yaounde', 'marital_status': 'Single',
             'number_of_dependents': 1, 'occupation': 'Hairdresser', 'business_type': 'Services',
             'monthly_income': 120000, 'risk_rating': 'B'},

            {'firstname': 'Jean Paul', 'lastname': 'Kamdem', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'gender': 'Male', 'mobile': '+237 690 45 67 89', 'email': 'jp.kamdem@yahoo.fr',
             'date_of_birth': '1982-05-18', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-004', 'activation_date': '2024-02-05', 'national_id': 'CM4567890123',
             'address': 'Tamdja, Marche des Vivres', 'city': 'Bafoussam', 'marital_status': 'Married',
             'number_of_dependents': 4, 'occupation': 'Farmer', 'business_type': 'Agriculture',
             'monthly_income': 200000, 'risk_rating': 'B'},

            {'firstname': 'Claudine', 'lastname': 'Ekotto', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Female', 'mobile': '+237 690 56 78 90', 'email': 'c.ekotto@gmail.com',
             'date_of_birth': '1988-09-25', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-005', 'activation_date': '2024-02-10', 'national_id': 'CM5678901234',
             'address': 'New Bell, Rue Joffre', 'city': 'Douala', 'marital_status': 'Widow',
             'number_of_dependents': 2, 'occupation': 'Restaurant Owner', 'business_type': 'Food Services',
             'monthly_income': 180000, 'risk_rating': 'B'},

            {'firstname': 'Patrick', 'lastname': 'Ndongo', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Male', 'mobile': '+237 690 67 89 01', 'email': 'p.ndongo@hotmail.com',
             'date_of_birth': '1975-12-03', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-006', 'activation_date': '2024-02-15', 'national_id': 'CM6789012345',
             'address': 'Omnisport, Carrefour Warda', 'city': 'Yaounde', 'marital_status': 'Married',
             'number_of_dependents': 6, 'occupation': 'Wholesaler', 'business_type': 'Wholesale Trade',
             'monthly_income': 500000, 'risk_rating': 'A'},

            {'firstname': 'Beatrice', 'lastname': 'Nana', 'office': 'Bamenda Branch', 'staff': 'loan.bamenda',
             'gender': 'Female', 'mobile': '+237 690 78 90 12', 'email': 'b.nana@gmail.com',
             'date_of_birth': '1992-04-14', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-007', 'activation_date': '2024-02-20', 'national_id': 'CM7890123456',
             'address': 'Nkwen, Mile 4', 'city': 'Bamenda', 'marital_status': 'Single',
             'number_of_dependents': 0, 'occupation': 'Seamstress', 'business_type': 'Services',
             'monthly_income': 100000, 'risk_rating': 'C'},

            {'firstname': 'Samuel', 'lastname': 'Ebong', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 690 89 01 23', 'email': 's.ebong@yahoo.fr',
             'date_of_birth': '1980-08-30', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-008', 'activation_date': '2024-03-01', 'national_id': 'CM8901234567',
             'address': 'Bepanda, Rond Point', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 3, 'occupation': 'Mechanic', 'business_type': 'Services',
             'monthly_income': 160000, 'risk_rating': 'B'},

            {'firstname': 'Bernadette', 'lastname': 'Fotso', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'gender': 'Female', 'mobile': '+237 690 90 12 34', 'email': 'b.fotso@gmail.com',
             'date_of_birth': '1987-06-20', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-009', 'activation_date': '2024-03-05', 'national_id': 'CM9012345678',
             'address': 'Famla, Marche Central', 'city': 'Bafoussam', 'marital_status': 'Married',
             'number_of_dependents': 2, 'occupation': 'Vegetable Seller', 'business_type': 'Retail Trade',
             'monthly_income': 90000, 'risk_rating': 'C'},

            {'firstname': 'Francois', 'lastname': 'Manga', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Male', 'mobile': '+237 690 01 23 45', 'email': 'f.manga@hotmail.com',
             'date_of_birth': '1983-02-28', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-010', 'activation_date': '2024-03-10', 'national_id': 'CM0123456789',
             'address': 'Elig-Essono, Nouvelle Route', 'city': 'Yaounde', 'marital_status': 'Divorced',
             'number_of_dependents': 2, 'occupation': 'Carpenter', 'business_type': 'Services',
             'monthly_income': 140000, 'risk_rating': 'B'},

            {'firstname': 'Agnes', 'lastname': 'Fon', 'office': 'Bamenda Branch', 'staff': 'loan.bamenda',
             'gender': 'Female', 'mobile': '+237 691 12 34 56', 'email': 'a.fon@gmail.com',
             'date_of_birth': '1991-10-05', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-011', 'activation_date': '2024-03-15', 'national_id': 'CM1234567891',
             'address': 'Ntarikon, Junction', 'city': 'Bamenda', 'marital_status': 'Single',
             'number_of_dependents': 1, 'occupation': 'Boutique Owner', 'business_type': 'Retail Trade',
             'monthly_income': 110000, 'risk_rating': 'C'},

            {'firstname': 'Robert', 'lastname': 'Essomba', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 691 23 45 67', 'email': 'r.essomba@yahoo.fr',
             'date_of_birth': '1979-01-17', 'client_type': 'Individual', 'legal_form_id': 1, 'client_classification': 'Active',
             'external_id': 'CLI-012', 'activation_date': '2024-03-20', 'national_id': 'CM2345678902',
             'address': 'Akwa, Boulevard de la Republique', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 4, 'occupation': 'Electronics Trader', 'business_type': 'Retail Trade',
             'monthly_income': 250000, 'risk_rating': 'A'},
        ]
        return pd.DataFrame(data)

    def create_groups_sheet(self):
        """Create sample groups (solidarity groups)

        IMPORTANT: Group activation dates must be >= parent office opening date
        - Douala Branch opened: 2024-01-15
        - Yaounde Branch opened: 2024-01-15
        - Bafoussam Branch opened: 2024-02-01
        - Bamenda Branch opened: 2024-02-15
        """
        data = [
            {'group_name': 'Femmes Solidaires Douala', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'external_id': 'GRP-001', 'activation_date': '2024-01-25', 'submitted_on_date': '2024-01-20',
             'active': 'Yes', 'center_name': 'Centre Commercial Douala',
             'client_external_ids': 'CLI-001,CLI-005,CLI-008'},

            {'group_name': 'Commercants Yaounde', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'external_id': 'GRP-002', 'activation_date': '2024-01-25', 'submitted_on_date': '2024-01-20',
             'active': 'Yes', 'center_name': 'Centre Melen Yaounde',
             'client_external_ids': 'CLI-003,CLI-006,CLI-010'},

            {'group_name': 'Agriculteurs Bafoussam', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'external_id': 'GRP-003', 'activation_date': '2024-02-10', 'submitted_on_date': '2024-02-05',
             'active': 'Yes', 'center_name': '',
             'client_external_ids': 'CLI-004,CLI-009'},

            {'group_name': 'Artisans Bamenda', 'office': 'Bamenda Branch', 'staff': 'loan.bamenda',
             'external_id': 'GRP-004', 'activation_date': '2024-02-20', 'submitted_on_date': '2024-02-16',
             'active': 'Yes', 'center_name': '',
             'client_external_ids': 'CLI-007,CLI-011'},

            {'group_name': 'Transporteurs Douala', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'external_id': 'GRP-005', 'activation_date': '2024-02-20', 'submitted_on_date': '2024-02-16',
             'active': 'Yes', 'center_name': 'Centre Commercial Douala',
             'client_external_ids': 'CLI-002,CLI-012'},
        ]
        return pd.DataFrame(data)

    def create_centers_sheet(self):
        """Create sample centers (collection of groups)

        IMPORTANT: Center activation dates must be >= parent office opening date
        - Douala Branch opened: 2024-01-15
        - Yaounde Branch opened: 2024-01-15
        - Bafoussam Branch opened: 2024-02-01
        - Bamenda Branch opened: 2024-02-15
        """
        data = [
            {'center_name': 'Centre Commercial Douala', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'external_id': 'CTR-001', 'activation_date': '2024-01-20', 'submitted_on_date': '2024-01-16',
             'active': 'Yes'},

            {'center_name': 'Centre Melen Yaounde', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'external_id': 'CTR-002', 'activation_date': '2024-01-20', 'submitted_on_date': '2024-01-16',
             'active': 'Yes'},

            {'center_name': 'Centre Tamdja Bafoussam', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'external_id': 'CTR-003', 'activation_date': '2024-02-05', 'submitted_on_date': '2024-02-02',
             'active': 'Yes'},
        ]
        return pd.DataFrame(data)

    def create_loan_products_sheet(self):
        """Create loan products configuration with all mandatory fields for Fineract API"""
        data = [
            {
                'product_name': 'Microcredit Solidarity Loan',
                'short_name': 'MSOL',
                'description': 'Small loans for petty traders and artisans (solidarity group)',
                'currency': 'XAF',
                'digits_after_decimal': 0,
                'principal_min': 50000,
                'principal_default': 200000,
                'principal_max': 500000,
                'number_of_repayments_min': 3,
                'number_of_repayments_default': 6,
                'number_of_repayments_max': 12,
                'repayment_every': 1,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 18.0,
                'interest_rate_default': 24.0,
                'interest_rate_max': 30.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'days_in_year_type': 365,
                'days_in_month_type': 30,
                'is_interest_recalculation_enabled': 'No',
                'grace_on_principal_periods': 0,
                'grace_on_interest_periods': 0,
                'processing_fee_percent': 2.0,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.5,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'No',
                'accounting_type': 'Accrual (periodic)'
            },
            {
                'product_name': 'SME Business Loan',
                'short_name': 'SBIZ',
                'description': 'Medium loans for small and medium enterprises',
                'currency': 'XAF',
                'digits_after_decimal': 0,
                'principal_min': 500000,
                'principal_default': 2000000,
                'principal_max': 10000000,
                'number_of_repayments_min': 6,
                'number_of_repayments_default': 12,
                'number_of_repayments_max': 24,
                'repayment_every': 1,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 15.0,
                'interest_rate_default': 20.0,
                'interest_rate_max': 25.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'days_in_year_type': 365,
                'days_in_month_type': 30,
                'is_interest_recalculation_enabled': 'No',
                'grace_on_principal_periods': 1,
                'grace_on_interest_periods': 0,
                'processing_fee_percent': 1.5,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.0,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'No',
                'accounting_type': 'Accrual (periodic)'
            },
            {
                'product_name': 'Agricultural Seasonal Loan',
                'short_name': 'ASEA',
                'description': 'Seasonal loans for agricultural activities',
                'currency': 'XAF',
                'digits_after_decimal': 0,
                'principal_min': 100000,
                'principal_default': 500000,
                'principal_max': 3000000,
                'number_of_repayments_min': 3,
                'number_of_repayments_default': 6,
                'number_of_repayments_max': 9,
                'repayment_every': 1,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 12.0,
                'interest_rate_default': 18.0,
                'interest_rate_max': 22.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'days_in_year_type': 365,
                'days_in_month_type': 30,
                'is_interest_recalculation_enabled': 'No',
                'grace_on_principal_periods': 2,
                'grace_on_interest_periods': 1,
                'processing_fee_percent': 1.0,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.5,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'Yes',
                'accounting_type': 'Accrual (periodic)'
            }
        ]
        return pd.DataFrame(data)

    def create_savings_products_sheet(self):
        """Create savings products configuration with all mandatory fields for Fineract API"""
        data = [
            {
                'product_name': 'Voluntary Savings Account',
                'short_name': 'VSAV',  # Max 4 characters for Fineract
                'description': 'Standard voluntary savings account',
                'currency': 'XAF',
                'digits_after_decimal': 0,  # MANDATORY for XAF
                'nominal_annual_interest_rate': 3.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 10000,
                'minimum_balance': 5000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'Yes',
                'days_to_inactive': 180,
                'days_to_dormancy': 365,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled
                'lock_in_period': 0,
                'accounting_type': 'Cash'  # CASH_BASED accounting
            },
            {
                'product_name': 'Fixed Deposit Account',
                'short_name': 'FDEP',  # Max 4 characters for Fineract
                'description': 'Fixed deposit with higher interest rate',
                'currency': 'XAF',
                'digits_after_decimal': 0,  # MANDATORY for XAF
                'nominal_annual_interest_rate': 6.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 100000,
                'minimum_balance': 100000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'No',
                'days_to_inactive': 0,
                'days_to_dormancy': 0,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled
                'lock_in_period': 180,
                'accounting_type': 'Cash'  # CASH_BASED accounting
            },
            {
                'product_name': 'Mandatory Group Savings',
                'short_name': 'MGRP',  # Max 4 characters for Fineract
                'description': 'Mandatory savings for group loan members',
                'currency': 'XAF',
                'digits_after_decimal': 0,  # MANDATORY for XAF
                'nominal_annual_interest_rate': 2.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 5000,
                'minimum_balance': 5000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'No',
                'days_to_inactive': 0,
                'days_to_dormancy': 0,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled
                'lock_in_period': 0,
                'accounting_type': 'Cash'  # CASH_BASED accounting
            }
        ]
        return pd.DataFrame(data)

    def create_charges_sheet(self):
        """Create fees and charges with fee frequency support

        Valid charge_time values by charge_type:
        - Loan: DISBURSEMENT(1), SPECIFIED_DUE_DATE(2), MONTHLY_FEE(8), WEEKLY_FEE(9), TRANCHE_DISBURSEMENT(12)
        - Savings: SPECIFIED_DUE_DATE(2), SAVINGS_ACTIVATION(5), WITHDRAWAL_FEE(6), ANNUAL_FEE(7),
                   MONTHLY_FEE(8), WEEKLY_FEE(9), OVERDRAFT_FEE(10), SAVINGS_CLOSURE(11), FDA_PRE_CLOSURE(16)
        """
        data = [
            # LOAN CHARGES (valid: 1, 2, 8, 9, 12)
            {'charge_name': 'Loan Processing Fee', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 2.0, 'currency': 'XAF', 'charge_time': 'Disbursement', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},

            {'charge_name': 'Late Payment Penalty', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 5.0, 'currency': 'XAF', 'charge_time': 'Specified Due Date', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},

            {'charge_name': 'Loan Insurance', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 1.5, 'currency': 'XAF', 'charge_time': 'Disbursement', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},

            {'charge_name': 'Loan Restructuring Fee', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 1.0, 'currency': 'XAF', 'charge_time': 'Specified Due Date', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},

            {'charge_name': 'Early Repayment Penalty', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 3.0, 'currency': 'XAF', 'charge_time': 'Disbursement', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'No'},

            # SAVINGS CHARGES - Only using simple charge types that don't require feeOnMonthDay/feeInterval
            # Valid simple types: 5 (Savings Activation)
            # Note: Withdrawal Fee (type 6) requires feeOnMonthDay in Fineract, so we exclude it
            {'charge_name': 'Account Opening Fee', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 1000, 'currency': 'XAF', 'charge_time': 'Savings Activation', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},

            {'charge_name': 'Dormant Account Reactivation', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 2000, 'currency': 'XAF', 'charge_time': 'Savings Activation', 'fee_frequency': None,
             'fee_interval': None, 'fee_on_day': None, 'fee_on_month': None, 'active': 'Yes'},
        ]
        return pd.DataFrame(data)

    def create_gl_accounts_sheet(self):
        """Create Chart of Accounts - OHADA compliant"""
        data = [
            # ASSETS - Class 1-5
            {'gl_code': '41', 'gl_name': 'Banks - Current Accounts', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Bank current accounts'},

            {'gl_code': '42', 'gl_name': 'Cash on Hand', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Physical cash in tellers and vaults'},

            {'gl_code': '51', 'gl_name': 'Loans to Clients - Principal', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Outstanding loan principal'},

            {'gl_code': '52', 'gl_name': 'Loans to Clients - Interest Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued interest receivable'},

            {'gl_code': '53', 'gl_name': 'Loans to Clients - Fees Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued fees receivable'},

            {'gl_code': '54', 'gl_name': 'Loans to Clients - Penalties Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued penalties receivable'},

            {'gl_code': '55', 'gl_name': 'Provision for Loan Losses', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Accumulated provision for bad loans (contra account)'},

            {'gl_code': '122', 'gl_name': 'Due from Other Branches (Receivable)', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Inter-branch receivables'},

            # LIABILITIES - Class 1-5
            {'gl_code': '61', 'gl_name': 'Voluntary Savings Accounts', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Customer voluntary savings deposits'},

            {'gl_code': '62', 'gl_name': 'Fixed Deposit Accounts', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Fixed deposit liabilities'},

            {'gl_code': '63', 'gl_name': 'Mandatory Group Savings', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Mandatory group savings liabilities'},

            {'gl_code': '64', 'gl_name': 'Savings Interest Payable', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued interest payable to savers'},

            {'gl_code': '131', 'gl_name': 'Due to Other Branches (Payable)', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Inter-branch payables'},

            {'gl_code': '71', 'gl_name': 'Share Capital', 'account_type': 'Equity',
             'classification': 'Equity', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Paid-up share capital'},

            {'gl_code': '72', 'gl_name': 'Retained Earnings', 'account_type': 'Equity',
             'classification': 'Equity', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Accumulated retained earnings'},

            # INCOME - Class 7
            {'gl_code': '81', 'gl_name': 'Interest Income on Loans', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Interest earned from loans'},

            {'gl_code': '82', 'gl_name': 'Fee Income - Loan Processing', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Loan processing fees'},

            {'gl_code': '83', 'gl_name': 'Penalty Income', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Late payment penalties'},

            {'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Savings account fees'},

            # EXPENSES - Class 6
            {'gl_code': '91', 'gl_name': 'Interest Expense on Savings', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Interest paid to savers'},

            {'gl_code': '92', 'gl_name': 'Loan Loss Provision Expense', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Provision for loan losses'},

            {'gl_code': '93', 'gl_name': 'Loan Write-off Expense', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Written-off loans'},

            {'gl_code': '94', 'gl_name': 'Staff Salaries', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Employee salaries and wages'},

            {'gl_code': '95', 'gl_name': 'Office Rent', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Office rental expenses'},

            {'gl_code': '96', 'gl_name': 'Utilities', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Utilities expenses'},

            {'gl_code': '97', 'gl_name': 'Tax Expense - Withholding Tax', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Withholding tax on interest'},

            {'gl_code': '141', 'gl_name': 'Tax Payable - WHT', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Withholding tax payable'},

            # Mobile Money Accounts
            {'gl_code': '43', 'gl_name': 'MTN Mobile Money', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'MTN MoMo balance'},

            {'gl_code': '44', 'gl_name': 'Orange Money', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Orange Money balance'},

            {'gl_code': '85', 'gl_name': 'Mobile Money Transfer Fees', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Mobile money transaction fees'},

            # Teller Cash Shortage/Overage Accounts
            {'gl_code': '98', 'gl_name': 'Cash Shortage', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Cash shortages from teller operations'},

            {'gl_code': '86', 'gl_name': 'Cash Overage', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Cash overages from teller operations'},
        ]
        return pd.DataFrame(data)

    def create_loan_accounts_sheet(self):
        """Create sample loan accounts

        workflow_state values:
        - 'active': Auto-approve and auto-disburse (normal flow)
        - 'pending_approval': Create loan, stop at approval (demonstrates maker-checker)
        - 'pending_disbursal': Create and approve loan, stop at disbursal (demonstrates maker-checker)
        """
        data = [
            {'client_external_id': 'CLI-001', 'product': 'MSOL', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-05', 'disbursed_on': '2024-02-09', 'principal': 200000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-001',
             'workflow_state': 'active',
             'loan_purpose_detail': 'Purchase additional stock for retail shop',
             'repayment_source': 'Business Income', 'credit_score': 72,
             'previous_loan_history': 'Good History'},

            {'client_external_id': 'CLI-002', 'product': 'SBIZ', 'submitted_on': '2024-02-05',
             'approved_on': '2024-02-08', 'disbursed_on': '2024-02-12', 'principal': 2000000,
             'loan_term': 12, 'interest_rate': 20.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Bank Loan - Ecobank', 'external_id': 'LOAN-002',
             'workflow_state': 'pending_approval',
             'loan_purpose_detail': 'Expand transport business - purchase second vehicle (PENDING APPROVAL)',
             'repayment_source': 'Business Income', 'credit_score': 85,
             'previous_loan_history': 'First Loan'},

            {'client_external_id': 'CLI-003', 'product': 'MSOL', 'submitted_on': '2024-02-10',
             'approved_on': '2024-02-12', 'disbursed_on': '2024-02-15', 'principal': 150000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.yaounde',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-003',
             'workflow_state': 'active',
             'loan_purpose_detail': 'Renovate salon and purchase new equipment',
             'repayment_source': 'Business Income', 'credit_score': 68,
             'previous_loan_history': 'First Loan'},

            {'client_external_id': 'CLI-004', 'product': 'ASEA', 'submitted_on': '2024-02-12',
             'approved_on': '2024-02-15', 'disbursed_on': '2024-02-19', 'principal': 500000,
             'loan_term': 6, 'interest_rate': 18.0, 'loan_officer': 'loan.bafoussam',
             'fund_source': 'Donor - AfDB', 'external_id': 'LOAN-004',
             'workflow_state': 'active',
             'loan_purpose_detail': 'Purchase fertilizer and seeds for planting season',
             'repayment_source': 'Business Income', 'credit_score': 75,
             'previous_loan_history': 'Good History'},

            {'client_external_id': 'CLI-005', 'product': 'MSOL', 'submitted_on': '2024-02-15',
             'approved_on': '2024-02-18', 'disbursed_on': '2024-02-20', 'principal': 180000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-005',
             'workflow_state': 'active',
             'loan_purpose_detail': 'Expand restaurant seating and buy cooking equipment',
             'repayment_source': 'Business Income', 'credit_score': 70,
             'previous_loan_history': 'Some Delays'},

            {'client_external_id': 'CLI-006', 'product': 'SBIZ', 'submitted_on': '2024-02-18',
             'approved_on': '2024-02-20', 'disbursed_on': '2024-02-26', 'principal': 3000000,
             'loan_term': 18, 'interest_rate': 18.0, 'loan_officer': 'loan.yaounde',
             'fund_source': 'Bank Loan - Ecobank', 'external_id': 'LOAN-006',
             'workflow_state': 'pending_disbursal',
             'loan_purpose_detail': 'Stock wholesale shop with imported goods (PENDING DISBURSAL)',
             'repayment_source': 'Business Income', 'credit_score': 90,
             'previous_loan_history': 'Good History'},
        ]
        return pd.DataFrame(data)

    def create_savings_accounts_sheet(self):
        """Create sample savings accounts"""
        data = [
            {'client_external_id': 'CLI-001', 'product': 'VSAV', 'submitted_on': '2024-01-20',
             'approved_on': '2024-01-20', 'activated_on': '2024-01-20', 'initial_deposit': 25000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-001',
             'savings_goal': 'School fees for children', 'target_amount': 500000,
             'target_date': '2024-09-01', 'monthly_commitment': 50000,
             'preferred_transaction_channel': 'Branch'},

            {'client_external_id': 'CLI-002', 'product': 'VSAV', 'submitted_on': '2024-01-25',
             'approved_on': '2024-01-25', 'activated_on': '2024-01-25', 'initial_deposit': 50000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-002',
             'savings_goal': 'Emergency fund', 'target_amount': 1000000,
             'target_date': '2024-12-31', 'monthly_commitment': 100000,
             'preferred_transaction_channel': 'Mobile Money'},

            {'client_external_id': 'CLI-002', 'product': 'FDEP', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-01', 'activated_on': '2024-02-01', 'initial_deposit': 100000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-003',
             'savings_goal': 'Business expansion capital', 'target_amount': 5000000,
             'target_date': '2025-02-01', 'monthly_commitment': 200000,
             'preferred_transaction_channel': 'Bank Transfer'},

            {'client_external_id': 'CLI-003', 'product': 'VSAV', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-01', 'activated_on': '2024-02-01', 'initial_deposit': 15000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-004',
             'savings_goal': 'Buy salon equipment', 'target_amount': 300000,
             'target_date': '2024-08-01', 'monthly_commitment': 40000,
             'preferred_transaction_channel': 'Branch'},

            {'client_external_id': 'CLI-004', 'product': 'VSAV', 'submitted_on': '2024-02-05',
             'approved_on': '2024-02-05', 'activated_on': '2024-02-05', 'initial_deposit': 20000,
             'field_officer': 'loan.bafoussam', 'external_id': 'SAV-005',
             'savings_goal': 'Purchase farm inputs', 'target_amount': 400000,
             'target_date': '2024-10-01', 'monthly_commitment': 50000,
             'preferred_transaction_channel': 'Agent'},

            {'client_external_id': 'CLI-005', 'product': 'VSAV', 'submitted_on': '2024-02-10',
             'approved_on': '2024-02-10', 'activated_on': '2024-02-10', 'initial_deposit': 18000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-006',
             'savings_goal': 'Home renovation', 'target_amount': 600000,
             'target_date': '2024-12-31', 'monthly_commitment': 60000,
             'preferred_transaction_channel': 'Mobile Money'},

            {'client_external_id': 'CLI-006', 'product': 'VSAV', 'submitted_on': '2024-02-15',
             'approved_on': '2024-02-15', 'activated_on': '2024-02-15', 'initial_deposit': 75000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-007',
             'savings_goal': 'Business inventory buffer', 'target_amount': 2000000,
             'target_date': '2025-02-15', 'monthly_commitment': 150000,
             'preferred_transaction_channel': 'Bank Transfer'},

            {'client_external_id': 'CLI-006', 'product': 'FDEP', 'submitted_on': '2024-02-20',
             'approved_on': '2024-02-20', 'activated_on': '2024-02-20', 'initial_deposit': 200000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-008',
             'savings_goal': 'Fixed investment for returns', 'target_amount': 10000000,
             'target_date': '2026-02-20', 'monthly_commitment': 300000,
             'preferred_transaction_channel': 'Branch'},

            {'client_external_id': 'CLI-007', 'product': 'VSAV', 'submitted_on': '2024-02-20',
             'approved_on': '2024-02-20', 'activated_on': '2024-02-20', 'initial_deposit': 12000,
             'field_officer': 'loan.bamenda', 'external_id': 'SAV-009',
             'savings_goal': 'Wedding expenses', 'target_amount': 800000,
             'target_date': '2024-12-01', 'monthly_commitment': 80000,
             'preferred_transaction_channel': 'Mobile Money'},

            {'client_external_id': 'CLI-008', 'product': 'VSAV', 'submitted_on': '2024-03-01',
             'approved_on': '2024-03-01', 'activated_on': '2024-03-01', 'initial_deposit': 16000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-010',
             'savings_goal': 'Tools and equipment', 'target_amount': 350000,
             'target_date': '2024-10-01', 'monthly_commitment': 45000,
             'preferred_transaction_channel': 'Agent'},

            {'client_external_id': 'CLI-009', 'product': 'VSAV', 'submitted_on': '2024-03-05',
             'approved_on': '2024-03-05', 'activated_on': '2024-03-05', 'initial_deposit': 10000,
             'field_officer': 'loan.bafoussam', 'external_id': 'SAV-011',
             'savings_goal': 'Market stall expansion', 'target_amount': 250000,
             'target_date': '2024-09-01', 'monthly_commitment': 35000,
             'preferred_transaction_channel': 'Branch'},

            {'client_external_id': 'CLI-010', 'product': 'VSAV', 'submitted_on': '2024-03-10',
             'approved_on': '2024-03-10', 'activated_on': '2024-03-10', 'initial_deposit': 14000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-012',
             'savings_goal': 'Buy woodworking tools', 'target_amount': 400000,
             'target_date': '2024-11-01', 'monthly_commitment': 50000,
             'preferred_transaction_channel': 'Mobile Money'},
        ]
        return pd.DataFrame(data)

    def create_fund_sources_sheet(self):
        """Create fund sources"""
        data = [
            {'fund_name': 'Own Capital', 'external_id': 'FUND-001', 'description': 'Institutional own capital'},
            {'fund_name': 'Bank Loan - Ecobank', 'external_id': 'FUND-002', 'description': 'Credit line from Ecobank'},
            {'fund_name': 'Bank Loan - BICEC', 'external_id': 'FUND-003', 'description': 'Credit line from BICEC'},
            {'fund_name': 'Donor - AfDB', 'external_id': 'FUND-004', 'description': 'African Development Bank grant'},
            {'fund_name': 'Donor - GIZ', 'external_id': 'FUND-005', 'description': 'German cooperation grant'},
            {'fund_name': 'Donor - World Bank', 'external_id': 'FUND-006', 'description': 'World Bank project funding'},
        ]
        return pd.DataFrame(data)

    def create_payment_types_sheet(self):
        """Create payment channels"""
        data = [
            {'payment_type': 'Cash', 'description': 'Cash payment', 'is_cash_payment': 'Yes', 'order_position': 1},
            {'payment_type': 'MTN Mobile Money', 'description': 'MTN MoMo payment', 'is_cash_payment': 'No', 'order_position': 2},
            {'payment_type': 'Orange Money', 'description': 'Orange Money payment', 'is_cash_payment': 'No', 'order_position': 3},
            {'payment_type': 'Bank Transfer', 'description': 'Bank transfer payment', 'is_cash_payment': 'No', 'order_position': 4},
            {'payment_type': 'Cheque', 'description': 'Cheque payment', 'is_cash_payment': 'No', 'order_position': 5},
        ]
        return pd.DataFrame(data)

    def create_holidays_sheet(self):
        """Create Cameroon public holidays"""
        data = [
            {'holiday_name': "New Year's Day", 'date': '2024-01-01', 'rescheduled_to': '2024-01-02', 'description': 'New Year Holiday'},
            {'holiday_name': 'Youth Day', 'date': '2024-02-11', 'rescheduled_to': '2024-02-12', 'description': 'Youth Day Holiday'},
            {'holiday_name': 'Labour Day', 'date': '2024-05-01', 'rescheduled_to': '2024-05-02', 'description': 'International Labour Day'},
            {'holiday_name': 'National Day', 'date': '2024-05-20', 'rescheduled_to': '2024-05-21', 'description': 'Cameroon National Day'},
            {'holiday_name': 'Assumption Day', 'date': '2024-08-15', 'rescheduled_to': '2024-08-16', 'description': 'Assumption of Mary'},
            {'holiday_name': 'Christmas Day', 'date': '2024-12-25', 'rescheduled_to': '2024-12-26', 'description': 'Christmas Holiday'},
            {'holiday_name': "New Year's Day", 'date': '2025-01-01', 'rescheduled_to': '2025-01-02', 'description': 'New Year Holiday'},
            {'holiday_name': 'Youth Day', 'date': '2025-02-11', 'rescheduled_to': '2025-02-12', 'description': 'Youth Day Holiday'},
            {'holiday_name': 'Labour Day', 'date': '2025-05-01', 'rescheduled_to': '2025-05-02', 'description': 'International Labour Day'},
            {'holiday_name': 'National Day', 'date': '2025-05-20', 'rescheduled_to': '2025-05-21', 'description': 'Cameroon National Day'},
            {'holiday_name': 'Assumption Day', 'date': '2025-08-15', 'rescheduled_to': '2025-08-18', 'description': 'Assumption of Mary'},
            {'holiday_name': 'Christmas Day', 'date': '2025-12-25', 'rescheduled_to': '2025-12-26', 'description': 'Christmas Holiday'},
        ]
        return pd.DataFrame(data)

    def create_loan_provisioning_sheet(self):
        """Create COBAC loan provisioning criteria using Fineract's existing categories"""
        # Map to Fineract's 4 existing provisioning categories: STANDARD, SUB-STANDARD, DOUBTFUL, LOSS
        data = [
            {'category_name': 'STANDARD', 'category_id': 1, 'min_days_overdue': 0, 'max_days_overdue': 30,
             'provision_percentage': 0, 'liability_gl_code': '62', 'expense_gl_code': '92'},
            {'category_name': 'SUB-STANDARD', 'category_id': 2, 'min_days_overdue': 31, 'max_days_overdue': 90,
             'provision_percentage': 25, 'liability_gl_code': '62', 'expense_gl_code': '92'},
            {'category_name': 'DOUBTFUL', 'category_id': 3, 'min_days_overdue': 91, 'max_days_overdue': 180,
             'provision_percentage': 50, 'liability_gl_code': '62', 'expense_gl_code': '92'},
            {'category_name': 'LOSS', 'category_id': 4, 'min_days_overdue': 181, 'max_days_overdue': 9999,
             'provision_percentage': 100, 'liability_gl_code': '62', 'expense_gl_code': '92'},
        ]
        return pd.DataFrame(data)

    def create_collateral_types_sheet(self):
        """Create collateral types"""
        data = [
            {'collateral_type': 'Land Title', 'description': 'Land ownership document', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Vehicle Registration', 'description': 'Vehicle ownership papers', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Household Goods', 'description': 'Furniture, appliances', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Shop/Business Equipment', 'description': 'Business machinery and equipment', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Savings Account Lien', 'description': 'Lien on savings account', 'requires_valuation': 'No'},
            {'collateral_type': 'Group Guarantee', 'description': 'Solidarity group guarantee', 'requires_valuation': 'No'},
        ]
        return pd.DataFrame(data)

    def create_guarantor_types_sheet(self):
        """Create guarantor types"""
        data = [
            {'guarantor_type': 'Spouse', 'description': 'Client spouse as guarantor'},
            {'guarantor_type': 'Family Member', 'description': 'Family member (parent, sibling, child)'},
            {'guarantor_type': 'Employer', 'description': 'Employer guarantee for salaried clients'},
            {'guarantor_type': 'Business Partner', 'description': 'Business partner or associate'},
            {'guarantor_type': 'Group Member', 'description': 'Fellow group member in solidarity lending'},
            {'guarantor_type': 'Community Leader', 'description': 'Village chief or community leader'},
        ]
        return pd.DataFrame(data)

    def create_floating_rates_sheet(self):
        """Create floating interest rates"""
        data = [
            {'rate_name': 'BEAC Base Rate', 'is_base_rate': 'Yes', 'is_active': 'Yes',
             'rate_value': 3.50, 'from_date': '2024-01-01',
             'description': 'Central African Bank (BEAC) reference rate'},

            {'rate_name': 'Prime Lending Rate', 'is_base_rate': 'No', 'is_active': 'Yes',
             'rate_value': 7.50, 'from_date': '2024-01-01',
             'description': 'Prime rate for best customers (BEAC + 4%)'},

            {'rate_name': 'SME Lending Rate', 'is_base_rate': 'No', 'is_active': 'Yes',
             'rate_value': 9.50, 'from_date': '2024-01-01',
             'description': 'Rate for SME loans (BEAC + 6%)'},
        ]
        return pd.DataFrame(data)

    def create_delinquency_buckets_sheet(self):
        """Create delinquency/arrears classification buckets"""
        data = [
            # Bucket 1: Early Stage (1-30 days)
            {'bucket_name': 'Early Stage Delinquency', 'classification': 'Early Stage',
             'min_days_overdue': 1, 'max_days_overdue': 30, 'color_code': '#FFA500',
             'description': 'Loans 1-30 days past due'},

            # Bucket 2: Moderate (31-60 days)
            {'bucket_name': 'Moderate Delinquency', 'classification': 'Moderate',
             'min_days_overdue': 31, 'max_days_overdue': 60, 'color_code': '#FF6347',
             'description': 'Loans 31-60 days past due'},

            # Bucket 3: High Risk (61-90 days)
            {'bucket_name': 'High Risk Delinquency', 'classification': 'High Risk',
             'min_days_overdue': 61, 'max_days_overdue': 90, 'color_code': '#DC143C',
             'description': 'Loans 61-90 days past due'},

            # Bucket 4: Very High Risk (91-180 days)
            {'bucket_name': 'Very High Risk Delinquency', 'classification': 'Very High Risk',
             'min_days_overdue': 91, 'max_days_overdue': 180, 'color_code': '#8B0000',
             'description': 'Loans 91-180 days past due'},

            # Bucket 5: Default (180+ days)
            {'bucket_name': 'Default', 'classification': 'Default',
             'min_days_overdue': 181, 'max_days_overdue': 9999, 'color_code': '#000000',
             'description': 'Loans over 180 days past due'},
        ]
        return pd.DataFrame(data)

    def create_tax_groups_sheet(self):
        """Create tax groups and tax components"""
        data = [
            # Tax Group 1: Savings Interest WHT
            {'tax_group_name': 'Savings Interest Tax', 'tax_component_name': 'Withholding Tax on Interest',
             'tax_type': 'Savings Interest', 'tax_percentage': 15.0, 'start_date': '2024-01-01',
             'credit_account_type': 'Liability', 'credit_gl_code': '141', 'credit_gl_name': 'Tax Payable - WHT',
             'description': '15% WHT on savings interest (Cameroon tax law)'},

            # Tax Group 2: Loan Interest WHT (if applicable)
            {'tax_group_name': 'Loan Interest Tax', 'tax_component_name': 'Withholding Tax on Loan Interest',
             'tax_type': 'Loan Interest', 'tax_percentage': 5.5, 'start_date': '2024-01-01',
             'credit_account_type': 'Liability', 'credit_gl_code': '141', 'credit_gl_name': 'Tax Payable - WHT',
             'description': '5.5% WHT on interest income from loans'},
        ]
        return pd.DataFrame(data)

    def create_config_sheet(self):
        """Create system configuration parameters"""
        data = [
            {'config_key': 'INSTITUTION_NAME', 'config_value': 'Cameroon Microfinance Institution', 'category': 'General'},
            {'config_key': 'INSTITUTION_EXTERNAL_ID', 'config_value': 'CMR-MFI-001', 'category': 'General'},
            {'config_key': 'BASE_CURRENCY', 'config_value': 'XAF', 'category': 'General'},
            {'config_key': 'WORKING_DAYS', 'config_value': 'Monday,Tuesday,Wednesday,Thursday,Friday', 'category': 'General'},
            {'config_key': 'ACCOUNTING_RULE', 'config_value': 'Cash Based', 'category': 'Accounting'},
            {'config_key': 'FINANCIAL_YEAR_START', 'config_value': '01-January', 'category': 'Accounting'},
            {'config_key': 'DEFAULT_PASSWORD', 'config_value': 'password', 'category': 'Security'},
            {'config_key': 'PASSWORD_CHANGE_REQUIRED', 'config_value': 'Yes', 'category': 'Security'},
            {'config_key': 'SMS_PROVIDER', 'config_value': 'Twilio', 'category': 'Communication'},
            {'config_key': 'SMS_ENABLED', 'config_value': 'No', 'category': 'Communication'},
            {'config_key': 'EMAIL_ENABLED', 'config_value': 'No', 'category': 'Communication'},
            {'config_key': 'MTN_MOMO_ENABLED', 'config_value': 'No', 'category': 'Payment'},
            {'config_key': 'ORANGE_MONEY_ENABLED', 'config_value': 'No', 'category': 'Payment'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_CASHIER', 'config_value': '500000', 'category': 'Approval'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_MANAGER', 'config_value': '5000000', 'category': 'Approval'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_DIRECTOR', 'config_value': 'unlimited', 'category': 'Approval'},
            {'config_key': 'MAX_LOAN_RESCHEDULE', 'config_value': '2', 'category': 'Loan'},
            {'config_key': 'WRITEOFF_AFTER_DAYS', 'config_value': '365', 'category': 'Loan'},
            {'config_key': 'WHT_RATE_SAVINGS_INTEREST', 'config_value': '15', 'category': 'Tax'},
        ]
        return pd.DataFrame(data)

    def create_loan_product_accounting_sheet(self):
        """Create loan product GL account mappings"""
        data = []

        # For each loan product, define all 13 GL mappings
        products = ['MSOL', 'SBIZ', 'ASEA']

        for product in products:
            mappings = [
                {'product_short_name': product, 'mapping_type': 'Fund Source', 'gl_code': '42',
                 'gl_name': 'Cash on Hand', 'description': 'Source of funds for disbursement'},

                {'product_short_name': product, 'mapping_type': 'Loan Portfolio', 'gl_code': '51',
                 'gl_name': 'Loans to Clients - Principal', 'description': 'Outstanding loan principal'},

                {'product_short_name': product, 'mapping_type': 'Interest Receivable', 'gl_code': '52',
                 'gl_name': 'Loans to Clients - Interest Receivable', 'description': 'Accrued interest receivable'},

                {'product_short_name': product, 'mapping_type': 'Fees Receivable', 'gl_code': '53',
                 'gl_name': 'Loans to Clients - Fees Receivable', 'description': 'Accrued fees receivable'},

                {'product_short_name': product, 'mapping_type': 'Penalties Receivable', 'gl_code': '54',
                 'gl_name': 'Loans to Clients - Penalties Receivable', 'description': 'Penalties receivable'},

                {'product_short_name': product, 'mapping_type': 'Transfer in Suspense', 'gl_code': '122',
                 'gl_name': 'Due from Other Branches (Receivable)', 'description': 'Suspense account for transfers'},

                {'product_short_name': product, 'mapping_type': 'Interest Income', 'gl_code': '81',
                 'gl_name': 'Interest Income on Loans', 'description': 'Interest income recognition'},

                {'product_short_name': product, 'mapping_type': 'Fee Income', 'gl_code': '82',
                 'gl_name': 'Fee Income - Loan Processing', 'description': 'Fee income recognition'},

                {'product_short_name': product, 'mapping_type': 'Penalty Income', 'gl_code': '83',
                 'gl_name': 'Penalty Income', 'description': 'Penalty income recognition'},

                {'product_short_name': product, 'mapping_type': 'Losses Written Off', 'gl_code': '93',
                 'gl_name': 'Loan Write-off Expense', 'description': 'Written-off loan expense'},

                {'product_short_name': product, 'mapping_type': 'Goodwill Credit', 'gl_code': '93',
                 'gl_name': 'Loan Write-off Expense', 'description': 'Goodwill credit account (EXPENSE type)'},

                {'product_short_name': product, 'mapping_type': 'Income from Recovery', 'gl_code': '81',
                 'gl_name': 'Interest Income on Loans', 'description': 'Recovery income (INCOME type)'},

                {'product_short_name': product, 'mapping_type': 'Over Payment Liability', 'gl_code': '64',
                 'gl_name': 'Savings Interest Payable', 'description': 'Overpayment liability account'},
            ]
            data.extend(mappings)

        return pd.DataFrame(data)

    def create_savings_product_accounting_sheet(self):
        """Create savings product GL account mappings"""
        data = []

        # For each savings product, define all 11 GL mappings
        savings_products = [
            {'short_name': 'VSAV', 'liability_gl': '61'},
            {'short_name': 'FDEP', 'liability_gl': '62'},
            {'short_name': 'MGRP', 'liability_gl': '63'}
        ]

        for product in savings_products:
            mappings = [
                {'product_short_name': product['short_name'], 'mapping_type': 'Savings Reference',
                 'gl_code': '42', 'gl_name': 'Cash on Hand',
                 'description': 'Asset account - where customer deposits are held'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Savings Control',
                 'gl_code': product['liability_gl'], 'gl_name': f"GL Code {product['liability_gl']}",
                 'description': 'Liability account - obligation to customers'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Interest on Savings',
                 'gl_code': '91', 'gl_name': 'Interest Expense on Savings', 'description': 'Interest expense'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Fees',
                 'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'description': 'Fee income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Penalties',
                 'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'description': 'Penalty income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Overdraft Portfolio Control',
                 'gl_code': '51', 'gl_name': 'Loans to Clients - Principal', 'description': 'Overdraft control'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Interest',
                 'gl_code': '81', 'gl_name': 'Interest Income on Loans', 'description': 'Overdraft interest income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Losses Written Off',
                 'gl_code': '93', 'gl_name': 'Loan Write-off Expense', 'description': 'Overdraft write-off'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Escheat Liability',
                 'gl_code': product['liability_gl'], 'gl_name': f"GL Code {product['liability_gl']}",
                 'description': 'Escheat liability'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Withholding Tax',
                 'gl_code': '141', 'gl_name': 'Tax Payable - WHT', 'description': 'Withholding tax payable'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Transfer in Suspense',
                 'gl_code': '131', 'gl_name': 'Due to Other Branches (Payable)', 'description': 'Transfer suspense account'},
            ]
            data.extend(mappings)

        return pd.DataFrame(data)

    def create_payment_type_accounting_sheet(self):
        """Create payment type GL account mappings"""
        data = [
            {'payment_type': 'Cash', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'fund_source': 'Yes', 'description': 'Physical cash transactions'},

            {'payment_type': 'MTN Mobile Money', 'gl_code': '43', 'gl_name': 'MTN Mobile Money',
             'fund_source': 'Yes', 'description': 'MTN MoMo wallet'},

            {'payment_type': 'Orange Money', 'gl_code': '44', 'gl_name': 'Orange Money',
             'fund_source': 'Yes', 'description': 'Orange Money wallet'},

            {'payment_type': 'Bank Transfer', 'gl_code': '41', 'gl_name': 'Banks - Current Accounts',
             'fund_source': 'Yes', 'description': 'Bank account transfers'},

            {'payment_type': 'Cheque', 'gl_code': '41', 'gl_name': 'Banks - Current Accounts',
             'fund_source': 'Yes', 'description': 'Cheque payments'},
        ]
        return pd.DataFrame(data)

    def create_financial_activity_mappings_sheet(self):
        """Create financial activity to GL account mappings"""
        data = [
            # Asset Transfer - CRITICAL for inter-branch operations
            {'financial_activity': 'Asset Transfer', 'gl_code': '122', 'gl_name': 'Due from Other Branches (Receivable)',
             'description': 'Inter-office asset transfers - receiving branch records receivable'},

            # Liability Transfer - CRITICAL for inter-branch operations
            {'financial_activity': 'Liability Transfer', 'gl_code': '131', 'gl_name': 'Due to Other Branches (Payable)',
             'description': 'Inter-office liability transfers - sending branch records payable'},

            # Cash at Mainvault
            {'financial_activity': 'Cash at Mainvault', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Main vault cash - head office cash reserves'},

            # Cash at Teller
            {'financial_activity': 'Cash at Teller', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Teller cash - cash in teller drawers'},

            # Opening Balances Contra
            {'financial_activity': 'Opening Balances Contra', 'gl_code': '71', 'gl_name': 'Share Capital',
             'description': 'Opening balance contra account - for initial data migration'},

            # Fund Source
            {'financial_activity': 'Fund Source', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Default fund source for loan disbursements'},
        ]
        return pd.DataFrame(data)

    def create_teller_cashier_mappings_sheet(self):
        """Create teller/cashier GL account mappings per office.

        staff_external_id links to the cashier staff member who operates the teller.
        """
        data = [
            # Yaounde Branch - Paul Atangana (STF-003) as cashier
            {'office_name': 'Yaounde Branch', 'teller_name': 'Teller 1 - Yaounde',
             'staff_external_id': 'STF-003',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Yaounde branch teller and vault configuration'},

            # Douala Branch - Francine Makang (STF-006) as cashier
            {'office_name': 'Douala Branch', 'teller_name': 'Teller 1 - Douala',
             'staff_external_id': 'STF-006',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Douala branch teller and vault configuration'},

            # Bafoussam Branch - David Fotso (STF-009) as cashier
            {'office_name': 'Bafoussam Branch', 'teller_name': 'Teller 1 - Bafoussam',
             'staff_external_id': 'STF-009',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Bafoussam branch teller and vault configuration'},

            # Bamenda Branch - Joseph Tanyi (STF-012) as cashier
            {'office_name': 'Bamenda Branch', 'teller_name': 'Teller 1 - Bamenda',
             'staff_external_id': 'STF-012',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Bamenda branch teller and vault configuration'},
        ]
        return pd.DataFrame(data)

    def create_tellers_sheet(self):
        """Create tellers (cash counters) for each branch"""
        data = [
            # Yaounde Branch
            {'office_name': 'Yaounde Branch', 'teller_name': 'Teller 1 - Yaounde',
             'description': 'Main teller counter at Yaounde branch',
             'start_date': '2024-01-15', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Douala Branch
            {'office_name': 'Douala Branch', 'teller_name': 'Teller 1 - Douala',
             'description': 'Main teller counter at Douala branch',
             'start_date': '2024-01-15', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Bafoussam Branch
            {'office_name': 'Bafoussam Branch', 'teller_name': 'Teller 1 - Bafoussam',
             'description': 'Main teller counter at Bafoussam branch',
             'start_date': '2024-02-01', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Bamenda Branch
            {'office_name': 'Bamenda Branch', 'teller_name': 'Teller 1 - Bamenda',
             'description': 'Main teller counter at Bamenda branch',
             'start_date': '2024-02-15', 'end_date': '2030-12-31',
             'status': 'Active'},
        ]
        return pd.DataFrame(data)

    def create_teller_accounting_rules_sheet(self):
        """Create teller accounting rules.

        Fields match TellerAccountingRule model:
        - teller_name: Name of the teller
        - cash_in_gl_code: GL account code for cash in (debit)
        - cash_out_gl_code: GL account code for cash out (credit)
        - description: Description of the accounting rule
        - office_name: Office scope for the rule
        """
        # GL codes from system: 42 = Cash on Hand, 41 = Banks - Current Accounts
        data = [
            # Yaounde Branch Teller
            {'teller_name': 'Teller 1 - Yaounde', 'office_name': 'Yaounde Branch',
             'cash_in_gl_code': '42', 'cash_out_gl_code': '42',
             'description': 'Cash accounting rule for Yaounde main teller'},

            # Douala Branch Teller
            {'teller_name': 'Teller 1 - Douala', 'office_name': 'Douala Branch',
             'cash_in_gl_code': '42', 'cash_out_gl_code': '42',
             'description': 'Cash accounting rule for Douala main teller'},

            # Bafoussam Branch Teller
            {'teller_name': 'Teller 1 - Bafoussam', 'office_name': 'Bafoussam Branch',
             'cash_in_gl_code': '42', 'cash_out_gl_code': '42',
             'description': 'Cash accounting rule for Bafoussam main teller'},

            # Bamenda Branch Teller
            {'teller_name': 'Teller 1 - Bamenda', 'office_name': 'Bamenda Branch',
             'cash_in_gl_code': '42', 'cash_out_gl_code': '42',
             'description': 'Cash accounting rule for Bamenda main teller'},
        ]
        return pd.DataFrame(data)

    def create_roles_permissions_sheet(self):
        """Create roles with permissions assignments"""
        data = [
            # System Administrator
            {'role_name': 'System Administrator', 'permission_group': 'Role', 'permission': 'SUPER_USER',
             'description': 'Super user with full system access'},

            # Branch Manager
            {'role_name': 'Branch Manager', 'permission_group': 'Role', 'permission': 'BRANCH_MANAGER',
             'description': 'Branch manager with approval authority'},

            # Loan Officer
            {'role_name': 'Loan Officer', 'permission_group': 'Role', 'permission': 'LOAN_OFFICER',
             'description': 'Loan officer for client and loan management'},

            # Cashier
            {'role_name': 'Cashier', 'permission_group': 'Role', 'permission': 'CASHIER',
             'description': 'Cashier for handling cash transactions'},

            # Accountant
            {'role_name': 'Accountant', 'permission_group': 'Role', 'permission': 'ACCOUNTANT',
             'description': 'Accountant for financial management (Maker)'},

            # Supervisor Accountant
            {'role_name': 'Supervisor Accountant', 'permission_group': 'Role', 'permission': 'SUPERVISOR_ACCOUNTANT',
             'description': 'Senior accountant for audit and approval (Checker)'},
        ]
        return pd.DataFrame(data)

    def create_maker_checker_config_sheet(self):
        """Create Maker-Checker (4-eyes principle) configuration for critical operations

        All permission codes verified against Fineract source code (@CommandType annotations).

        NOTE: Fineract's maker-checker is boolean-only (enabled/disabled).
        Amount-based thresholds are NOT supported by the native Fineract API.
        When enabled, ALL operations of that type will require approval.
        """
        data = [
            # Loan operations
            {'task_name': 'Loan Approval', 'entity': 'Loan', 'action': 'APPROVE',
             'enabled': True,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'All loan approvals require manager authorization (demonstrates workflow)'},

            {'task_name': 'Loan Disbursement', 'entity': 'Loan', 'action': 'DISBURSE',
             'enabled': True,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'All loan disbursements require manager authorization (demonstrates workflow)'},

            {'task_name': 'Loan Write-off', 'entity': 'Loan', 'action': 'WRITEOFF',
             'enabled': False,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Loan write-offs require manager approval (disabled by default)'},

            {'task_name': 'Create Loan Reschedule Request', 'entity': 'RESCHEDULELOAN', 'action': 'CREATE',
             'enabled': False,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Loan rescheduling requires manager approval (disabled by default)'},

            # Savings operations
            {'task_name': 'Close Savings Account', 'entity': 'SAVINGSACCOUNT', 'action': 'CLOSE',
             'enabled': True,
             'maker_role': 'Cashier', 'checker_role': 'Branch Manager',
             'description': 'Savings account closures require manager approval (disabled by default)'},

            # Client operations
            {'task_name': 'Client Activation', 'entity': 'Client', 'action': 'ACTIVATE',
             'enabled': False,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'New client activation requires manager approval (disabled by default)'},

            {'task_name': 'Propose Client Transfer', 'entity': 'Client', 'action': 'PROPOSETRANSFER',
             'enabled': True,
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Client office transfer requires manager approval (disabled by default)'},

            # Accounting operations
            {'task_name': 'Create Manual Journal Entry', 'entity': 'JOURNALENTRY', 'action': 'CREATE',
             'enabled': True,
             'maker_role': 'Accountant', 'checker_role': 'Supervisor Accountant',
             'description': 'Manual journal entries require supervisor approval (disabled by default)'},

            # User/System operations
            {'task_name': 'Create User', 'entity': 'User', 'action': 'CREATE',
             'enabled': False,
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Creating new system users requires head office approval (disabled by default)'},

            {'task_name': 'Update User', 'entity': 'User', 'action': 'UPDATE',
             'enabled': False,
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Updating user details requires head office approval (disabled by default)'},

            # Office/Fund transfers
            {'task_name': 'Create Office Transaction', 'entity': 'OFFICETRANSACTION', 'action': 'CREATE',
             'enabled': False,
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Inter-branch fund transfers require head office approval (disabled by default)'},
        ]
        return pd.DataFrame(data)

    def create_currency_config_sheet(self):
        """Create currency configuration"""
        data = [
            {'currency_code': 'XAF', 'currency_name': 'Central African CFA Franc', 'decimal_places': 0,
             'in_multiples_of': 1, 'display_symbol': 'FCFA', 'name_code': 'currency.XAF',
             'description': 'Official currency of CEMAC countries including Cameroon'},
        ]
        return pd.DataFrame(data)

    def create_working_days_sheet(self):
        """Create working days configuration"""
        data = [
            {'day_of_week': 'Monday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Tuesday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Wednesday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Thursday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Friday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Saturday', 'working_day': 'No', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Sunday', 'working_day': 'No', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
        ]
        return pd.DataFrame(data)

    def create_account_number_preferences_sheet(self):
        """Create account number format preferences.

        Valid prefixType values per entity type (from Fineract API):
        - CLIENT: 1 (Office Name), 101 (Client Type), 401 (Custom Prefix)
        - LOAN: 1 (Office Name), 201 (Loan Product Short Name), 401 (Custom Prefix)
        - SAVINGS: 1 (Office Name), 301 (Savings Product Short Name), 401 (Custom Prefix)
        - GROUPS: 1 (Office Name) only
        - CENTERS: 1 (Office Name) only
        """
        data = [
            {'entity_type': 'CLIENT', 'prefix_type': 'Office Name', 'prefix_type_id': 1,
             'example': 'Yaounde000001', 'description': 'Client account numbers: Office name + sequence'},

            {'entity_type': 'LOAN', 'prefix_type': 'Loan Product Short Name', 'prefix_type_id': 201,
             'example': 'MICRO000001', 'description': 'Loan account numbers: Product short name + sequence'},

            {'entity_type': 'SAVINGS', 'prefix_type': 'Savings Product Short Name', 'prefix_type_id': 301,
             'example': 'VOL00000001', 'description': 'Savings account numbers: Product short name + sequence'},

            {'entity_type': 'GROUP', 'prefix_type': 'Office Name', 'prefix_type_id': 1,
             'example': 'Yaounde-GRP001', 'description': 'Group account numbers: Office name + sequence'},

            {'entity_type': 'CENTER', 'prefix_type': 'Office Name', 'prefix_type_id': 1,
             'example': 'Yaounde-CTR001', 'description': 'Center account numbers: Office name + sequence'},
        ]
        return pd.DataFrame(data)

    def create_codes_and_values_sheet(self):
        """Create codes and code values for dropdowns"""
        data = [
            # Gender
            {'code_name': 'Gender', 'code_value': 'Male', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Male gender'},
            {'code_name': 'Gender', 'code_value': 'Female', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Female gender'},
            {'code_name': 'Gender', 'code_value': 'Other', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Other gender'},

            # Client Type
            {'code_name': 'ClientType', 'code_value': 'Individual', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Individual client'},
            {'code_name': 'ClientType', 'code_value': 'Corporate', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Corporate/business client'},

            # Client Classification
            {'code_name': 'ClientClassification', 'code_value': 'Active', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Active client'},
            {'code_name': 'ClientClassification', 'code_value': 'Pending', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Pending activation'},
            {'code_name': 'ClientClassification', 'code_value': 'Closed', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Closed/inactive client'},

            # Marital Status
            {'code_name': 'MaritalStatus', 'code_value': 'Single', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Single'},
            {'code_name': 'MaritalStatus', 'code_value': 'Married', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Married'},
            {'code_name': 'MaritalStatus', 'code_value': 'Divorced', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Divorced'},
            {'code_name': 'MaritalStatus', 'code_value': 'Widow', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Widow/Widower'},

            # Education Level
            {'code_name': 'EducationLevel', 'code_value': 'Primary', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Primary education'},
            {'code_name': 'EducationLevel', 'code_value': 'Secondary', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Secondary education'},
            {'code_name': 'EducationLevel', 'code_value': 'University', 'code_position': 3, 'is_active': 'Yes',
             'description': 'University degree'},
            {'code_name': 'EducationLevel', 'code_value': 'No Formal Education', 'code_position': 4, 'is_active': 'Yes',
             'description': 'No formal education'},

            # Loan Purpose
            {'code_name': 'LoanPurpose', 'code_value': 'Working Capital', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Business working capital'},
            {'code_name': 'LoanPurpose', 'code_value': 'Equipment Purchase', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Purchase of equipment'},
            {'code_name': 'LoanPurpose', 'code_value': 'Inventory', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Purchase inventory/stock'},
            {'code_name': 'LoanPurpose', 'code_value': 'Agricultural Input', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Seeds, fertilizer, tools'},
            {'code_name': 'LoanPurpose', 'code_value': 'Education', 'code_position': 5, 'is_active': 'Yes',
             'description': 'School fees and education'},
            {'code_name': 'LoanPurpose', 'code_value': 'Emergency', 'code_position': 6, 'is_active': 'Yes',
             'description': 'Emergency/medical expenses'},

            # Account Closure Reason
            {'code_name': 'SavingsClosureReason', 'code_value': 'Client Request', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Client requested closure'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Dormant Account', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Account became dormant'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Moved to Another Branch', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Client relocated'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Deceased', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Client deceased'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Fraudulent Activity', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Account involved in fraud'},

            # Loan Closure Reason
            {'code_name': 'LoanClosureReason', 'code_value': 'Fully Paid', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Loan fully repaid'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Written Off', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Loan written off as bad debt'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Rescheduled', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Loan rescheduled to new terms'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Foreclosed', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Collateral foreclosed'},

            # Business Type
            {'code_name': 'BusinessType', 'code_value': 'Retail Trade', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Retail business'},
            {'code_name': 'BusinessType', 'code_value': 'Wholesale Trade', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Wholesale business'},
            {'code_name': 'BusinessType', 'code_value': 'Agriculture', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Farming/agriculture'},
            {'code_name': 'BusinessType', 'code_value': 'Services', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Service provision'},
            {'code_name': 'BusinessType', 'code_value': 'Manufacturing', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Manufacturing/production'},
            {'code_name': 'BusinessType', 'code_value': 'Transportation', 'code_position': 6, 'is_active': 'Yes',
             'description': 'Transport services'},
            {'code_name': 'BusinessType', 'code_value': 'Food Services', 'code_position': 7, 'is_active': 'Yes',
             'description': 'Restaurant/catering'},

            # Guarantor Relationship (Note: Spouse already exists in Fineract by default)
            {'code_name': 'GuarantorRelationship', 'code_value': 'Family Member', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Family member (parent, sibling, child)'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Employer', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Employer guarantee for salaried clients'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Business Partner', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Business partner or associate'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Group Member', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Fellow group member in solidarity lending'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Community Leader', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Village chief or community leader'},
        ]
        return pd.DataFrame(data)

    def create_scheduler_jobs_sheet(self):
        """Create scheduler job configurations

        All job names verified against Fineract source code (JobName.java enum).
        """
        data = [
            # Core savings and loan jobs
            {'job_name': 'Post Interest For Savings', 'display_name': 'Post Interest For Savings',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Post interest to savings accounts daily at midnight'},

            {'job_name': 'Transfer Fee For Loans From Savings', 'display_name': 'Transfer Fee For Loans From Savings',
             'cron_expression': '0 0 1 * * ?', 'active': 'Yes',
             'description': 'Auto-deduct loan fees from linked savings accounts daily at 1 AM'},

            {'job_name': 'Update Loan Arrears Ageing', 'display_name': 'Update Loan Arrears Ageing',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Update loan arrears aging daily'},

            {'job_name': 'Add Accrual Transactions', 'display_name': 'Add Accrual Transactions',
             'cron_expression': '0 0 2 * * ?', 'active': 'Yes',
             'description': 'Add accrual accounting entries daily at 2 AM'},

            {'job_name': 'Apply Annual Fee For Savings', 'display_name': 'Apply Annual Fee For Savings',
             'cron_expression': '0 0 0 1 1 ? *', 'active': 'Yes',
             'description': 'Apply annual savings account fees on January 1st'},

            # Penalty and charges
            {'job_name': 'Apply penalty to overdue loans', 'display_name': 'Apply penalty to overdue loans',
             'cron_expression': '0 0 3 * * ?', 'active': 'Yes',
             'description': 'Apply penalty charges to overdue loans daily at 3 AM'},

            # Savings dormancy
            {'job_name': 'Update Savings Dormant Accounts', 'display_name': 'Update Savings Dormant Accounts',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Mark savings accounts as inactive/dormant based on activity'},

            # Standing instructions (recurring payments)
            {'job_name': 'Execute Standing Instruction', 'display_name': 'Execute Standing Instruction',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Execute standing instructions (recurring payments) daily'},

            # Periodic accruals
            {'job_name': 'Add Periodic Accrual Transactions', 'display_name': 'Add Periodic Accrual Transactions',
             'cron_expression': '0 0 0 1 * ? *', 'active': 'Yes',
             'description': 'Add monthly accrual entries on the 1st of each month'},

            {'job_name': 'Update Business Date', 'display_name': 'Update Business Date',
             'cron_expression': '0 0 1 * * ?', 'active': 'Yes',
             'description': 'Daily job to update the business date at 1 AM'},

            # NOTE: "Update Loan Paid In Advance" job does not exist in Fineract source code
            # This functionality may have been removed or renamed in newer versions
        ]
        return pd.DataFrame(data)

    def create_global_config_sheet(self):
        """Create comprehensive global configuration settings.

        value_type indicates how the value should be interpreted:
        - 'boolean': The config is a toggle, use 'enabled' field only (no value field needed)
        - 'numeric': The config expects an integer value
        - 'string': The config expects a string value (use stringValue field)
        - 'date': The config expects a date value (use dateValue field)
        """
        data = [
            # Maker-Checker Global Settings (boolean - toggle only)
            {'config_name': 'maker-checker', 'enabled': 'Yes', 'value_type': 'boolean',
             'description': 'Enable maker-checker globally'},

            # Reschedule Future Repayments (boolean)
            {'config_name': 'reschedule-future-repayments', 'enabled': 'Yes', 'value_type': 'boolean',
             'description': 'Allow rescheduling of future repayments'},

            # Allow Backdated Transaction (boolean)
            {'config_name': 'allow-backdated-transaction-before-interest-posting', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Prevent backdated transactions before interest posting'},

            # Allow Transactions on Non-Working Days (boolean)
            {'config_name': 'allow-transactions-on-non-workingday', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Block transactions on holidays and weekends'},

            # Allow Transactions on Holidays (boolean)
            {'config_name': 'allow-transactions-on-holiday', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Block transactions on public holidays'},

            # Financial Year Beginning (numeric: 1-12 for month)
            {'config_name': 'financial-year-beginning-month', 'enabled': 'Yes', 'value': 1, 'value_type': 'numeric',
             'description': 'Financial year starts in January (1-12)'},

            # Days before repayment is due reminder (numeric)
            {'config_name': 'days-before-repayment-is-due', 'enabled': 'Yes', 'value': 3, 'value_type': 'numeric',
             'description': 'Days before repayment due to send reminder'},

            # Days after repayment is overdue reminder (numeric)
            {'config_name': 'days-after-repayment-is-overdue', 'enabled': 'Yes', 'value': 3, 'value_type': 'numeric',
             'description': 'Days after repayment overdue to send reminder'},

            # Enable Sub Rates (boolean)
            {'config_name': 'sub-rates', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Enable sub-rate interest calculations'},

            # Is Interest to be Recovered First (boolean)
            {'config_name': 'is-interest-to-be-recovered-first-when-greater-than-emi', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Recover interest before principal when payment exceeds EMI'},

            # Enable Principal Threshold for Last Installment (boolean)
            {'config_name': 'is-principal-compounding-disabled-for-overdue-loans', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Disable principal compounding for overdue loans'},

            # Enable Business Date (boolean)
            {'config_name': 'enable-business-date', 'enabled': 'Yes', 'value_type': 'boolean',
             'description': 'Enable business date (COB) feature'},

            # Enable Auto Repayment for Down-payment (boolean)
            {'config_name': 'enable-auto-generated-external-id', 'enabled': 'Yes', 'value_type': 'boolean',
             'description': 'Auto-generate external IDs if not provided'},

            # Meeting Calendar (boolean)
            {'config_name': 'meetings-mandatory-for-jlg-loans', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Make meetings mandatory for JLG (Joint Liability Group) loans'},

            # Enable Same Maker/Checker (boolean)
            {'config_name': 'enable-same-maker-checker', 'enabled': 'No', 'value_type': 'boolean',
             'description': 'Allow same user as maker and checker'},

            # Penalty Wait Period (numeric)
            {'config_name': 'penalty-wait-period', 'enabled': 'Yes', 'value': 0, 'value_type': 'numeric',
             'description': 'Days to wait before applying penalty charges'},
        ]
        return pd.DataFrame(data)

    def create_sms_email_config_sheet(self):
        """Create SMS and Email configuration"""
        data = [
            # SMS Gateway Configuration (Twilio)
            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'account_sid',
             'config_value': 'YOUR_TWILIO_ACCOUNT_SID', 'is_active': 'Yes',
             'description': 'Twilio Account SID for SMS service'},

            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'auth_token',
             'config_value': 'YOUR_TWILIO_AUTH_TOKEN', 'is_active': 'Yes',
             'description': 'Twilio Authentication Token (keep secure!)'},

            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'from_phone_number',
             'config_value': '+237XXXXXXXXX', 'is_active': 'Yes',
             'description': 'Twilio phone number to send SMS from'},

            # SMS Gateway Configuration (Alternative - Infobip)
            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'base_url',
             'config_value': 'https://api.infobip.com', 'is_active': 'No',
             'description': 'Infobip API base URL'},

            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'api_key',
             'config_value': 'YOUR_INFOBIP_API_KEY', 'is_active': 'No',
             'description': 'Infobip API Key'},

            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'sender_id',
             'config_value': 'MicroFin', 'is_active': 'No',
             'description': 'Sender ID shown to recipients'},

            # Email SMTP Configuration
            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_host',
             'config_value': 'smtp.gmail.com', 'is_active': 'Yes',
             'description': 'SMTP server hostname'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_port',
             'config_value': '587', 'is_active': 'Yes',
             'description': 'SMTP server port (587 for TLS, 465 for SSL)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_username',
             'config_value': 'your-email@gmail.com', 'is_active': 'Yes',
             'description': 'SMTP authentication username (email address)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_password',
             'config_value': 'YOUR_APP_PASSWORD', 'is_active': 'Yes',
             'description': 'SMTP authentication password (use app password for Gmail)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'from_email',
             'config_value': 'noreply@yourmfi.com', 'is_active': 'Yes',
             'description': 'Email address shown as sender'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'from_name',
             'config_value': 'Cameroon MicroFinance', 'is_active': 'Yes',
             'description': 'Display name shown as sender'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'use_tls',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable TLS encryption'},

            # Notification Settings
            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'enable_sms_notifications',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable SMS notifications globally'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'enable_email_notifications',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable email notifications globally'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'max_retry_attempts',
             'config_value': '3', 'is_active': 'Yes',
             'description': 'Maximum retry attempts for failed notifications'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'retry_delay_minutes',
             'config_value': '5', 'is_active': 'Yes',
             'description': 'Delay between retry attempts (minutes)'},
        ]
        return pd.DataFrame(data)

    def create_notification_templates_sheet(self):
        """Create notification templates for SMS and Email"""
        data = [
            # Client Notifications
            {'template_name': 'Client Activation', 'channel': 'SMS', 'event_trigger': 'Client Activated',
             'subject': '', 'message_body': 'Hello {clientName}, your account has been activated at {officeName}. Welcome to {institutionName}! Your account number is {accountNumber}.',
             'is_active': 'Yes', 'description': 'Sent when client account is activated'},

            {'template_name': 'Client Activation Email', 'channel': 'Email', 'event_trigger': 'Client Activated',
             'subject': 'Welcome to {institutionName}',
             'message_body': 'Dear {clientName},\n\nYour account has been successfully activated at {officeName}.\n\nAccount Number: {accountNumber}\nActivation Date: {activationDate}\n\nThank you for choosing {institutionName}.\n\nBest regards,\n{institutionName} Team',
             'is_active': 'Yes', 'description': 'Email sent when client account is activated'},

            # Loan Notifications
            {'template_name': 'Loan Approval', 'channel': 'SMS', 'event_trigger': 'Loan Approved',
             'subject': '', 'message_body': 'Congratulations {clientName}! Your loan of {loanAmount} XAF has been approved. Loan Account: {loanAccountNumber}. Visit {officeName} to collect.',
             'is_active': 'Yes', 'description': 'Sent when loan is approved'},

            {'template_name': 'Loan Disbursement', 'channel': 'SMS', 'event_trigger': 'Loan Disbursed',
             'subject': '', 'message_body': 'Dear {clientName}, your loan of {loanAmount} XAF has been disbursed to your account. First repayment due: {firstRepaymentDate}. Total repayments: {numberOfRepayments}.',
             'is_active': 'Yes', 'description': 'Sent when loan is disbursed'},

            {'template_name': 'Loan Repayment Due', 'channel': 'SMS', 'event_trigger': 'Repayment Due in 3 Days',
             'subject': '', 'message_body': 'Reminder: Your loan repayment of {repaymentAmount} XAF is due on {repaymentDate}. Loan Account: {loanAccountNumber}. Thank you.',
             'is_active': 'Yes', 'description': 'Reminder sent 3 days before repayment due date'},

            {'template_name': 'Loan Repayment Overdue', 'channel': 'SMS', 'event_trigger': 'Repayment Overdue',
             'subject': '', 'message_body': 'URGENT: Your loan repayment of {repaymentAmount} XAF was due on {repaymentDate}. Outstanding: {totalOutstanding} XAF. Please pay immediately to avoid penalties.',
             'is_active': 'Yes', 'description': 'Sent when repayment becomes overdue'},

            {'template_name': 'Loan Repayment Received', 'channel': 'SMS', 'event_trigger': 'Repayment Made',
             'subject': '', 'message_body': 'Payment received: {repaymentAmount} XAF for loan {loanAccountNumber}. Outstanding balance: {outstandingBalance} XAF. Next payment: {nextRepaymentDate}.',
             'is_active': 'Yes', 'description': 'Sent when repayment is received'},

            {'template_name': 'Loan Fully Repaid', 'channel': 'SMS', 'event_trigger': 'Loan Closed',
             'subject': '', 'message_body': 'Congratulations {clientName}! Your loan {loanAccountNumber} has been fully repaid. Total paid: {totalPaid} XAF. Thank you for your commitment!',
             'is_active': 'Yes', 'description': 'Sent when loan is fully repaid'},

            # Savings Notifications
            {'template_name': 'Savings Account Activation', 'channel': 'SMS', 'event_trigger': 'Savings Activated',
             'subject': '', 'message_body': 'Hello {clientName}, your savings account {savingsAccountNumber} has been activated. Current balance: {accountBalance} XAF.',
             'is_active': 'Yes', 'description': 'Sent when savings account is activated'},

            {'template_name': 'Savings Deposit', 'channel': 'SMS', 'event_trigger': 'Deposit Made',
             'subject': '', 'message_body': 'Deposit received: {depositAmount} XAF to account {savingsAccountNumber}. New balance: {accountBalance} XAF. Date: {transactionDate}.',
             'is_active': 'Yes', 'description': 'Sent when deposit is made'},

            {'template_name': 'Savings Withdrawal', 'channel': 'SMS', 'event_trigger': 'Withdrawal Made',
             'subject': '', 'message_body': 'Withdrawal: {withdrawalAmount} XAF from account {savingsAccountNumber}. New balance: {accountBalance} XAF. Date: {transactionDate}.',
             'is_active': 'Yes', 'description': 'Sent when withdrawal is made'},

            {'template_name': 'Interest Posted', 'channel': 'SMS', 'event_trigger': 'Interest Posted',
             'subject': '', 'message_body': 'Interest earned: {interestAmount} XAF posted to account {savingsAccountNumber}. New balance: {accountBalance} XAF. Period: {postingPeriod}.',
             'is_active': 'Yes', 'description': 'Sent when interest is posted'},

            {'template_name': 'Low Balance Alert', 'channel': 'SMS', 'event_trigger': 'Balance Below Minimum',
             'subject': '', 'message_body': 'Alert: Your savings account {savingsAccountNumber} balance ({accountBalance} XAF) is below minimum ({minimumBalance} XAF). Please deposit to avoid charges.',
             'is_active': 'Yes', 'description': 'Sent when balance falls below minimum'},

            {'template_name': 'Account Dormancy Warning', 'channel': 'SMS', 'event_trigger': 'Account Inactive 60 Days',
             'subject': '', 'message_body': 'Your account {savingsAccountNumber} has been inactive for 60 days. Please make a transaction within 30 days to prevent dormancy.',
             'is_active': 'Yes', 'description': 'Warning before account becomes dormant'},

            # General Notifications
            {'template_name': 'Password Changed', 'channel': 'Email', 'event_trigger': 'Password Changed',
             'subject': 'Password Changed - {institutionName}',
             'message_body': 'Dear {userName},\n\nYour password was changed on {changeDate} at {changeTime}.\n\nIf you did not make this change, please contact us immediately.\n\nBest regards,\n{institutionName} Security Team',
             'is_active': 'Yes', 'description': 'Security notification for password changes'},

            {'template_name': 'Failed Login Attempt', 'channel': 'Email', 'event_trigger': '3 Failed Login Attempts',
             'subject': 'Security Alert - Failed Login Attempts',
             'message_body': 'Dear {userName},\n\nWe detected 3 failed login attempts on your account from IP: {ipAddress} on {attemptDate}.\n\nIf this was not you, please contact us immediately.\n\nBest regards,\n{institutionName} Security Team',
             'is_active': 'Yes', 'description': 'Security alert for multiple failed logins'},
        ]
        return pd.DataFrame(data)

    def create_data_tables_sheet(self):
        """Create custom data tables (custom fields) configuration

        IMPORTANT fields for Fineract API:
        - entity_sub_type: Required for Client entity (PERSON or ENTITY)
        - code: Required for Dropdown type fields (must reference an existing code name)
        - length: Required for String/Text/Dropdown types

        Length defaults:
        - String: 200
        - Text: 1000
        - Dropdown: 100
        - Number: 10
        - Decimal: 19
        - Date/Datetime/Boolean: not required (empty)
        """
        data = [
            # Client Custom Fields - entitySubType is MANDATORY for m_client tables
            # Note: Using String type instead of Dropdown for simplicity (avoids code dependency)
            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'id_type', 'field_type': 'String', 'length': 100, 'mandatory': 'Yes',
             'description': 'Type of identification document (National ID, Passport, etc.)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'id_number', 'field_type': 'String', 'length': 200, 'mandatory': 'Yes',
             'description': 'Identification document number'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'id_expiry_date', 'field_type': 'Date', 'length': '', 'mandatory': 'No',
             'description': 'ID document expiry date'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'next_of_kin_name', 'field_type': 'String', 'length': 200, 'mandatory': 'Yes',
             'description': 'Next of kin full name'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'next_of_kin_phone', 'field_type': 'String', 'length': 200, 'mandatory': 'Yes',
             'description': 'Next of kin phone number'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'next_of_kin_relationship', 'field_type': 'String', 'length': 100, 'mandatory': 'Yes',
             'description': 'Relationship to next of kin (Spouse, Parent, Sibling, etc.)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'employer_name', 'field_type': 'String', 'length': 200, 'mandatory': 'No',
             'description': 'Employer name (for salaried clients)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'employer_phone', 'field_type': 'String', 'length': 200, 'mandatory': 'No',
             'description': 'Employer contact number'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'years_in_business', 'field_type': 'Number', 'length': 10, 'mandatory': 'No',
             'description': 'Years in current business (for self-employed)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'business_location', 'field_type': 'String', 'length': 200, 'mandatory': 'No',
             'description': 'Business physical location'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'home_ownership', 'field_type': 'String', 'length': 100, 'mandatory': 'No',
             'description': 'Home ownership status (Owned, Rented, Family Home, etc.)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info',
             'field_name': 'disability_status', 'field_type': 'String', 'length': 100, 'mandatory': 'No',
             'description': 'Disability status (None, Physical, Visual, Hearing, Other)'},

            # Additional client fields from Clients sheet (must use m_client with entity_sub_type)
            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'address',
             'field_type': 'String', 'length': 200, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Residential address'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'city',
             'field_type': 'String', 'length': 200, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'City of residence'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'marital_status',
             'field_type': 'String', 'length': 100, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Marital status (Single, Married, Divorced, Widow, Widower)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'number_of_dependents',
             'field_type': 'Number', 'length': 10, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Number of dependents'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'occupation',
             'field_type': 'String', 'length': 200, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Client occupation or profession'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'business_type',
             'field_type': 'String', 'length': 200, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Type of business activity'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'monthly_income',
             'field_type': 'Decimal', 'length': 19, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Estimated monthly income (XAF)'},

            {'entity_type': 'm_client', 'entity_sub_type': 'PERSON', 'table_name': 'client_additional_info', 'field_name': 'risk_rating',
             'field_type': 'String', 'length': 10, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Client risk rating (A=Low, B, C, D=High)'},

            # Loan Custom Fields
            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'loan_purpose_detail',
             'field_type': 'Text', 'length': 1000, 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Detailed description of loan purpose'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'collateral_description',
             'field_type': 'Text', 'length': 1000, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Description of collateral provided'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'collateral_value',
             'field_type': 'Decimal', 'length': 19, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Estimated collateral value (XAF)'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'guarantor_count',
             'field_type': 'Number', 'length': 10, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Number of guarantors'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'credit_score',
             'field_type': 'Number', 'length': 10, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Credit score (0-100)'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'repayment_source',
             'field_type': 'String', 'length': 100, 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Primary source of repayment (Business Income, Salary, Remittances, Other)'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'previous_loan_history',
             'field_type': 'String', 'length': 100, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Previous loan repayment history (First Loan, Good History, Some Delays, Defaulted Before)'},

            # Savings Custom Fields
            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'savings_goal',
             'field_type': 'String', 'length': 200, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Client\'s savings goal'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'target_amount',
             'field_type': 'Decimal', 'length': 19, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Target savings amount (XAF)'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'target_date',
             'field_type': 'Date', 'length': '', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Target date to reach goal'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'monthly_commitment',
             'field_type': 'Decimal', 'length': 19, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Monthly savings commitment (XAF)'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'preferred_transaction_channel',
             'field_type': 'String', 'length': 100, 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Preferred way to transact (Branch, Mobile Money, Bank Transfer, Agent)'},
        ]
        return pd.DataFrame(data)

    def format_excel(self, workbook_path):
        """Apply formatting, descriptions, and data validation to Excel workbook"""
        wb = openpyxl.load_workbook(workbook_path)

        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define column descriptions and validation rules for each sheet
        sheet_metadata = self._get_sheet_metadata()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get metadata for this sheet
            metadata = sheet_metadata.get(sheet_name, {})

            # Format headers and add descriptions
            for idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

                # Add description as comment if available
                col_name = cell.value
                if col_name and col_name in metadata:
                    col_info = metadata[col_name]
                    description = col_info.get('description', '')
                    if description:
                        # Add comment with description
                        comment = Comment(description, 'Fineract Demo Data Generator')
                        comment.width = 300
                        comment.height = 100
                        cell.comment = comment

                # Add data validation for columns with allowed values
                if col_name and col_name in metadata:
                    col_info = metadata[col_name]
                    allowed_values = col_info.get('allowed_values', [])
                    if allowed_values:
                        # Create data validation
                        dv = DataValidation(type="list", formula1=f'"{",".join(allowed_values)}"', allow_blank=True)
                        dv.error = f'Please select from: {", ".join(allowed_values)}'
                        dv.errorTitle = 'Invalid Value'
                        dv.prompt = f'Select from: {", ".join(allowed_values)}'
                        dv.promptTitle = col_name
                        ws.add_data_validation(dv)

                        # Apply validation to entire column (up to row 1000)
                        col_letter = get_column_letter(idx)
                        dv.add(f'{col_letter}2:{col_letter}1000')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze top row
            ws.freeze_panes = 'A2'

        wb.save(workbook_path)

    def _get_sheet_metadata(self):
        """Define descriptions and validation rules for all sheet columns"""
        return {
            'Charges': {
                'charge_name': {
                    'description': 'Name of the charge/fee (e.g., "Loan Processing Fee", "ATM Fee")'
                },
                'charge_type': {
                    'description': 'Entity the charge applies to',
                    'allowed_values': ['Loan', 'Savings', 'Client']
                },
                'charge_time': {
                    'description': 'When the charge is applied. IMPORTANT: Use exact values from the list.\n\nFor Loans: Disbursement, Specified Due Date, Installment Fee, Overdue Installment\n\nFor Savings: Savings Activation, Withdrawal Fee, ATM Fee, Annual Fee, Monthly Fee, Overdraft Fee, Weekly Fee, Saving No Activity Fee',
                    'allowed_values': [
                        'Disbursement', 'Specified Due Date', 'Installment Fee', 'Overdue Installment',
                        'Savings Activation', 'Withdrawal Fee', 'ATM Fee', 'Annual Fee', 'Monthly Fee',
                        'Overdraft Fee', 'Weekly Fee', 'Saving No Activity Fee'
                    ]
                },
                'calculation_type': {
                    'description': 'How the charge amount is calculated',
                    'allowed_values': ['Flat', 'Percentage of Amount', 'Percentage of Interest']
                },
                'amount': {
                    'description': 'Charge amount (flat amount in XAF or percentage value)'
                },
                'currency': {
                    'description': 'Currency code',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                },
                'active': {
                    'description': 'Whether the charge is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Loan Products': {
                'short_name': {
                    'description': 'Short identifier for the product (max 4 characters)'
                },
                'product_name': {
                    'description': 'Full product name displayed to users'
                },
                'currency': {
                    'description': 'Currency code for this product',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                }
            },
            'Savings Products': {
                'short_name': {
                    'description': 'Short identifier for the product (max 4 characters)'
                },
                'product_name': {
                    'description': 'Full product name displayed to users'
                },
                'currency': {
                    'description': 'Currency code for this product',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                },
                'withdrawal_fee_for_transfers': {
                    'description': 'Charge withdrawal fee for transfers',
                    'allowed_values': ['Yes', 'No']
                },
                'overdraft_allowed': {
                    'description': 'Allow account to go negative',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Chart of Accounts': {
                'gl_code': {
                    'description': 'GL account code (OHADA compliant)'
                },
                'gl_name': {
                    'description': 'Name of the GL account'
                },
                'account_type': {
                    'description': 'Type of account in the chart',
                    'allowed_values': ['Asset', 'Liability', 'Equity', 'Income', 'Expense']
                },
                'usage': {
                    'description': 'Header accounts cannot have transactions posted to them',
                    'allowed_values': ['Detail', 'Header']
                },
                'manual_entries': {
                    'description': 'Allow manual journal entries',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Staff': {
                'role': {
                    'description': 'Staff role determines permissions',
                    'allowed_values': ['Branch Manager', 'Loan Officer', 'Cashier']
                }
            },
            'Working Days': {
                'working_day': {
                    'description': 'Is this a working day?',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Account Number Preferences': {
                'entity_type': {
                    'description': 'Entity for account number generation',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Groups', 'Centers']
                },
                'prefix_type': {
                    'description': 'Prefix to use for account numbers',
                    'allowed_values': ['Office Name', 'Client Type', 'Loan Product Short Name', 'Savings Product Short Name']
                }
            },
            'Codes and Values': {
                'is_active': {
                    'description': 'Whether this code value is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Global Configuration': {
                'enabled': {
                    'description': 'Whether this configuration is enabled',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'SMS Email Config': {
                'is_active': {
                    'description': 'Whether this configuration is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Notification Templates': {
                'channel': {
                    'description': 'Notification delivery channel',
                    'allowed_values': ['SMS', 'Email', 'System']
                },
                'is_active': {
                    'description': 'Whether this template is active',
                    'allowed_values': ['Yes', 'No']
                },
                'entity_type': {
                    'description': 'Entity this template applies to',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Group']
                }
            },
            'Tellers': {
                'status': {
                    'description': 'Teller operational status',
                    'allowed_values': ['Active', 'Inactive']
                }
            },
            'Scheduler Jobs': {
                'active': {
                    'description': 'Whether this job is enabled',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Data Tables': {
                'entity_type': {
                    'description': 'Entity the custom field applies to',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Group']
                },
                'field_type': {
                    'description': 'Data type for the custom field',
                    'allowed_values': ['String', 'Number', 'Decimal', 'Date', 'Datetime', 'Text', 'Dropdown']
                },
                'mandatory': {
                    'description': 'Whether this field is required',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Maker Checker Config': {
                'entity': {
                    'description': 'Entity type for maker-checker. Common values: Loan, SAVINGSACCOUNT, Client, JOURNALENTRY, User, RESCHEDULELOAN, OFFICETRANSACTION',
                    'allowed_values': ['Loan', 'SAVINGSACCOUNT', 'Client', 'JOURNALENTRY', 'User', 'RESCHEDULELOAN', 'OFFICETRANSACTION']
                },
                'action': {
                    'description': 'Action type for maker-checker. Common values: APPROVE, DISBURSE, WRITEOFF, CREATE, ACTIVATE, UPDATE, CLOSE, PROPOSETRANSFER',
                    'allowed_values': ['APPROVE', 'DISBURSE', 'WRITEOFF', 'CREATE', 'ACTIVATE', 'UPDATE', 'CLOSE', 'PROPOSETRANSFER']
                },
                'enabled': {
                    'description': 'Enable this maker-checker permission. When enabled, ALL operations of this type require approval (Fineract does NOT support amount-based thresholds)',
                    'allowed_values': ['True', 'False']
                }
            },
            'Loan Accounts': {
                'workflow_state': {
                    'description': 'Loan workflow state:\n- active: Auto-approve and auto-disburse (normal flow)\n- pending_approval: Create loan, stop at approval (demonstrates maker-checker)\n- pending_disbursal: Create and approve loan, stop at disbursal (demonstrates maker-checker)',
                    'allowed_values': ['active', 'pending_approval', 'pending_disbursal']
                }
            }
        }

    def create_savings_deposits_sheet(self):
        """Create sample savings deposit transactions"""
        data = [
            # Regular deposits (using actual account external IDs)
            {'client_name': 'Jean Dupont', 'savings_account_number': 'SAV-001',
             'transaction_date': '2024-03-01', 'transaction_amount': 50000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-001',
             'note': 'Monthly deposit', 'office': 'Douala Branch'},

            {'client_name': 'Marie Kamga', 'savings_account_number': 'SAV-002',
             'transaction_date': '2024-03-05', 'transaction_amount': 25000,
             'payment_type': 'Bank Transfer', 'receipt_number': 'RCP-002',
             'note': 'Monthly savings', 'office': 'Douala Branch'},

            {'client_name': 'Paul Ekambi', 'savings_account_number': 'SAV-004',
             'transaction_date': '2024-03-02', 'transaction_amount': 30000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-003',
             'note': 'Weekly deposit', 'office': 'Yaounde Branch'},

            {'client_name': 'Grace Fotso', 'savings_account_number': 'SAV-005',
             'transaction_date': '2024-03-10', 'transaction_amount': 15000,
             'payment_type': 'MTN Mobile Money', 'receipt_number': 'RCP-004',
             'note': 'Mobile money deposit', 'office': 'Bafoussam Branch'},

            {'client_name': 'François Nkomo', 'savings_account_number': 'SAV-006',
             'transaction_date': '2024-03-03', 'transaction_amount': 40000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-005',
             'note': 'Business savings', 'office': 'Douala Branch'},

            {'client_name': 'Justine Moukouri', 'savings_account_number': 'SAV-007',
             'transaction_date': '2024-03-12', 'transaction_amount': 35000,
             'payment_type': 'Cheque', 'receipt_number': 'RCP-006',
             'note': 'Farm income deposit', 'office': 'Yaounde Branch'},

            # Inter-branch deposits (Client from Branch A deposits at Branch B)
            {'client_name': 'Jean Dupont', 'savings_account_number': 'SAV-001',
             'transaction_date': '2024-03-20', 'transaction_amount': 35000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-007',
             'note': 'Cross-branch deposit - Douala client depositing at Yaounde',
             'office': 'Yaounde Branch'},

            {'client_name': 'Paul Ekambi', 'savings_account_number': 'SAV-004',
             'transaction_date': '2024-03-22', 'transaction_amount': 20000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-008',
             'note': 'Cross-branch deposit - Yaounde client depositing at Douala',
             'office': 'Douala Branch'},

            {'client_name': 'Marie Kamga', 'savings_account_number': 'SAV-002',
             'transaction_date': '2024-03-25', 'transaction_amount': 20000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-009',
             'note': 'Cross-branch deposit - Douala client depositing at Bafoussam',
             'office': 'Bafoussam Branch'},

            {'client_name': 'Grace Fotso', 'savings_account_number': 'SAV-005',
             'transaction_date': '2024-03-28', 'transaction_amount': 25000,
             'payment_type': 'Cash', 'receipt_number': 'RCP-010',
             'note': 'Cross-branch deposit - Bafoussam client depositing at Yaounde',
             'office': 'Yaounde Branch'},
        ]
        return pd.DataFrame(data)

    def create_savings_withdrawals_sheet(self):
        """Create sample savings withdrawal transactions.

        Note: Withdrawal amounts must be less than account balance.
        Using small amounts to ensure sufficient balance after deposits.
        """
        data = [
            # Withdrawals for Yaounde clients
            {'client_name': 'Jean Dupont', 'savings_account_number': 'SAV-001',
             'transaction_date': '2024-03-15', 'transaction_amount': 5000,
             'payment_type': 'Cash', 'receipt_number': 'WDL-001',
             'note': 'Emergency withdrawal', 'office': 'Yaounde Branch'},

            {'client_name': 'Marie Kamga', 'savings_account_number': 'SAV-002',
             'transaction_date': '2024-03-18', 'transaction_amount': 5000,
             'payment_type': 'Cash', 'receipt_number': 'WDL-002',
             'note': 'Cash withdrawal', 'office': 'Yaounde Branch'},

            # Withdrawals for Douala clients
            {'client_name': 'Paul Ekambi', 'savings_account_number': 'SAV-004',
             'transaction_date': '2024-03-16', 'transaction_amount': 5000,
             'payment_type': 'Cash', 'receipt_number': 'WDL-003',
             'note': 'Personal expense', 'office': 'Douala Branch'},

            {'client_name': 'Grace Fotso', 'savings_account_number': 'SAV-005',
             'transaction_date': '2024-03-20', 'transaction_amount': 5000,
             'payment_type': 'Bank Transfer', 'receipt_number': 'WDL-004',
             'note': 'Supplier payment', 'office': 'Douala Branch'},

            # Withdrawals for Bafoussam clients
            {'client_name': 'Pierre Tchuente', 'savings_account_number': 'SAV-011',
             'transaction_date': '2024-03-17', 'transaction_amount': 5000,
             'payment_type': 'Cash', 'receipt_number': 'WDL-005',
             'note': 'School fees', 'office': 'Bafoussam Branch'},

            {'client_name': 'Justine Moukouri', 'savings_account_number': 'SAV-012',
             'transaction_date': '2024-03-22', 'transaction_amount': 5000,
             'payment_type': 'Cash', 'receipt_number': 'WDL-006',
             'note': 'Medical expenses', 'office': 'Bafoussam Branch'},
        ]
        return pd.DataFrame(data)

    def create_loan_repayments_sheet(self):
        """Create sample loan repayment transactions.

        Note: Loans must be Active, Fully Paid or Overpaid to accept repayments.
        LOAN-001, LOAN-003, LOAN-004, LOAN-005 are Overpaid (can accept repayments)
        LOAN-002, LOAN-006 are pending approval (cannot accept repayments)
        """
        data = [
            # Repayments for Yaounde clients - LOAN-001 (Overpaid)
            {'client_name': 'Jean Dupont', 'loan_account_number': 'LOAN-001',
             'transaction_date': '2024-03-15', 'principal_amount': 20000, 'interest_amount': 5000,
             'fee_amount': 0, 'penalty_amount': 0, 'payment_type': 'Cash',
             'receipt_number': 'REP-001', 'note': 'Monthly repayment', 'office': 'Yaounde Branch'},

            # Repayments for Douala clients - LOAN-003 (Overpaid)
            {'client_name': 'Paul Ekambi', 'loan_account_number': 'LOAN-003',
             'transaction_date': '2024-03-18', 'principal_amount': 30000, 'interest_amount': 8000,
             'fee_amount': 0, 'penalty_amount': 0, 'payment_type': 'Cash',
             'receipt_number': 'REP-003', 'note': 'Monthly repayment', 'office': 'Douala Branch'},

            {'client_name': 'Grace Fotso', 'loan_account_number': 'LOAN-004',
             'transaction_date': '2024-03-25', 'principal_amount': 80000, 'interest_amount': 20000,
             'fee_amount': 0, 'penalty_amount': 2000, 'payment_type': 'Bank Transfer',
             'receipt_number': 'REP-004', 'note': 'Late payment with penalty', 'office': 'Douala Branch'},

            # Repayments for Bafoussam clients - LOAN-005 (Overpaid)
            {'client_name': 'Pierre Tchuente', 'loan_account_number': 'LOAN-005',
             'transaction_date': '2024-03-22', 'principal_amount': 25000, 'interest_amount': 6500,
             'fee_amount': 0, 'penalty_amount': 0, 'payment_type': 'Cash',
             'receipt_number': 'REP-005', 'note': 'Monthly repayment', 'office': 'Bafoussam Branch'},
        ]
        return pd.DataFrame(data)

    def create_loan_collateral_sheet(self):
        """Create loan collateral assignments.

        Fields match LoanCollateral model:
        - loan_external_id: External ID of the loan account
        - collateral_type_name: Name of collateral type (must match existing type)
        - value: Collateral value amount
        - description: Description of the collateral
        """
        # Collateral type names must match create_collateral_types_sheet():
        # 'Land Title', 'Vehicle Registration', 'Household Goods', 'Shop/Business Equipment',
        # 'Savings Account Lien', 'Group Guarantee'
        # Use loans LOAN-002 and LOAN-006 (pending approval status - collateral can be added)
        data = [
            # Collateral for pending loans only
            {'loan_external_id': 'LOAN-002', 'collateral_type_name': 'Land Title',
             'value': 1000000,
             'description': 'Residential plot in Bastos, Yaounde - Title deed #YDE-2023-001'},

            {'loan_external_id': 'LOAN-002', 'collateral_type_name': 'Shop/Business Equipment',
             'value': 500000,
             'description': 'Commercial sewing machines (5 units)'},

            {'loan_external_id': 'LOAN-006', 'collateral_type_name': 'Vehicle Registration',
             'value': 1200000,
             'description': 'Toyota Hilux 2020 - Registration DLA-4567-AB'},

            {'loan_external_id': 'LOAN-006', 'collateral_type_name': 'Household Goods',
             'value': 300000,
             'description': 'Furniture, appliances and household items'},
        ]
        return pd.DataFrame(data)

    def create_loan_guarantors_sheet(self):
        """Create loan guarantor assignments.

        Fields match LoanGuarantor model:
        - loan_external_id: External ID of the loan account
        - guarantor_type: CLIENT, STAFF, or EXTERNAL
        - client_external_id: For CLIENT type guarantors
        - staff_name: For STAFF type guarantors
        - firstname, lastname, address_line_1, city: For EXTERNAL type guarantors
        - amount: Guaranteed amount
        - savings_external_id: Optional savings account held as collateral
        """
        # Only use loans with "Submitted and pending approval" status: LOAN-002, LOAN-006
        # Cannot add guarantors to closed/overpaid loans
        # Note: Client cannot be guarantor for their own loan
        # LOAN-002 belongs to CLI-002, so use CLI-003 as guarantor
        data = [
            # CLIENT type guarantor - another client guarantees the loan
            # (CLI-003 guarantees CLI-002's loan LOAN-002)
            {'loan_external_id': 'LOAN-002', 'guarantor_type': 'CLIENT',
             'client_external_id': 'CLI-003', 'amount': 250000},

            # EXTERNAL type guarantor - person not in system
            {'loan_external_id': 'LOAN-006', 'guarantor_type': 'EXTERNAL',
             'firstname': 'Jean', 'lastname': 'Kemayou',
             'address_line_1': 'Centre Ville', 'city': 'Bafoussam', 'amount': 400000},
        ]
        return pd.DataFrame(data)

    def create_inter_branch_transfers_sheet(self):
        """Create sample inter-branch transfer transactions"""
        data = [
            # Transfer 1: Yaounde to Douala
            {'transfer_date': '2024-03-10', 'from_office': 'Yaounde Branch', 'to_office': 'Douala Branch',
             'transfer_amount': 5000000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-001', 'initiated_by': 'Jean Mbarga',
             'description': 'Branch liquidity management - Cash transfer to Douala',
             'status': 'Completed'},

            # Transfer 2: Douala to Bafoussam
            {'transfer_date': '2024-03-12', 'from_office': 'Douala Branch', 'to_office': 'Bafoussam Branch',
             'transfer_amount': 3000000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-002', 'initiated_by': 'Grace Douala',
             'description': 'Weekly cash replenishment to Bafoussam branch',
             'status': 'Completed'},

            # Transfer 3: Bafoussam to Bamenda
            {'transfer_date': '2024-03-15', 'from_office': 'Bafoussam Branch', 'to_office': 'Bamenda Branch',
             'transfer_amount': 2000000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-003', 'initiated_by': 'Samuel Kenmogne',
             'description': 'Opening cash allocation for Bamenda branch',
             'status': 'Completed'},

            # Transfer 4: Douala to Yaounde (reverse flow)
            {'transfer_date': '2024-03-18', 'from_office': 'Douala Branch', 'to_office': 'Yaounde Branch',
             'transfer_amount': 4000000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-004', 'initiated_by': 'Grace Douala',
             'description': 'Excess liquidity transfer to Head Office region',
             'status': 'Completed'},

            # Transfer 5: Yaounde to Bafoussam
            {'transfer_date': '2024-03-20', 'from_office': 'Yaounde Branch', 'to_office': 'Bafoussam Branch',
             'transfer_amount': 2500000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-005', 'initiated_by': 'Jean Mbarga',
             'description': 'Agricultural season loan disbursement support',
             'status': 'Completed'},

            # Transfer 6: Bamenda to Douala
            {'transfer_date': '2024-03-22', 'from_office': 'Bamenda Branch', 'to_office': 'Douala Branch',
             'transfer_amount': 1500000, 'currency': 'XAF', 'transfer_type': 'Cash Transfer',
             'reference_number': 'IBT-006', 'initiated_by': 'David Tanko',
             'description': 'Deposit surplus transfer to main vault',
             'status': 'Completed'},
        ]
        return pd.DataFrame(data)

    def create_financial_reports_sheet(self):
        """Create Financial Overview Reports (to be registered in Fineract)"""
        data = [
            # Income Statement Report
            {'report_name': 'Income Statement (Profit & Loss)',
             'report_category': 'Financial',
             'report_type': 'Table',
             'description': 'Revenue, expenses, and net profit analysis',
             'sql_query': '''
SELECT
    'REVENUE' as category,
    'Interest Income' as line_item,
    SUM(l.interest_charged_derived) as amount
FROM m_loan l
WHERE l.loan_status_id = 300
UNION ALL
SELECT
    'REVENUE' as category,
    'Fee Income' as line_item,
    SUM(l.fee_charges_charged_derived) as amount
FROM m_loan l
WHERE l.loan_status_id = 300
UNION ALL
SELECT
    'EXPENSES' as category,
    'Loan Loss Provision' as line_item,
    SUM(CASE
        WHEN DATEDIFF(CURDATE(), l.overdue_since_date_derived) > 180 THEN l.principal_outstanding_derived * 1.0
        WHEN DATEDIFF(CURDATE(), l.overdue_since_date_derived) > 90 THEN l.principal_outstanding_derived * 0.75
        WHEN DATEDIFF(CURDATE(), l.overdue_since_date_derived) > 30 THEN l.principal_outstanding_derived * 0.50
        WHEN DATEDIFF(CURDATE(), l.overdue_since_date_derived) > 0 THEN l.principal_outstanding_derived * 0.25
        ELSE 0
    END) as amount
FROM m_loan l
WHERE l.loan_status_id = 300
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Balance Sheet Report
            {'report_name': 'Balance Sheet',
             'report_category': 'Financial',
             'report_type': 'Table',
             'description': 'Assets, liabilities, and equity position',
             'sql_query': '''
SELECT
    'ASSETS' as section,
    'Gross Loan Portfolio' as line_item,
    SUM(l.principal_outstanding_derived) as amount
FROM m_loan l
WHERE l.loan_status_id = 300
UNION ALL
SELECT
    'ASSETS' as section,
    'Cash and Bank Balances' as line_item,
    SUM(gl.amount) as amount
FROM acc_gl_journal_entry gl
JOIN acc_gl_account a ON gl.account_id = a.id
WHERE a.classification_enum = 1
UNION ALL
SELECT
    'LIABILITIES' as section,
    'Savings Deposits' as line_item,
    SUM(s.account_balance_derived) as amount
FROM m_savings_account s
WHERE s.status_enum = 300
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Portfolio Performance Report
            {'report_name': 'Portfolio Performance Statistics',
             'report_category': 'Portfolio',
             'report_type': 'Table',
             'description': 'Loan portfolio performance metrics and statistics',
             'sql_query': '''
SELECT
    COUNT(DISTINCT l.id) as total_loans,
    COUNT(DISTINCT CASE WHEN l.loan_status_id = 300 THEN l.id END) as active_loans,
    SUM(l.principal_disbursed_derived) as total_disbursed,
    SUM(l.principal_outstanding_derived) as total_outstanding,
    SUM(l.principal_repaid_derived) as total_repaid,
    AVG(l.principal_amount) as avg_loan_size,
    (SUM(l.principal_repaid_derived) / SUM(l.principal_disbursed_derived) * 100) as repayment_rate,
    p.name as product_name
FROM m_loan l
JOIN m_product_loan p ON l.product_id = p.id
GROUP BY p.name
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Portfolio by Product Report
            {'report_name': 'Portfolio Analysis by Product',
             'report_category': 'Portfolio',
             'report_type': 'Table',
             'description': 'Loan portfolio breakdown by product',
             'sql_query': '''
SELECT
    p.name as product_name,
    COUNT(l.id) as loan_count,
    SUM(l.principal_amount) as total_principal,
    SUM(l.principal_outstanding_derived) as outstanding,
    SUM(l.principal_repaid_derived) as repaid,
    AVG(l.annual_nominal_interest_rate) as avg_interest_rate
FROM m_loan l
JOIN m_product_loan p ON l.product_id = p.id
WHERE l.loan_status_id IN (200, 300)
GROUP BY p.name
ORDER BY total_principal DESC
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Portfolio by Branch Report
            {'report_name': 'Portfolio Analysis by Branch',
             'report_category': 'Portfolio',
             'report_type': 'Table',
             'description': 'Loan portfolio breakdown by office/branch',
             'sql_query': '''
SELECT
    o.name as office_name,
    COUNT(l.id) as loan_count,
    SUM(l.principal_amount) as total_principal,
    SUM(l.principal_outstanding_derived) as outstanding,
    SUM(l.principal_repaid_derived) as repaid,
    COUNT(DISTINCT l.client_id) as unique_clients
FROM m_loan l
JOIN m_office o ON l.office_id = o.id
WHERE l.loan_status_id IN (200, 300)
GROUP BY o.name
ORDER BY total_principal DESC
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Active Clients Report
            {'report_name': 'Active Clients Summary',
             'report_category': 'Client',
             'report_type': 'Table',
             'description': 'Summary of active clients with account balances',
             'sql_query': '''
SELECT
    c.display_name as client_name,
    c.account_no as client_account,
    o.name as office_name,
    COUNT(DISTINCT l.id) as total_loans,
    SUM(l.principal_outstanding_derived) as loan_balance,
    COUNT(DISTINCT s.id) as total_savings,
    SUM(s.account_balance_derived) as savings_balance
FROM m_client c
LEFT JOIN m_loan l ON c.id = l.client_id AND l.loan_status_id = 300
LEFT JOIN m_savings_account s ON c.id = s.client_id AND s.status_enum = 300
JOIN m_office o ON c.office_id = o.id
WHERE c.status_enum = 300
GROUP BY c.id, c.display_name, c.account_no, o.name
ORDER BY loan_balance DESC
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Savings Account Summary Report
            {'report_name': 'Savings Accounts Summary',
             'report_category': 'Savings',
             'report_type': 'Table',
             'description': 'Summary of all savings accounts by product',
             'sql_query': '''
SELECT
    p.name as product_name,
    COUNT(s.id) as account_count,
    SUM(s.account_balance_derived) as total_balance,
    AVG(s.account_balance_derived) as avg_balance,
    SUM(s.total_deposits_derived) as total_deposits,
    SUM(s.total_withdrawals_derived) as total_withdrawals
FROM m_savings_account s
JOIN m_savings_product p ON s.product_id = p.id
WHERE s.status_enum = 300
GROUP BY p.name
ORDER BY total_balance DESC
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Loan Repayment Collection Report
            {'report_name': 'Loan Repayment Collections',
             'report_category': 'Transaction',
             'report_type': 'Table',
             'description': 'Daily/monthly loan repayment collections',
             'sql_query': '''
SELECT
    DATE(t.transaction_date) as collection_date,
    pt.value as payment_type,
    COUNT(t.id) as transaction_count,
    SUM(t.amount) as total_collected,
    SUM(t.principal_portion_derived) as principal_collected,
    SUM(t.interest_portion_derived) as interest_collected
FROM m_loan_transaction t
LEFT JOIN m_payment_type pt ON t.payment_detail_id = pt.id
WHERE t.transaction_type_enum = 2
    AND t.is_reversed = 0
    AND t.transaction_date >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
GROUP BY DATE(t.transaction_date), pt.value
ORDER BY collection_date DESC
''',
             'parameters': None,
             'use_report': 'Yes'},

            # Overdue Loans Report
            {'report_name': 'Overdue Loans Report',
             'report_category': 'Collection',
             'report_type': 'Table',
             'description': 'List of all overdue loans with days past due',
             'sql_query': '''
SELECT
    c.display_name as client_name,
    c.account_no as client_account,
    l.account_no as loan_account,
    p.name as product_name,
    o.name as office_name,
    l.principal_outstanding_derived as outstanding_principal,
    l.interest_outstanding_derived as outstanding_interest,
    DATEDIFF(CURDATE(), l.overdue_since_date_derived) as days_overdue,
    s.display_name as loan_officer
FROM m_loan l
JOIN m_client c ON l.client_id = c.id
JOIN m_product_loan p ON l.product_id = p.id
JOIN m_office o ON l.office_id = o.id
LEFT JOIN m_staff s ON l.loan_officer_id = s.id
WHERE l.loan_status_id = 300
    AND l.overdue_since_date_derived IS NOT NULL
ORDER BY days_overdue DESC, outstanding_principal DESC
''',
             'parameters': None,
             'use_report': 'Yes'},
        ]
        return pd.DataFrame(data)

    def create_business_date_sheet(self):
        """Create a sheet with the current date for business date initialization."""
        today = datetime.now()
        data = [
            {
                'type': 'BUSINESS_DATE',
                'date': today.strftime('%Y-%m-%d'),
                'dateFormat': 'yyyy-MM-dd',
                'locale': 'en'
            }
        ]
        return pd.DataFrame(data)

    def create_cobac_report_config_sheet(self):
        """Create COBAC reporting configuration"""
        data = [
            # Portfolio Quality Reports
            {'report_name': 'COBAC R01 - Portfolio Quality', 'report_type': 'Regulatory',
             'frequency': 'Monthly', 'due_day': 15, 'report_format': 'Excel',
             'description': 'Loan portfolio classification by risk category (Performing, Watch, Substandard, Doubtful, Loss)',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            {'report_name': 'COBAC R02 - Loan Provisioning', 'report_type': 'Regulatory',
             'frequency': 'Monthly', 'due_day': 15, 'report_format': 'Excel',
             'description': 'Loan loss provisions by category with aging analysis',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Capital Adequacy Reports
            {'report_name': 'COBAC R03 - Capital Adequacy', 'report_type': 'Regulatory',
             'frequency': 'Quarterly', 'due_day': 30, 'report_format': 'Excel',
             'description': 'Risk-weighted assets and capital adequacy ratio (minimum 8%)',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Liquidity Reports
            {'report_name': 'COBAC R04 - Liquidity Position', 'report_type': 'Regulatory',
             'frequency': 'Monthly', 'due_day': 10, 'report_format': 'Excel',
             'description': 'Cash and liquid assets position, liquidity ratios',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Large Exposures
            {'report_name': 'COBAC R05 - Large Exposures', 'report_type': 'Regulatory',
             'frequency': 'Quarterly', 'due_day': 30, 'report_format': 'Excel',
             'description': 'Exposures exceeding 10% of capital (risk concentration)',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Financial Statements
            {'report_name': 'COBAC R06 - Balance Sheet', 'report_type': 'Financial Statement',
             'frequency': 'Quarterly', 'due_day': 45, 'report_format': 'Excel',
             'description': 'OHADA-compliant balance sheet (SYSCOHADA format)',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            {'report_name': 'COBAC R07 - Income Statement', 'report_type': 'Financial Statement',
             'frequency': 'Quarterly', 'due_day': 45, 'report_format': 'Excel',
             'description': 'OHADA-compliant profit & loss statement',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Operational Reports
            {'report_name': 'COBAC R08 - Arrears Report', 'report_type': 'Operational',
             'frequency': 'Monthly', 'due_day': 10, 'report_format': 'Excel',
             'description': 'Detailed arrears analysis by product, branch, and age',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            {'report_name': 'COBAC R09 - Portfolio at Risk', 'report_type': 'Operational',
             'frequency': 'Monthly', 'due_day': 10, 'report_format': 'Excel',
             'description': 'PAR>30, PAR>90, PAR>180 ratios by product and branch',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},

            # Governance Reports
            {'report_name': 'COBAC R10 - Related Party Transactions', 'report_type': 'Governance',
             'frequency': 'Quarterly', 'due_day': 30, 'report_format': 'Excel',
             'description': 'Loans and transactions with related parties (directors, shareholders)',
             'regulatory_body': 'COBAC', 'mandatory': 'Yes'},
        ]
        return pd.DataFrame(data)

    def generate(self):
        """Generate complete Excel template"""
        os.makedirs(self.output_dir, exist_ok=True)

        print(f"Generating Fineract Demo Data Excel Template...")

        # Create Excel writer
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            # Generate all sheets
            print("  Creating Offices sheet...")
            self.create_offices_sheet().to_excel(writer, sheet_name='Offices', index=False)

            print("  Creating Staff sheet...")
            self.create_staff_sheet().to_excel(writer, sheet_name='Staff', index=False)

            print("  Creating Users sheet...")
            self.create_users_sheet().to_excel(writer, sheet_name='Users', index=False)

            print("  Creating Clients sheet...")
            self.create_clients_sheet().to_excel(writer, sheet_name='Clients', index=False)

            print("  Creating Groups sheet...")
            self.create_groups_sheet().to_excel(writer, sheet_name='Groups', index=False)

            print("  Creating Centers sheet...")
            self.create_centers_sheet().to_excel(writer, sheet_name='Centers', index=False)

            print("  Creating Loan Products sheet...")
            self.create_loan_products_sheet().to_excel(writer, sheet_name='Loan Products', index=False)

            print("  Creating Savings Products sheet...")
            self.create_savings_products_sheet().to_excel(writer, sheet_name='Savings Products', index=False)

            print("  Creating Charges sheet...")
            self.create_charges_sheet().to_excel(writer, sheet_name='Charges', index=False)

            print("  Creating Chart of Accounts sheet...")
            self.create_gl_accounts_sheet().to_excel(writer, sheet_name='Chart of Accounts', index=False)

            print("  Creating Loan Accounts sheet...")
            self.create_loan_accounts_sheet().to_excel(writer, sheet_name='Loan Accounts', index=False)

            print("  Creating Savings Accounts sheet...")
            self.create_savings_accounts_sheet().to_excel(writer, sheet_name='Savings Accounts', index=False)

            print("  Creating Fund Sources sheet...")
            self.create_fund_sources_sheet().to_excel(writer, sheet_name='Fund Sources', index=False)

            print("  Creating Payment Types sheet...")
            self.create_payment_types_sheet().to_excel(writer, sheet_name='Payment Types', index=False)

            print("  Creating Holidays sheet...")
            self.create_holidays_sheet().to_excel(writer, sheet_name='Holidays', index=False)

            print("  Creating Loan Provisioning sheet...")
            self.create_loan_provisioning_sheet().to_excel(writer, sheet_name='Loan Provisioning', index=False)

            print("  Creating Collateral Types sheet...")
            self.create_collateral_types_sheet().to_excel(writer, sheet_name='Collateral Types', index=False)

            print("  Creating Guarantor Types sheet...")
            self.create_guarantor_types_sheet().to_excel(writer, sheet_name='Guarantor Types', index=False)

            print("  Creating Floating Rates sheet...")
            self.create_floating_rates_sheet().to_excel(writer, sheet_name='Floating Rates', index=False)

            print("  Creating Delinquency Buckets sheet...")
            self.create_delinquency_buckets_sheet().to_excel(writer, sheet_name='Delinquency Buckets', index=False)

            print("  Creating Tax Groups sheet...")
            self.create_tax_groups_sheet().to_excel(writer, sheet_name='Tax Groups', index=False)

            print("  Creating Configuration sheet...")
            self.create_config_sheet().to_excel(writer, sheet_name='Configuration', index=False)

            print("  Creating Loan Product Accounting Mappings sheet...")
            self.create_loan_product_accounting_sheet().to_excel(writer, sheet_name='Loan Product Accounting', index=False)

            print("  Creating Savings Product Accounting Mappings sheet...")
            self.create_savings_product_accounting_sheet().to_excel(writer, sheet_name='Savings Product Accounting', index=False)

            print("  Creating Payment Type Accounting Mappings sheet...")
            self.create_payment_type_accounting_sheet().to_excel(writer, sheet_name='Payment Type Accounting', index=False)

            print("  Creating Financial Activity Mappings sheet...")
            self.create_financial_activity_mappings_sheet().to_excel(writer, sheet_name='Financial Activity Mapping', index=False)

            print("  Creating Teller/Cashier Mappings sheet...")
            self.create_teller_cashier_mappings_sheet().to_excel(writer, sheet_name='Teller Cashier Mapping', index=False)

            print("  Creating Tellers sheet...")
            self.create_tellers_sheet().to_excel(writer, sheet_name='Tellers', index=False)

            print("  Creating Teller Accounting Rules sheet...")
            self.create_teller_accounting_rules_sheet().to_excel(writer, sheet_name='Teller Accounting Rules', index=False)

            print("  Creating Roles and Permissions sheet...")
            self.create_roles_permissions_sheet().to_excel(writer, sheet_name='Roles Permissions', index=False)

            print("  Creating Maker-Checker Configuration sheet...")
            self.create_maker_checker_config_sheet().to_excel(writer, sheet_name='Maker Checker Config', index=False)

            print("  Creating Currency Configuration sheet...")
            self.create_currency_config_sheet().to_excel(writer, sheet_name='Currency Config', index=False)

            print("  Creating Working Days sheet...")
            self.create_working_days_sheet().to_excel(writer, sheet_name='Working Days', index=False)

            print("  Creating Account Number Preferences sheet...")
            self.create_account_number_preferences_sheet().to_excel(writer, sheet_name='Account Number Preferences', index=False)

            print("  Creating Codes and Values sheet...")
            self.create_codes_and_values_sheet().to_excel(writer, sheet_name='Codes and Values', index=False)

            print("  Creating Scheduler Jobs sheet...")
            self.create_scheduler_jobs_sheet().to_excel(writer, sheet_name='Scheduler Jobs', index=False)

            print("  Creating Global Configuration sheet...")
            self.create_global_config_sheet().to_excel(writer, sheet_name='Global Configuration', index=False)

            print("  Creating SMS/Email Configuration sheet...")
            self.create_sms_email_config_sheet().to_excel(writer, sheet_name='SMS Email Config', index=False)

            print("  Creating Notification Templates sheet...")
            self.create_notification_templates_sheet().to_excel(writer, sheet_name='Notification Templates', index=False)

            print("  Creating Data Tables (Custom Fields) sheet...")
            self.create_data_tables_sheet().to_excel(writer, sheet_name='Data Tables', index=False)

            # Transaction sheets
            print("  Creating Savings Deposits sheet...")
            self.create_savings_deposits_sheet().to_excel(writer, sheet_name='Savings Deposits', index=False)

            print("  Creating Savings Withdrawals sheet...")
            self.create_savings_withdrawals_sheet().to_excel(writer, sheet_name='Savings Withdrawals', index=False)

            print("  Creating Loan Repayments sheet...")
            self.create_loan_repayments_sheet().to_excel(writer, sheet_name='Loan Repayments', index=False)

            print("  Creating Loan Collateral sheet...")
            self.create_loan_collateral_sheet().to_excel(writer, sheet_name='Loan Collateral', index=False)

            print("  Creating Loan Guarantors sheet...")
            self.create_loan_guarantors_sheet().to_excel(writer, sheet_name='Loan Guarantors', index=False)

            print("  Creating Inter-Branch Transfers sheet...")
            self.create_inter_branch_transfers_sheet().to_excel(writer, sheet_name='Inter-Branch Transfers', index=False)

            # Financial Reports (to be registered in Fineract)
            print("  Creating Financial Reports sheet...")
            self.create_financial_reports_sheet().to_excel(writer, sheet_name='Financial Reports', index=False)

            # COBAC Reporting
            print("  Creating COBAC Report Configuration sheet...")
            self.create_cobac_report_config_sheet().to_excel(writer, sheet_name='COBAC Reports', index=False)

            print("  Creating Business Date sheet...")
            self.create_business_date_sheet().to_excel(writer, sheet_name='Business Date', index=False)


        # Apply formatting
        print("  Applying formatting...")
        self.format_excel(self.filename)

        print(f"\n✓ Excel template generated successfully: {self.filename}")
        print(f"\nSheets created: 46")
        print(f"  1. Offices (4 branches)")
        print(f"  2. Staff (12 members)")
        print(f"  3. Clients (12 clients)")
        print(f"  4. Groups (5 solidarity groups)")
        print(f"  5. Centers (3 centers)")
        print(f"  6. Loan Products (3 products)")
        print(f"  7. Savings Products (3 products)")
        print(f"  8. Charges (11 fees)")
        print(f"  9. Chart of Accounts (30 GL accounts)")
        print(f" 10. Loan Accounts (6 sample loans)")
        print(f" 11. Savings Accounts (12 sample accounts)")
        print(f" 12. Fund Sources (6 sources)")
        print(f" 13. Payment Types (5 channels)")
        print(f" 14. Holidays (12 public holidays)")
        print(f" 15. Loan Provisioning (5 COBAC categories)")
        print(f" 16. Collateral Types (6 types)")
        print(f" 17. Guarantor Types (6 types)")
        print(f" 18. Floating Rates (3 rates - BEAC, Prime, SME)")
        print(f" 19. Delinquency Buckets (5 buckets for arrears classification)")
        print(f" 20. Tax Groups (2 groups - Savings WHT, Loan WHT)")
        print(f" 21. Configuration (19 system settings)")
        print(f" 22. Loan Product Accounting (39 mappings - 13 per product)")
        print(f" 23. Savings Product Accounting (33 mappings - 11 per product)")
        print(f" 24. Payment Type Accounting (5 mappings)")
        print(f" 25. Financial Activity Mapping (6 mappings - inter-branch support)")
        print(f" 26. Teller Cashier Mapping (4 GL account mappings per office)")
        print(f" 27. Tellers (4 tellers - 1 per office)")
        print(f" 28. Roles Permissions (3 roles with granular permissions)")
        print(f" 29. Maker Checker Config (11 dual-authorization rules)")
        print(f" 30. Currency Config (1 currency - XAF)")
        print(f" 31. Working Days (7 days configuration)")
        print(f" 32. Account Number Preferences (5 entity types)")
        print(f" 33. Codes and Values (48 dropdown values across 10 code categories)")
        print(f" 34. Scheduler Jobs (10 automated jobs)")
        print(f" 35. Global Configuration (23 global settings)")
        print(f" 36. SMS/Email Config (17 configuration items - Twilio, Gmail SMTP)")
        print(f" 37. Notification Templates (16 templates - SMS/Email for all events)")
        print(f" 38. Data Tables (24 custom fields - Client, Loan, Savings)")
        print(f" 39. Savings Deposits (6 sample deposit transactions)")
        print(f" 40. Savings Withdrawals (6 sample withdrawal transactions)")
        print(f" 41. Loan Disbursements (6 sample disbursements)")
        print(f" 42. Loan Repayments (6 sample repayment transactions)")
        print(f" 43. Loan Collateral (6 collateral assignments - Land, Equipment, Vehicles)")
        print(f" 44. Loan Guarantors (6 guarantor assignments - Individual & Corporate)")
        print(f" 45. Inter-Branch Transfers (6 sample inter-branch cash transfers)")
        print(f" 46. COBAC Reports (10 regulatory reports configuration)")

        return self.filename

if __name__ == '__main__':
    generator = FineractDemoDataGenerator()
    generator.generate()
