#!/usr/bin/env python3
"""
Fineract Demo Data Excel Generator
Creates a comprehensive Excel template for Cameroon OHADA-compliant microfinance setup
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import os

class FineractDemoDataGenerator:
    def __init__(self, output_dir='../output'):
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.filename = f'{output_dir}/fineract_demo_data_{self.timestamp}.xlsx'
        self.writer = None

    def create_offices_sheet(self):
        """Create offices/branches data"""
        data = [
            {
                'office_name': 'Yaounde Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-01-15',
                'external_id': 'BR-YDE-001',
                'address': 'Avenue Kennedy, Quartier Bastos',
                'city': 'Yaounde',
                'region': 'Centre',
                'phone': '+237 222 20 45 67',
                'email': 'yaounde@mfi.cm'
            },
            {
                'office_name': 'Douala Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-01-15',
                'external_id': 'BR-DLA-001',
                'address': 'Boulevard de la Liberte, Akwa',
                'city': 'Douala',
                'region': 'Littoral',
                'phone': '+237 233 42 56 78',
                'email': 'douala@mfi.cm'
            },
            {
                'office_name': 'Bafoussam Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-02-01',
                'external_id': 'BR-BAF-001',
                'address': 'Marche A, Centre Ville',
                'city': 'Bafoussam',
                'region': 'Ouest',
                'phone': '+237 233 44 12 34',
                'email': 'bafoussam@mfi.cm'
            },
            {
                'office_name': 'Bamenda Branch',
                'parent_office': 'Head Office',
                'opening_date': '2024-02-15',
                'external_id': 'BR-BAM-001',
                'address': 'Commercial Avenue, Nkwen',
                'city': 'Bamenda',
                'region': 'Nord-Ouest',
                'phone': '+237 233 36 45 67',
                'email': 'bamenda@mfi.cm'
            }
        ]
        return pd.DataFrame(data)

    def create_staff_sheet(self):
        """Create staff members with role assignments"""
        data = [
            # Yaounde Branch Staff
            {'firstname': 'Jean', 'lastname': 'Mbarga', 'office': 'Yaounde Branch', 'role': 'Branch Manager',
             'username': 'manager.yaounde', 'email': 'j.mbarga@mfi.cm', 'mobile': '+237 677 12 34 56',
             'joining_date': '2024-01-15', 'external_id': 'STF-001'},

            {'firstname': 'Marie', 'lastname': 'Ngo Balla', 'office': 'Yaounde Branch', 'role': 'Loan Officer',
             'username': 'loan.yaounde', 'email': 'm.ngoballa@mfi.cm', 'mobile': '+237 677 23 45 67',
             'joining_date': '2024-01-20', 'external_id': 'STF-002'},

            {'firstname': 'Paul', 'lastname': 'Atangana', 'office': 'Yaounde Branch', 'role': 'Cashier',
             'username': 'cashier.yaounde', 'email': 'p.atangana@mfi.cm', 'mobile': '+237 677 34 56 78',
             'joining_date': '2024-01-20', 'external_id': 'STF-003'},

            # Douala Branch Staff
            {'firstname': 'Grace', 'lastname': 'Douala', 'office': 'Douala Branch', 'role': 'Branch Manager',
             'username': 'manager.douala', 'email': 'g.douala@mfi.cm', 'mobile': '+237 677 45 67 89',
             'joining_date': '2024-01-15', 'external_id': 'STF-004'},

            {'firstname': 'Emmanuel', 'lastname': 'Ewondo', 'office': 'Douala Branch', 'role': 'Loan Officer',
             'username': 'loan.douala', 'email': 'e.ewondo@mfi.cm', 'mobile': '+237 677 56 78 90',
             'joining_date': '2024-01-20', 'external_id': 'STF-005'},

            {'firstname': 'Francine', 'lastname': 'Makang', 'office': 'Douala Branch', 'role': 'Cashier',
             'username': 'cashier.douala', 'email': 'f.makang@mfi.cm', 'mobile': '+237 677 67 89 01',
             'joining_date': '2024-01-20', 'external_id': 'STF-006'},

            # Bafoussam Branch Staff
            {'firstname': 'Bernard', 'lastname': 'Kamga', 'office': 'Bafoussam Branch', 'role': 'Branch Manager',
             'username': 'manager.bafoussam', 'email': 'b.kamga@mfi.cm', 'mobile': '+237 677 78 90 12',
             'joining_date': '2024-02-01', 'external_id': 'STF-007'},

            {'firstname': 'Justine', 'lastname': 'Tchuente', 'office': 'Bafoussam Branch', 'role': 'Loan Officer',
             'username': 'loan.bafoussam', 'email': 'j.tchuente@mfi.cm', 'mobile': '+237 677 89 01 23',
             'joining_date': '2024-02-05', 'external_id': 'STF-008'},

            {'firstname': 'David', 'lastname': 'Fotso', 'office': 'Bafoussam Branch', 'role': 'Cashier',
             'username': 'cashier.bafoussam', 'email': 'd.fotso@mfi.cm', 'mobile': '+237 677 90 12 34',
             'joining_date': '2024-02-05', 'external_id': 'STF-009'},

            # Bamenda Branch Staff
            {'firstname': 'Peter', 'lastname': 'Nkeng', 'office': 'Bamenda Branch', 'role': 'Branch Manager',
             'username': 'manager.bamenda', 'email': 'p.nkeng@mfi.cm', 'mobile': '+237 677 01 23 45',
             'joining_date': '2024-02-15', 'external_id': 'STF-010'},

            {'firstname': 'Alice', 'lastname': 'Fon', 'office': 'Bamenda Branch', 'role': 'Loan Officer',
             'username': 'loan.bamenda', 'email': 'a.fon@mfi.cm', 'mobile': '+237 677 13 35 57',
             'joining_date': '2024-02-20', 'external_id': 'STF-011'},

            {'firstname': 'Joseph', 'lastname': 'Tanyi', 'office': 'Bamenda Branch', 'role': 'Cashier',
             'username': 'cashier.bamenda', 'email': 'j.tanyi@mfi.cm', 'mobile': '+237 677 24 46 68',
             'joining_date': '2024-02-20', 'external_id': 'STF-012'},
        ]
        return pd.DataFrame(data)

    def create_clients_sheet(self):
        """Create sample clients"""
        data = [
            {'firstname': 'Akoumba', 'lastname': 'Ngono', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Female', 'mobile': '+237 690 12 34 56', 'email': 'a.ngono@gmail.com',
             'date_of_birth': '1985-03-15', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-001', 'activation_date': '2024-01-20', 'national_id': 'CM1234567890',
             'address': 'Bonaberi, Rue des Fleurs', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 3, 'occupation': 'Petty Trader', 'business_type': 'Retail Trade',
             'monthly_income': 150000, 'risk_rating': 'B'},

            {'firstname': 'Ibrahim', 'lastname': 'Mahamat', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 690 23 45 67', 'email': 'i.mahamat@yahoo.fr',
             'date_of_birth': '1978-07-22', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-002', 'activation_date': '2024-01-25', 'national_id': 'CM2345678901',
             'address': 'Ndokoti, Carrefour Ange Raphael', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 5, 'occupation': 'Transport Owner', 'business_type': 'Transportation',
             'monthly_income': 300000, 'risk_rating': 'A'},

            {'firstname': 'Marie Claire', 'lastname': 'Tchouta', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Female', 'mobile': '+237 690 34 56 78', 'email': 'm.tchouta@gmail.com',
             'date_of_birth': '1990-11-10', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-003', 'activation_date': '2024-02-01', 'national_id': 'CM3456789012',
             'address': 'Melen, Quartier Fouda', 'city': 'Yaounde', 'marital_status': 'Single',
             'number_of_dependents': 1, 'occupation': 'Hairdresser', 'business_type': 'Services',
             'monthly_income': 120000, 'risk_rating': 'B'},

            {'firstname': 'Jean Paul', 'lastname': 'Kamdem', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'gender': 'Male', 'mobile': '+237 690 45 67 89', 'email': 'jp.kamdem@yahoo.fr',
             'date_of_birth': '1982-05-18', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-004', 'activation_date': '2024-02-05', 'national_id': 'CM4567890123',
             'address': 'Tamdja, Marche des Vivres', 'city': 'Bafoussam', 'marital_status': 'Married',
             'number_of_dependents': 4, 'occupation': 'Farmer', 'business_type': 'Agriculture',
             'monthly_income': 200000, 'risk_rating': 'B'},

            {'firstname': 'Claudine', 'lastname': 'Ekotto', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Female', 'mobile': '+237 690 56 78 90', 'email': 'c.ekotto@gmail.com',
             'date_of_birth': '1988-09-25', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-005', 'activation_date': '2024-02-10', 'national_id': 'CM5678901234',
             'address': 'New Bell, Rue Joffre', 'city': 'Douala', 'marital_status': 'Widow',
             'number_of_dependents': 2, 'occupation': 'Restaurant Owner', 'business_type': 'Food Services',
             'monthly_income': 180000, 'risk_rating': 'B'},

            {'firstname': 'Patrick', 'lastname': 'Ndongo', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Male', 'mobile': '+237 690 67 89 01', 'email': 'p.ndongo@hotmail.com',
             'date_of_birth': '1975-12-03', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-006', 'activation_date': '2024-02-15', 'national_id': 'CM6789012345',
             'address': 'Omnisport, Carrefour Warda', 'city': 'Yaounde', 'marital_status': 'Married',
             'number_of_dependents': 6, 'occupation': 'Wholesaler', 'business_type': 'Wholesale Trade',
             'monthly_income': 500000, 'risk_rating': 'A'},

            {'firstname': 'Beatrice', 'lastname': 'Nana', 'office': 'Bamenda Branch', 'staff': 'loan.bamenda',
             'gender': 'Female', 'mobile': '+237 690 78 90 12', 'email': 'b.nana@gmail.com',
             'date_of_birth': '1992-04-14', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-007', 'activation_date': '2024-02-20', 'national_id': 'CM7890123456',
             'address': 'Nkwen, Mile 4', 'city': 'Bamenda', 'marital_status': 'Single',
             'number_of_dependents': 0, 'occupation': 'Seamstress', 'business_type': 'Services',
             'monthly_income': 100000, 'risk_rating': 'C'},

            {'firstname': 'Samuel', 'lastname': 'Ebong', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 690 89 01 23', 'email': 's.ebong@yahoo.fr',
             'date_of_birth': '1980-08-30', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-008', 'activation_date': '2024-03-01', 'national_id': 'CM8901234567',
             'address': 'Bepanda, Rond Point', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 3, 'occupation': 'Mechanic', 'business_type': 'Services',
             'monthly_income': 160000, 'risk_rating': 'B'},

            {'firstname': 'Bernadette', 'lastname': 'Fotso', 'office': 'Bafoussam Branch', 'staff': 'loan.bafoussam',
             'gender': 'Female', 'mobile': '+237 690 90 12 34', 'email': 'b.fotso@gmail.com',
             'date_of_birth': '1987-06-20', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-009', 'activation_date': '2024-03-05', 'national_id': 'CM9012345678',
             'address': 'Famla, Marche Central', 'city': 'Bafoussam', 'marital_status': 'Married',
             'number_of_dependents': 2, 'occupation': 'Vegetable Seller', 'business_type': 'Retail Trade',
             'monthly_income': 90000, 'risk_rating': 'C'},

            {'firstname': 'Francois', 'lastname': 'Manga', 'office': 'Yaounde Branch', 'staff': 'loan.yaounde',
             'gender': 'Male', 'mobile': '+237 690 01 23 45', 'email': 'f.manga@hotmail.com',
             'date_of_birth': '1983-02-28', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-010', 'activation_date': '2024-03-10', 'national_id': 'CM0123456789',
             'address': 'Elig-Essono, Nouvelle Route', 'city': 'Yaounde', 'marital_status': 'Divorced',
             'number_of_dependents': 2, 'occupation': 'Carpenter', 'business_type': 'Services',
             'monthly_income': 140000, 'risk_rating': 'B'},

            {'firstname': 'Agnes', 'lastname': 'Fon', 'office': 'Bamenda Branch', 'staff': 'loan.bamenda',
             'gender': 'Female', 'mobile': '+237 691 12 34 56', 'email': 'a.fon@gmail.com',
             'date_of_birth': '1991-10-05', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-011', 'activation_date': '2024-03-15', 'national_id': 'CM1234567891',
             'address': 'Ntarikon, Junction', 'city': 'Bamenda', 'marital_status': 'Single',
             'number_of_dependents': 1, 'occupation': 'Boutique Owner', 'business_type': 'Retail Trade',
             'monthly_income': 110000, 'risk_rating': 'C'},

            {'firstname': 'Robert', 'lastname': 'Essomba', 'office': 'Douala Branch', 'staff': 'loan.douala',
             'gender': 'Male', 'mobile': '+237 691 23 45 67', 'email': 'r.essomba@yahoo.fr',
             'date_of_birth': '1979-01-17', 'client_type': 'Individual', 'client_classification': 'Active',
             'external_id': 'CLI-012', 'activation_date': '2024-03-20', 'national_id': 'CM2345678902',
             'address': 'Akwa, Boulevard de la Republique', 'city': 'Douala', 'marital_status': 'Married',
             'number_of_dependents': 4, 'occupation': 'Electronics Trader', 'business_type': 'Retail Trade',
             'monthly_income': 250000, 'risk_rating': 'A'},
        ]
        return pd.DataFrame(data)

    def create_loan_products_sheet(self):
        """Create loan products configuration"""
        data = [
            {
                'product_name': 'Microcredit Solidarity Loan',
                'short_name': 'MSOL',
                'description': 'Small loans for petty traders and artisans (solidarity group)',
                'currency': 'XAF',
                'principal_min': 50000,
                'principal_default': 200000,
                'principal_max': 500000,
                'number_of_repayments_min': 3,
                'number_of_repayments_default': 6,
                'number_of_repayments_max': 12,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 18.0,
                'interest_rate_default': 24.0,
                'interest_rate_max': 30.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'grace_on_principal_periods': 0,
                'grace_on_interest_periods': 0,
                'processing_fee_percent': 2.0,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.5,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'No'
            },
            {
                'product_name': 'SME Business Loan',
                'short_name': 'SBIZ',
                'description': 'Medium loans for small and medium enterprises',
                'currency': 'XAF',
                'principal_min': 500000,
                'principal_default': 2000000,
                'principal_max': 10000000,
                'number_of_repayments_min': 6,
                'number_of_repayments_default': 12,
                'number_of_repayments_max': 24,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 15.0,
                'interest_rate_default': 20.0,
                'interest_rate_max': 25.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'grace_on_principal_periods': 1,
                'grace_on_interest_periods': 0,
                'processing_fee_percent': 1.5,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.0,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'No'
            },
            {
                'product_name': 'Agricultural Seasonal Loan',
                'short_name': 'ASEA',
                'description': 'Seasonal loans for agricultural activities',
                'currency': 'XAF',
                'principal_min': 100000,
                'principal_default': 500000,
                'principal_max': 3000000,
                'number_of_repayments_min': 3,
                'number_of_repayments_default': 6,
                'number_of_repayments_max': 9,
                'repayment_frequency': 'Monthly',
                'interest_rate_min': 12.0,
                'interest_rate_default': 18.0,
                'interest_rate_max': 22.0,
                'interest_calculation_period': 'Same as repayment period',
                'amortization_type': 'Equal installments',
                'interest_type': 'Declining Balance',
                'grace_on_principal_periods': 2,
                'grace_on_interest_periods': 1,
                'processing_fee_percent': 1.0,
                'insurance_mandatory': 'Yes',
                'insurance_percent': 1.5,
                'minimum_gap_between_installments': 30,
                'allow_partial_period_interest_calculation': 'Yes'
            }
        ]
        return pd.DataFrame(data)

    def create_savings_products_sheet(self):
        """Create savings products configuration"""
        data = [
            {
                'product_name': 'Voluntary Savings Account',
                'short_name': 'VSAV',  # Max 4 characters for Fineract
                'description': 'Standard voluntary savings account',
                'currency': 'XAF',
                'nominal_annual_interest_rate': 3.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 10000,
                'minimum_balance': 5000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'Yes',
                'days_to_inactive': 180,
                'days_to_dormancy': 365,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled, but tax groups need to be created first
                'lock_in_period': 0
            },
            {
                'product_name': 'Fixed Deposit Account',
                'short_name': 'FDEP',  # Max 4 characters for Fineract
                'description': 'Fixed deposit with higher interest rate',
                'currency': 'XAF',
                'nominal_annual_interest_rate': 6.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 100000,
                'minimum_balance': 100000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'No',
                'days_to_inactive': 0,
                'days_to_dormancy': 0,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled, but tax groups need to be created first
                'lock_in_period': 180
            },
            {
                'product_name': 'Mandatory Group Savings',
                'short_name': 'MGRP',  # Max 4 characters for Fineract
                'description': 'Mandatory savings for group loan members',
                'currency': 'XAF',
                'nominal_annual_interest_rate': 2.0,
                'interest_compounding_period': 'Monthly',
                'interest_posting_period': 'Monthly',
                'interest_calculation_type': 'Daily Balance',
                'interest_calculation_days_in_year': 365,
                'minimum_opening_balance': 5000,
                'minimum_balance': 5000,
                'overdraft_allowed': 'No',
                'withdrawal_fee_for_transfers': 'No',
                'allow_dormancy_tracking': 'No',
                'days_to_inactive': 0,
                'days_to_dormancy': 0,
                'withhold_tax': 0,  # Disabled - taxGroupId is mandatory if enabled, but tax groups need to be created first
                'lock_in_period': 0
            }
        ]
        return pd.DataFrame(data)

    def create_charges_sheet(self):
        """Create fees and charges"""
        data = [
            {'charge_name': 'Loan Processing Fee', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 2.0, 'currency': 'XAF', 'charge_time': 'Disbursement', 'active': 'Yes'},

            {'charge_name': 'Late Payment Penalty', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 5.0, 'currency': 'XAF', 'charge_time': 'Overdue Installment', 'active': 'Yes'},

            {'charge_name': 'Loan Insurance', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 1.5, 'currency': 'XAF', 'charge_time': 'Disbursement', 'active': 'Yes'},

            {'charge_name': 'Account Opening Fee', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 1000, 'currency': 'XAF', 'charge_time': 'Activation', 'active': 'Yes'},

            {'charge_name': 'Monthly Account Maintenance', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 500, 'currency': 'XAF', 'charge_time': 'Monthly', 'active': 'Yes'},

            {'charge_name': 'Withdrawal Fee', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 200, 'currency': 'XAF', 'charge_time': 'Withdrawal', 'active': 'Yes'},

            {'charge_name': 'ATM Card Fee', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 5000, 'currency': 'XAF', 'charge_time': 'Annual', 'active': 'Yes'},

            # Note: 'On Demand' charge time type is not valid for Savings charges in Fineract
            # Valid charge time types for savings: Activation, Withdrawal, Annual, Monthly, etc.
            # {'charge_name': 'Account Statement Fee', 'charge_type': 'Savings', 'calculation_type': 'Flat',
            #  'amount': 1000, 'currency': 'XAF', 'charge_time': 'On Demand', 'active': 'Yes'},

            {'charge_name': 'Loan Restructuring Fee', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 1.0, 'currency': 'XAF', 'charge_time': 'Specified Due Date', 'active': 'Yes'},

            {'charge_name': 'Early Repayment Penalty', 'charge_type': 'Loan', 'calculation_type': 'Percentage of Amount',
             'amount': 3.0, 'currency': 'XAF', 'charge_time': 'Disbursement', 'active': 'No'},

            {'charge_name': 'Dormant Account Reactivation', 'charge_type': 'Savings', 'calculation_type': 'Flat',
             'amount': 2000, 'currency': 'XAF', 'charge_time': 'Activation', 'active': 'Yes'},
        ]
        return pd.DataFrame(data)

    def create_gl_accounts_sheet(self):
        """Create Chart of Accounts - OHADA compliant"""
        data = [
            # ASSETS - Class 1-5
            {'gl_code': '41', 'gl_name': 'Banks - Current Accounts', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Bank current accounts'},

            {'gl_code': '42', 'gl_name': 'Cash on Hand', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Physical cash in tellers and vaults'},

            {'gl_code': '51', 'gl_name': 'Loans to Clients - Principal', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Outstanding loan principal'},

            {'gl_code': '52', 'gl_name': 'Loans to Clients - Interest Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued interest receivable'},

            {'gl_code': '53', 'gl_name': 'Loans to Clients - Fees Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued fees receivable'},

            {'gl_code': '54', 'gl_name': 'Loans to Clients - Penalties Receivable', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued penalties receivable'},

            {'gl_code': '55', 'gl_name': 'Provision for Loan Losses', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Accumulated provision for bad loans (contra account)'},

            {'gl_code': '122', 'gl_name': 'Due from Other Branches (Receivable)', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Inter-branch receivables'},

            # LIABILITIES - Class 1-5
            {'gl_code': '61', 'gl_name': 'Voluntary Savings Accounts', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Customer voluntary savings deposits'},

            {'gl_code': '62', 'gl_name': 'Fixed Deposit Accounts', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Fixed deposit liabilities'},

            {'gl_code': '63', 'gl_name': 'Mandatory Group Savings', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Mandatory group savings liabilities'},

            {'gl_code': '64', 'gl_name': 'Savings Interest Payable', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Accrued interest payable to savers'},

            {'gl_code': '131', 'gl_name': 'Due to Other Branches (Payable)', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Inter-branch payables'},

            {'gl_code': '71', 'gl_name': 'Share Capital', 'account_type': 'Equity',
             'classification': 'Equity', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Paid-up share capital'},

            {'gl_code': '72', 'gl_name': 'Retained Earnings', 'account_type': 'Equity',
             'classification': 'Equity', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Accumulated retained earnings'},

            # INCOME - Class 7
            {'gl_code': '81', 'gl_name': 'Interest Income on Loans', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Interest earned from loans'},

            {'gl_code': '82', 'gl_name': 'Fee Income - Loan Processing', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Loan processing fees'},

            {'gl_code': '83', 'gl_name': 'Penalty Income', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Late payment penalties'},

            {'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Savings account fees'},

            # EXPENSES - Class 6
            {'gl_code': '91', 'gl_name': 'Interest Expense on Savings', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'No',
             'description': 'Interest paid to savers'},

            {'gl_code': '92', 'gl_name': 'Loan Loss Provision Expense', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Provision for loan losses'},

            {'gl_code': '93', 'gl_name': 'Loan Write-off Expense', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Written-off loans'},

            {'gl_code': '94', 'gl_name': 'Staff Salaries', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Employee salaries and wages'},

            {'gl_code': '95', 'gl_name': 'Office Rent', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Office rental expenses'},

            {'gl_code': '96', 'gl_name': 'Utilities', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Utilities expenses'},

            {'gl_code': '97', 'gl_name': 'Tax Expense - Withholding Tax', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Withholding tax on interest'},

            {'gl_code': '141', 'gl_name': 'Tax Payable - WHT', 'account_type': 'Liability',
             'classification': 'Current Liability', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Withholding tax payable'},

            # Mobile Money Accounts
            {'gl_code': '43', 'gl_name': 'MTN Mobile Money', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'MTN MoMo balance'},

            {'gl_code': '44', 'gl_name': 'Orange Money', 'account_type': 'Asset',
             'classification': 'Current Asset', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Orange Money balance'},

            {'gl_code': '85', 'gl_name': 'Mobile Money Transfer Fees', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Mobile money transaction fees'},

            # Teller Cash Shortage/Overage Accounts
            {'gl_code': '98', 'gl_name': 'Cash Shortage', 'account_type': 'Expense',
             'classification': 'Operating Expense', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Cash shortages from teller operations'},

            {'gl_code': '86', 'gl_name': 'Cash Overage', 'account_type': 'Income',
             'classification': 'Revenue', 'usage': 'Detail', 'manual_entries': 'Yes',
             'description': 'Cash overages from teller operations'},
        ]
        return pd.DataFrame(data)

    def create_loan_accounts_sheet(self):
        """Create sample loan accounts"""
        data = [
            {'client_external_id': 'CLI-001', 'product': 'MSOL', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-05', 'disbursed_on': '2024-02-09', 'principal': 200000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-001'},

            {'client_external_id': 'CLI-002', 'product': 'SBIZ', 'submitted_on': '2024-02-05',
             'approved_on': '2024-02-08', 'disbursed_on': '2024-02-12', 'principal': 2000000,
             'loan_term': 12, 'interest_rate': 20.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Bank Loan - Ecobank', 'external_id': 'LOAN-002'},

            {'client_external_id': 'CLI-003', 'product': 'MSOL', 'submitted_on': '2024-02-10',
             'approved_on': '2024-02-12', 'disbursed_on': '2024-02-15', 'principal': 150000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.yaounde',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-003'},

            {'client_external_id': 'CLI-004', 'product': 'ASEA', 'submitted_on': '2024-02-12',
             'approved_on': '2024-02-15', 'disbursed_on': '2024-02-19', 'principal': 500000,
             'loan_term': 6, 'interest_rate': 18.0, 'loan_officer': 'loan.bafoussam',
             'fund_source': 'Donor - AfDB', 'external_id': 'LOAN-004'},

            {'client_external_id': 'CLI-005', 'product': 'MSOL', 'submitted_on': '2024-02-15',
             'approved_on': '2024-02-18', 'disbursed_on': '2024-02-20', 'principal': 180000,
             'loan_term': 6, 'interest_rate': 24.0, 'loan_officer': 'loan.douala',
             'fund_source': 'Own Capital', 'external_id': 'LOAN-005'},

            {'client_external_id': 'CLI-006', 'product': 'SBIZ', 'submitted_on': '2024-02-18',
             'approved_on': '2024-02-20', 'disbursed_on': '2024-02-25', 'principal': 3000000,
             'loan_term': 18, 'interest_rate': 18.0, 'loan_officer': 'loan.yaounde',
             'fund_source': 'Bank Loan - Ecobank', 'external_id': 'LOAN-006'},
        ]
        return pd.DataFrame(data)

    def create_savings_accounts_sheet(self):
        """Create sample savings accounts"""
        data = [
            {'client_external_id': 'CLI-001', 'product': 'VSAV', 'submitted_on': '2024-01-20',
             'approved_on': '2024-01-20', 'activated_on': '2024-01-20', 'initial_deposit': 25000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-001'},

            {'client_external_id': 'CLI-002', 'product': 'VSAV', 'submitted_on': '2024-01-25',
             'approved_on': '2024-01-25', 'activated_on': '2024-01-25', 'initial_deposit': 50000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-002'},

            {'client_external_id': 'CLI-002', 'product': 'FDEP', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-01', 'activated_on': '2024-02-01', 'initial_deposit': 100000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-003'},

            {'client_external_id': 'CLI-003', 'product': 'VSAV', 'submitted_on': '2024-02-01',
             'approved_on': '2024-02-01', 'activated_on': '2024-02-01', 'initial_deposit': 15000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-004'},

            {'client_external_id': 'CLI-004', 'product': 'VSAV', 'submitted_on': '2024-02-05',
             'approved_on': '2024-02-05', 'activated_on': '2024-02-05', 'initial_deposit': 20000,
             'field_officer': 'loan.bafoussam', 'external_id': 'SAV-005'},

            {'client_external_id': 'CLI-005', 'product': 'VSAV', 'submitted_on': '2024-02-10',
             'approved_on': '2024-02-10', 'activated_on': '2024-02-10', 'initial_deposit': 18000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-006'},

            {'client_external_id': 'CLI-006', 'product': 'VSAV', 'submitted_on': '2024-02-15',
             'approved_on': '2024-02-15', 'activated_on': '2024-02-15', 'initial_deposit': 75000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-007'},

            {'client_external_id': 'CLI-006', 'product': 'FDEP', 'submitted_on': '2024-02-20',
             'approved_on': '2024-02-20', 'activated_on': '2024-02-20', 'initial_deposit': 200000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-008'},

            {'client_external_id': 'CLI-007', 'product': 'VSAV', 'submitted_on': '2024-02-20',
             'approved_on': '2024-02-20', 'activated_on': '2024-02-20', 'initial_deposit': 12000,
             'field_officer': 'loan.bamenda', 'external_id': 'SAV-009'},

            {'client_external_id': 'CLI-008', 'product': 'VSAV', 'submitted_on': '2024-03-01',
             'approved_on': '2024-03-01', 'activated_on': '2024-03-01', 'initial_deposit': 16000,
             'field_officer': 'loan.douala', 'external_id': 'SAV-010'},

            {'client_external_id': 'CLI-009', 'product': 'VSAV', 'submitted_on': '2024-03-05',
             'approved_on': '2024-03-05', 'activated_on': '2024-03-05', 'initial_deposit': 10000,
             'field_officer': 'loan.bafoussam', 'external_id': 'SAV-011'},

            {'client_external_id': 'CLI-010', 'product': 'VSAV', 'submitted_on': '2024-03-10',
             'approved_on': '2024-03-10', 'activated_on': '2024-03-10', 'initial_deposit': 14000,
             'field_officer': 'loan.yaounde', 'external_id': 'SAV-012'},
        ]
        return pd.DataFrame(data)

    def create_fund_sources_sheet(self):
        """Create fund sources"""
        data = [
            {'fund_name': 'Own Capital', 'external_id': 'FUND-001', 'description': 'Institutional own capital'},
            {'fund_name': 'Bank Loan - Ecobank', 'external_id': 'FUND-002', 'description': 'Credit line from Ecobank'},
            {'fund_name': 'Bank Loan - BICEC', 'external_id': 'FUND-003', 'description': 'Credit line from BICEC'},
            {'fund_name': 'Donor - AfDB', 'external_id': 'FUND-004', 'description': 'African Development Bank grant'},
            {'fund_name': 'Donor - GIZ', 'external_id': 'FUND-005', 'description': 'German cooperation grant'},
            {'fund_name': 'Donor - World Bank', 'external_id': 'FUND-006', 'description': 'World Bank project funding'},
        ]
        return pd.DataFrame(data)

    def create_payment_types_sheet(self):
        """Create payment channels"""
        data = [
            {'payment_type': 'Cash', 'description': 'Cash payment', 'is_cash_payment': 'Yes', 'order_position': 1},
            {'payment_type': 'MTN Mobile Money', 'description': 'MTN MoMo payment', 'is_cash_payment': 'No', 'order_position': 2},
            {'payment_type': 'Orange Money', 'description': 'Orange Money payment', 'is_cash_payment': 'No', 'order_position': 3},
            {'payment_type': 'Bank Transfer', 'description': 'Bank transfer payment', 'is_cash_payment': 'No', 'order_position': 4},
            {'payment_type': 'Cheque', 'description': 'Cheque payment', 'is_cash_payment': 'No', 'order_position': 5},
        ]
        return pd.DataFrame(data)

    def create_holidays_sheet(self):
        """Create Cameroon public holidays"""
        data = [
            {'holiday_name': "New Year's Day", 'date': '2024-01-01', 'rescheduled_to': '2024-01-02', 'description': 'New Year Holiday'},
            {'holiday_name': 'Youth Day', 'date': '2024-02-11', 'rescheduled_to': '2024-02-12', 'description': 'Youth Day Holiday'},
            {'holiday_name': 'Labour Day', 'date': '2024-05-01', 'rescheduled_to': '2024-05-02', 'description': 'International Labour Day'},
            {'holiday_name': 'National Day', 'date': '2024-05-20', 'rescheduled_to': '2024-05-21', 'description': 'Cameroon National Day'},
            {'holiday_name': 'Assumption Day', 'date': '2024-08-15', 'rescheduled_to': '2024-08-16', 'description': 'Assumption of Mary'},
            {'holiday_name': 'Christmas Day', 'date': '2024-12-25', 'rescheduled_to': '2024-12-26', 'description': 'Christmas Holiday'},
            {'holiday_name': "New Year's Day", 'date': '2025-01-01', 'rescheduled_to': '2025-01-02', 'description': 'New Year Holiday'},
            {'holiday_name': 'Youth Day', 'date': '2025-02-11', 'rescheduled_to': '2025-02-12', 'description': 'Youth Day Holiday'},
            {'holiday_name': 'Labour Day', 'date': '2025-05-01', 'rescheduled_to': '2025-05-02', 'description': 'International Labour Day'},
            {'holiday_name': 'National Day', 'date': '2025-05-20', 'rescheduled_to': '2025-05-21', 'description': 'Cameroon National Day'},
            {'holiday_name': 'Assumption Day', 'date': '2025-08-15', 'rescheduled_to': '2025-08-18', 'description': 'Assumption of Mary'},
            {'holiday_name': 'Christmas Day', 'date': '2025-12-25', 'rescheduled_to': '2025-12-26', 'description': 'Christmas Holiday'},
        ]
        return pd.DataFrame(data)

    def create_loan_provisioning_sheet(self):
        """Create COBAC loan provisioning criteria"""
        data = [
            {'category_name': 'Performing', 'min_days_overdue': 0, 'max_days_overdue': 0, 'provision_percentage': 0},
            {'category_name': 'Watch', 'min_days_overdue': 1, 'max_days_overdue': 30, 'provision_percentage': 25},
            {'category_name': 'Substandard', 'min_days_overdue': 31, 'max_days_overdue': 90, 'provision_percentage': 50},
            {'category_name': 'Doubtful', 'min_days_overdue': 91, 'max_days_overdue': 180, 'provision_percentage': 75},
            {'category_name': 'Loss', 'min_days_overdue': 181, 'max_days_overdue': 9999, 'provision_percentage': 100},
        ]
        return pd.DataFrame(data)

    def create_collateral_types_sheet(self):
        """Create collateral types"""
        data = [
            {'collateral_type': 'Land Title', 'description': 'Land ownership document', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Vehicle Registration', 'description': 'Vehicle ownership papers', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Household Goods', 'description': 'Furniture, appliances', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Shop/Business Equipment', 'description': 'Business machinery and equipment', 'requires_valuation': 'Yes'},
            {'collateral_type': 'Savings Account Lien', 'description': 'Lien on savings account', 'requires_valuation': 'No'},
            {'collateral_type': 'Group Guarantee', 'description': 'Solidarity group guarantee', 'requires_valuation': 'No'},
        ]
        return pd.DataFrame(data)

    def create_guarantor_types_sheet(self):
        """Create guarantor types"""
        data = [
            {'guarantor_type': 'Spouse', 'description': 'Client spouse as guarantor'},
            {'guarantor_type': 'Family Member', 'description': 'Family member (parent, sibling, child)'},
            {'guarantor_type': 'Employer', 'description': 'Employer guarantee for salaried clients'},
            {'guarantor_type': 'Business Partner', 'description': 'Business partner or associate'},
            {'guarantor_type': 'Group Member', 'description': 'Fellow group member in solidarity lending'},
            {'guarantor_type': 'Community Leader', 'description': 'Village chief or community leader'},
        ]
        return pd.DataFrame(data)

    def create_config_sheet(self):
        """Create system configuration parameters"""
        data = [
            {'config_key': 'INSTITUTION_NAME', 'config_value': 'Cameroon Microfinance Institution', 'category': 'General'},
            {'config_key': 'INSTITUTION_EXTERNAL_ID', 'config_value': 'CMR-MFI-001', 'category': 'General'},
            {'config_key': 'BASE_CURRENCY', 'config_value': 'XAF', 'category': 'General'},
            {'config_key': 'WORKING_DAYS', 'config_value': 'Monday,Tuesday,Wednesday,Thursday,Friday', 'category': 'General'},
            {'config_key': 'ACCOUNTING_RULE', 'config_value': 'Cash Based', 'category': 'Accounting'},
            {'config_key': 'FINANCIAL_YEAR_START', 'config_value': '01-January', 'category': 'Accounting'},
            {'config_key': 'DEFAULT_PASSWORD', 'config_value': 'password', 'category': 'Security'},
            {'config_key': 'PASSWORD_CHANGE_REQUIRED', 'config_value': 'Yes', 'category': 'Security'},
            {'config_key': 'SMS_PROVIDER', 'config_value': 'Twilio', 'category': 'Communication'},
            {'config_key': 'SMS_ENABLED', 'config_value': 'No', 'category': 'Communication'},
            {'config_key': 'EMAIL_ENABLED', 'config_value': 'No', 'category': 'Communication'},
            {'config_key': 'MTN_MOMO_ENABLED', 'config_value': 'No', 'category': 'Payment'},
            {'config_key': 'ORANGE_MONEY_ENABLED', 'config_value': 'No', 'category': 'Payment'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_CASHIER', 'config_value': '500000', 'category': 'Approval'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_MANAGER', 'config_value': '5000000', 'category': 'Approval'},
            {'config_key': 'LOAN_APPROVAL_LIMIT_DIRECTOR', 'config_value': 'unlimited', 'category': 'Approval'},
            {'config_key': 'MAX_LOAN_RESCHEDULE', 'config_value': '2', 'category': 'Loan'},
            {'config_key': 'WRITEOFF_AFTER_DAYS', 'config_value': '365', 'category': 'Loan'},
            {'config_key': 'WHT_RATE_SAVINGS_INTEREST', 'config_value': '15', 'category': 'Tax'},
        ]
        return pd.DataFrame(data)

    def create_loan_product_accounting_sheet(self):
        """Create loan product GL account mappings"""
        data = []

        # For each loan product, define all 13 GL mappings
        products = ['MSOL', 'SBIZ', 'ASEA']

        for product in products:
            mappings = [
                {'product_short_name': product, 'mapping_type': 'Fund Source', 'gl_code': '42',
                 'gl_name': 'Cash on Hand', 'description': 'Source of funds for disbursement'},

                {'product_short_name': product, 'mapping_type': 'Loan Portfolio', 'gl_code': '51',
                 'gl_name': 'Loans to Clients - Principal', 'description': 'Outstanding loan principal'},

                {'product_short_name': product, 'mapping_type': 'Interest Receivable', 'gl_code': '52',
                 'gl_name': 'Loans to Clients - Interest Receivable', 'description': 'Accrued interest receivable'},

                {'product_short_name': product, 'mapping_type': 'Fees Receivable', 'gl_code': '53',
                 'gl_name': 'Loans to Clients - Fees Receivable', 'description': 'Accrued fees receivable'},

                {'product_short_name': product, 'mapping_type': 'Penalties Receivable', 'gl_code': '54',
                 'gl_name': 'Loans to Clients - Penalties Receivable', 'description': 'Penalties receivable'},

                {'product_short_name': product, 'mapping_type': 'Transfer in Suspense', 'gl_code': '42',
                 'gl_name': 'Cash on Hand', 'description': 'Suspense account for transfers'},

                {'product_short_name': product, 'mapping_type': 'Interest Income', 'gl_code': '81',
                 'gl_name': 'Interest Income on Loans', 'description': 'Interest income recognition'},

                {'product_short_name': product, 'mapping_type': 'Fee Income', 'gl_code': '82',
                 'gl_name': 'Fee Income - Loan Processing', 'description': 'Fee income recognition'},

                {'product_short_name': product, 'mapping_type': 'Penalty Income', 'gl_code': '83',
                 'gl_name': 'Penalty Income', 'description': 'Penalty income recognition'},

                {'product_short_name': product, 'mapping_type': 'Losses Written Off', 'gl_code': '93',
                 'gl_name': 'Loan Write-off Expense', 'description': 'Written-off loan expense'},

                {'product_short_name': product, 'mapping_type': 'Goodwill Credit', 'gl_code': '81',
                 'gl_name': 'Interest Income on Loans', 'description': 'Goodwill credit account'},

                {'product_short_name': product, 'mapping_type': 'Income from Recovery', 'gl_code': '81',
                 'gl_name': 'Interest Income on Loans', 'description': 'Recovery income'},

                {'product_short_name': product, 'mapping_type': 'Over Payment Liability', 'gl_code': '64',
                 'gl_name': 'Savings Interest Payable', 'description': 'Overpayment liability account'},
            ]
            data.extend(mappings)

        return pd.DataFrame(data)

    def create_savings_product_accounting_sheet(self):
        """Create savings product GL account mappings"""
        data = []

        # For each savings product, define all 11 GL mappings
        savings_products = [
            {'short_name': 'VSAV', 'liability_gl': '61'},
            {'short_name': 'FDEP', 'liability_gl': '62'},
            {'short_name': 'MGRP', 'liability_gl': '63'}
        ]

        for product in savings_products:
            mappings = [
                {'product_short_name': product['short_name'], 'mapping_type': 'Savings Reference',
                 'gl_code': '42', 'gl_name': 'Cash on Hand',
                 'description': 'Asset account - where customer deposits are held'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Savings Control',
                 'gl_code': product['liability_gl'], 'gl_name': f"GL Code {product['liability_gl']}",
                 'description': 'Liability account - obligation to customers'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Interest on Savings',
                 'gl_code': '91', 'gl_name': 'Interest Expense on Savings', 'description': 'Interest expense'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Fees',
                 'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'description': 'Fee income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Penalties',
                 'gl_code': '84', 'gl_name': 'Fee Income - Savings Accounts', 'description': 'Penalty income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Overdraft Portfolio Control',
                 'gl_code': '51', 'gl_name': 'Loans to Clients - Principal', 'description': 'Overdraft control'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Income from Interest',
                 'gl_code': '81', 'gl_name': 'Interest Income on Loans', 'description': 'Overdraft interest income'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Losses Written Off',
                 'gl_code': '93', 'gl_name': 'Loan Write-off Expense', 'description': 'Overdraft write-off'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Escheat Liability',
                 'gl_code': product['liability_gl'], 'gl_name': f"GL Code {product['liability_gl']}",
                 'description': 'Escheat liability'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Withholding Tax',
                 'gl_code': '141', 'gl_name': 'Tax Payable - WHT', 'description': 'Withholding tax payable'},

                {'product_short_name': product['short_name'], 'mapping_type': 'Transfer in Suspense',
                 'gl_code': '42', 'gl_name': 'Cash on Hand', 'description': 'Transfer suspense account'},
            ]
            data.extend(mappings)

        return pd.DataFrame(data)

    def create_payment_type_accounting_sheet(self):
        """Create payment type GL account mappings"""
        data = [
            {'payment_type': 'Cash', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'fund_source': 'Yes', 'description': 'Physical cash transactions'},

            {'payment_type': 'MTN Mobile Money', 'gl_code': '43', 'gl_name': 'MTN Mobile Money',
             'fund_source': 'Yes', 'description': 'MTN MoMo wallet'},

            {'payment_type': 'Orange Money', 'gl_code': '44', 'gl_name': 'Orange Money',
             'fund_source': 'Yes', 'description': 'Orange Money wallet'},

            {'payment_type': 'Bank Transfer', 'gl_code': '41', 'gl_name': 'Banks - Current Accounts',
             'fund_source': 'Yes', 'description': 'Bank account transfers'},

            {'payment_type': 'Cheque', 'gl_code': '41', 'gl_name': 'Banks - Current Accounts',
             'fund_source': 'Yes', 'description': 'Cheque payments'},
        ]
        return pd.DataFrame(data)

    def create_financial_activity_mappings_sheet(self):
        """Create financial activity to GL account mappings"""
        data = [
            # Asset Transfer - CRITICAL for inter-branch operations
            {'financial_activity': 'Asset Transfer', 'gl_code': '122', 'gl_name': 'Due from Other Branches (Receivable)',
             'description': 'Inter-office asset transfers - receiving branch records receivable'},

            # Liability Transfer - CRITICAL for inter-branch operations
            {'financial_activity': 'Liability Transfer', 'gl_code': '131', 'gl_name': 'Due to Other Branches (Payable)',
             'description': 'Inter-office liability transfers - sending branch records payable'},

            # Cash at Mainvault
            {'financial_activity': 'Cash at Mainvault', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Main vault cash - head office cash reserves'},

            # Cash at Teller
            {'financial_activity': 'Cash at Teller', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Teller cash - cash in teller drawers'},

            # Opening Balances Contra
            {'financial_activity': 'Opening Balances Contra', 'gl_code': '71', 'gl_name': 'Share Capital',
             'description': 'Opening balance contra account - for initial data migration'},

            # Fund Source
            {'financial_activity': 'Fund Source', 'gl_code': '42', 'gl_name': 'Cash on Hand',
             'description': 'Default fund source for loan disbursements'},
        ]
        return pd.DataFrame(data)

    def create_teller_cashier_mappings_sheet(self):
        """Create teller/cashier GL account mappings per office"""
        data = [
            # Yaounde Branch
            {'office_name': 'Yaounde Branch', 'teller_name': 'Teller - Yaounde',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Yaounde branch teller and vault configuration'},

            # Douala Branch
            {'office_name': 'Douala Branch', 'teller_name': 'Teller - Douala',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Douala branch teller and vault configuration'},

            # Bafoussam Branch
            {'office_name': 'Bafoussam Branch', 'teller_name': 'Teller - Bafoussam',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Bafoussam branch teller and vault configuration'},

            # Bamenda Branch
            {'office_name': 'Bamenda Branch', 'teller_name': 'Teller - Bamenda',
             'cash_gl_code': '42', 'vault_gl_code': '42',
             'shortage_gl_code': '98', 'overage_gl_code': '86',
             'description': 'Bamenda branch teller and vault configuration'},
        ]
        return pd.DataFrame(data)

    def create_tellers_sheet(self):
        """Create tellers (cash counters) for each branch"""
        data = [
            # Yaounde Branch
            {'office_name': 'Yaounde Branch', 'teller_name': 'Teller 1 - Yaounde',
             'description': 'Main teller counter at Yaounde branch',
             'start_date': '2024-01-15', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Douala Branch
            {'office_name': 'Douala Branch', 'teller_name': 'Teller 1 - Douala',
             'description': 'Main teller counter at Douala branch',
             'start_date': '2024-01-15', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Bafoussam Branch
            {'office_name': 'Bafoussam Branch', 'teller_name': 'Teller 1 - Bafoussam',
             'description': 'Main teller counter at Bafoussam branch',
             'start_date': '2024-02-01', 'end_date': '2030-12-31',
             'status': 'Active'},

            # Bamenda Branch
            {'office_name': 'Bamenda Branch', 'teller_name': 'Teller 1 - Bamenda',
             'description': 'Main teller counter at Bamenda branch',
             'start_date': '2024-02-15', 'end_date': '2030-12-31',
             'status': 'Active'},
        ]
        return pd.DataFrame(data)

    def create_roles_permissions_sheet(self):
        """Create roles with permissions assignments"""
        data = [
            # Branch Manager Role - Full branch operations access
            {'role_name': 'Branch Manager', 'permission_group': 'Client', 'permission': 'ALL_FUNCTIONS',
             'description': 'Full client management'},
            {'role_name': 'Branch Manager', 'permission_group': 'Loan', 'permission': 'ALL_FUNCTIONS',
             'description': 'Full loan management'},
            {'role_name': 'Branch Manager', 'permission_group': 'Savings', 'permission': 'ALL_FUNCTIONS',
             'description': 'Full savings management'},
            {'role_name': 'Branch Manager', 'permission_group': 'Transaction', 'permission': 'ALL_FUNCTIONS',
             'description': 'All transaction operations'},
            {'role_name': 'Branch Manager', 'permission_group': 'Report', 'permission': 'READ',
             'description': 'View all reports'},
            {'role_name': 'Branch Manager', 'permission_group': 'Office', 'permission': 'READ',
             'description': 'View office information'},
            {'role_name': 'Branch Manager', 'permission_group': 'Staff', 'permission': 'READ',
             'description': 'View staff information'},
            {'role_name': 'Branch Manager', 'permission_group': 'Accounting', 'permission': 'READ',
             'description': 'View accounting entries'},
            {'role_name': 'Branch Manager', 'permission_group': 'Maker-Checker', 'permission': 'APPROVE',
             'description': 'Can approve maker-checker tasks (checker role)'},

            # Loan Officer Role - Loan and client operations
            {'role_name': 'Loan Officer', 'permission_group': 'Client', 'permission': 'CREATE_UPDATE_READ',
             'description': 'Create and update clients'},
            {'role_name': 'Loan Officer', 'permission_group': 'Loan', 'permission': 'CREATE_UPDATE_READ',
             'description': 'Create and manage loans (approval requires manager)'},
            {'role_name': 'Loan Officer', 'permission_group': 'Savings', 'permission': 'CREATE_UPDATE_READ',
             'description': 'Create and manage savings accounts'},
            {'role_name': 'Loan Officer', 'permission_group': 'Transaction', 'permission': 'CREATE_READ',
             'description': 'Process transactions (deposits, repayments)'},
            {'role_name': 'Loan Officer', 'permission_group': 'Report', 'permission': 'READ',
             'description': 'View operational reports'},
            {'role_name': 'Loan Officer', 'permission_group': 'Office', 'permission': 'READ',
             'description': 'View office information'},
            {'role_name': 'Loan Officer', 'permission_group': 'Maker-Checker', 'permission': 'CREATE',
             'description': 'Can create maker-checker tasks (maker role)'},

            # Cashier Role - Transaction processing only
            {'role_name': 'Cashier', 'permission_group': 'Client', 'permission': 'READ',
             'description': 'View client information only'},
            {'role_name': 'Cashier', 'permission_group': 'Loan', 'permission': 'READ',
             'description': 'View loan information only'},
            {'role_name': 'Cashier', 'permission_group': 'Savings', 'permission': 'READ',
             'description': 'View savings information only'},
            {'role_name': 'Cashier', 'permission_group': 'Transaction', 'permission': 'CREATE_READ',
             'description': 'Process deposits, withdrawals, repayments'},
            {'role_name': 'Cashier', 'permission_group': 'Teller', 'permission': 'ALL_FUNCTIONS',
             'description': 'Full teller/cashier operations'},
            {'role_name': 'Cashier', 'permission_group': 'Report', 'permission': 'READ',
             'description': 'View transaction reports'},
            {'role_name': 'Cashier', 'permission_group': 'Maker-Checker', 'permission': 'CREATE',
             'description': 'Can create maker-checker tasks (maker role)'},
        ]
        return pd.DataFrame(data)

    def create_maker_checker_config_sheet(self):
        """Create Maker-Checker (4-eyes principle) configuration for critical operations

        All permission codes verified against Fineract source code (@CommandType annotations).
        """
        data = [
            # Loan operations
            {'task_name': 'Loan Approval', 'entity': 'Loan', 'action': 'APPROVE',
             'threshold_amount': 2000000, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Loans above 2M XAF require manager approval'},

            {'task_name': 'Loan Disbursement', 'entity': 'Loan', 'action': 'DISBURSE',
             'threshold_amount': 5000000, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Disbursements above 5M XAF require manager approval'},

            {'task_name': 'Loan Write-off', 'entity': 'Loan', 'action': 'WRITEOFF',
             'threshold_amount': 0, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'All loan write-offs require manager approval (any amount)'},

            {'task_name': 'Create Loan Reschedule Request', 'entity': 'RESCHEDULELOAN', 'action': 'CREATE',
             'threshold_amount': 1000000, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Loan rescheduling above 1M XAF requires manager approval'},

            # Savings operations
            {'task_name': 'Close Savings Account', 'entity': 'SAVINGSACCOUNT', 'action': 'CLOSE',
             'threshold_amount': 500000, 'threshold_currency': 'XAF',
             'maker_role': 'Cashier', 'checker_role': 'Branch Manager',
             'description': 'Account closures with balance above 500K XAF require manager approval'},

            # Client operations
            {'task_name': 'Client Activation', 'entity': 'Client', 'action': 'ACTIVATE',
             'threshold_amount': 0, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'New client activation requires manager approval'},

            {'task_name': 'Propose Client Transfer', 'entity': 'Client', 'action': 'PROPOSETRANSFER',
             'threshold_amount': 0, 'threshold_currency': 'XAF',
             'maker_role': 'Loan Officer', 'checker_role': 'Branch Manager',
             'description': 'Client office transfer requires manager approval'},

            # Accounting operations
            {'task_name': 'Create Manual Journal Entry', 'entity': 'JOURNALENTRY', 'action': 'CREATE',
             'threshold_amount': 500000, 'threshold_currency': 'XAF',
             'maker_role': 'Accountant', 'checker_role': 'Branch Manager',
             'description': 'Manual journal entries above 500K XAF require manager approval'},

            # User/System operations
            {'task_name': 'Create User', 'entity': 'User', 'action': 'CREATE',
             'threshold_amount': 0, 'threshold_currency': 'XAF',
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Creating new system users requires head office approval'},

            {'task_name': 'Update User', 'entity': 'User', 'action': 'UPDATE',
             'threshold_amount': 0, 'threshold_currency': 'XAF',
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Updating user details requires head office approval'},

            # Office/Fund transfers
            {'task_name': 'Create Office Transaction', 'entity': 'OFFICETRANSACTION', 'action': 'CREATE',
             'threshold_amount': 3000000, 'threshold_currency': 'XAF',
             'maker_role': 'Branch Manager', 'checker_role': 'Head Office Manager',
             'description': 'Inter-branch fund transfers above 3M XAF require head office approval'},
        ]
        return pd.DataFrame(data)

    def create_currency_config_sheet(self):
        """Create currency configuration"""
        data = [
            {'currency_code': 'XAF', 'currency_name': 'Central African CFA Franc', 'decimal_places': 0,
             'in_multiples_of': 1, 'display_symbol': 'FCFA', 'name_code': 'currency.XAF',
             'description': 'Official currency of CEMAC countries including Cameroon'},
        ]
        return pd.DataFrame(data)

    def create_working_days_sheet(self):
        """Create working days configuration"""
        data = [
            {'day_of_week': 'Monday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Tuesday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Wednesday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Thursday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Friday', 'working_day': 'Yes', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Saturday', 'working_day': 'No', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
            {'day_of_week': 'Sunday', 'working_day': 'No', 'recurrence_type': 'Weekly', 'recurrence_frequency': 1},
        ]
        return pd.DataFrame(data)

    def create_account_number_preferences_sheet(self):
        """Create account number format preferences"""
        data = [
            {'entity_type': 'Client', 'prefix_type': 'Office Short Name', 'account_number_length': 9,
             'example': 'DLA000001', 'description': 'Client account numbers: Office code + 6 digits'},

            {'entity_type': 'Loan', 'prefix_type': 'Product Short Name', 'account_number_length': 12,
             'example': 'MICRO000001', 'description': 'Loan account numbers: Product code + 6 digits'},

            {'entity_type': 'Savings', 'prefix_type': 'Product Short Name', 'account_number_length': 12,
             'example': 'VOL00000001', 'description': 'Savings account numbers: Product code + 8 digits'},

            {'entity_type': 'Groups', 'prefix_type': 'Office Short Name', 'account_number_length': 10,
             'example': 'GRP-DLA001', 'description': 'Group account numbers: GRP-Office code + 3 digits'},

            {'entity_type': 'Centers', 'prefix_type': 'Office Short Name', 'account_number_length': 10,
             'example': 'CTR-DLA001', 'description': 'Center account numbers: CTR-Office code + 3 digits'},
        ]
        return pd.DataFrame(data)

    def create_codes_and_values_sheet(self):
        """Create codes and code values for dropdowns"""
        data = [
            # Gender
            {'code_name': 'Gender', 'code_value': 'Male', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Male gender'},
            {'code_name': 'Gender', 'code_value': 'Female', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Female gender'},
            {'code_name': 'Gender', 'code_value': 'Other', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Other gender'},

            # Client Type
            {'code_name': 'ClientType', 'code_value': 'Individual', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Individual client'},
            {'code_name': 'ClientType', 'code_value': 'Corporate', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Corporate/business client'},

            # Client Classification
            {'code_name': 'ClientClassification', 'code_value': 'Active', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Active client'},
            {'code_name': 'ClientClassification', 'code_value': 'Pending', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Pending activation'},
            {'code_name': 'ClientClassification', 'code_value': 'Closed', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Closed/inactive client'},

            # Marital Status
            {'code_name': 'MaritalStatus', 'code_value': 'Single', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Single'},
            {'code_name': 'MaritalStatus', 'code_value': 'Married', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Married'},
            {'code_name': 'MaritalStatus', 'code_value': 'Divorced', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Divorced'},
            {'code_name': 'MaritalStatus', 'code_value': 'Widow', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Widow/Widower'},

            # Education Level
            {'code_name': 'EducationLevel', 'code_value': 'Primary', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Primary education'},
            {'code_name': 'EducationLevel', 'code_value': 'Secondary', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Secondary education'},
            {'code_name': 'EducationLevel', 'code_value': 'University', 'code_position': 3, 'is_active': 'Yes',
             'description': 'University degree'},
            {'code_name': 'EducationLevel', 'code_value': 'No Formal Education', 'code_position': 4, 'is_active': 'Yes',
             'description': 'No formal education'},

            # Loan Purpose
            {'code_name': 'LoanPurpose', 'code_value': 'Working Capital', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Business working capital'},
            {'code_name': 'LoanPurpose', 'code_value': 'Equipment Purchase', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Purchase of equipment'},
            {'code_name': 'LoanPurpose', 'code_value': 'Inventory', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Purchase inventory/stock'},
            {'code_name': 'LoanPurpose', 'code_value': 'Agricultural Input', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Seeds, fertilizer, tools'},
            {'code_name': 'LoanPurpose', 'code_value': 'Education', 'code_position': 5, 'is_active': 'Yes',
             'description': 'School fees and education'},
            {'code_name': 'LoanPurpose', 'code_value': 'Emergency', 'code_position': 6, 'is_active': 'Yes',
             'description': 'Emergency/medical expenses'},

            # Account Closure Reason
            {'code_name': 'SavingsClosureReason', 'code_value': 'Client Request', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Client requested closure'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Dormant Account', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Account became dormant'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Moved to Another Branch', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Client relocated'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Deceased', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Client deceased'},
            {'code_name': 'SavingsClosureReason', 'code_value': 'Fraudulent Activity', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Account involved in fraud'},

            # Loan Closure Reason
            {'code_name': 'LoanClosureReason', 'code_value': 'Fully Paid', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Loan fully repaid'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Written Off', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Loan written off as bad debt'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Rescheduled', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Loan rescheduled to new terms'},
            {'code_name': 'LoanClosureReason', 'code_value': 'Foreclosed', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Collateral foreclosed'},

            # Business Type
            {'code_name': 'BusinessType', 'code_value': 'Retail Trade', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Retail business'},
            {'code_name': 'BusinessType', 'code_value': 'Wholesale Trade', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Wholesale business'},
            {'code_name': 'BusinessType', 'code_value': 'Agriculture', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Farming/agriculture'},
            {'code_name': 'BusinessType', 'code_value': 'Services', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Service provision'},
            {'code_name': 'BusinessType', 'code_value': 'Manufacturing', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Manufacturing/production'},
            {'code_name': 'BusinessType', 'code_value': 'Transportation', 'code_position': 6, 'is_active': 'Yes',
             'description': 'Transport services'},
            {'code_name': 'BusinessType', 'code_value': 'Food Services', 'code_position': 7, 'is_active': 'Yes',
             'description': 'Restaurant/catering'},

            # Guarantor Relationship (Note: Spouse already exists in Fineract by default)
            {'code_name': 'GuarantorRelationship', 'code_value': 'Family Member', 'code_position': 1, 'is_active': 'Yes',
             'description': 'Family member (parent, sibling, child)'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Employer', 'code_position': 2, 'is_active': 'Yes',
             'description': 'Employer guarantee for salaried clients'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Business Partner', 'code_position': 3, 'is_active': 'Yes',
             'description': 'Business partner or associate'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Group Member', 'code_position': 4, 'is_active': 'Yes',
             'description': 'Fellow group member in solidarity lending'},
            {'code_name': 'GuarantorRelationship', 'code_value': 'Community Leader', 'code_position': 5, 'is_active': 'Yes',
             'description': 'Village chief or community leader'},
        ]
        return pd.DataFrame(data)

    def create_scheduler_jobs_sheet(self):
        """Create scheduler job configurations

        All job names verified against Fineract source code (JobName.java enum).
        """
        data = [
            # Core savings and loan jobs
            {'job_name': 'Post Interest For Savings', 'display_name': 'Post Interest For Savings',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Post interest to savings accounts daily at midnight'},

            {'job_name': 'Transfer Fee For Loans From Savings', 'display_name': 'Transfer Fee For Loans From Savings',
             'cron_expression': '0 0 1 * * ?', 'active': 'Yes',
             'description': 'Auto-deduct loan fees from linked savings accounts daily at 1 AM'},

            {'job_name': 'Update Loan Arrears Ageing', 'display_name': 'Update Loan Arrears Ageing',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Update loan arrears aging daily'},

            {'job_name': 'Add Accrual Transactions', 'display_name': 'Add Accrual Transactions',
             'cron_expression': '0 0 2 * * ?', 'active': 'Yes',
             'description': 'Add accrual accounting entries daily at 2 AM'},

            {'job_name': 'Apply Annual Fee For Savings', 'display_name': 'Apply Annual Fee For Savings',
             'cron_expression': '0 0 0 1 1 ? *', 'active': 'Yes',
             'description': 'Apply annual savings account fees on January 1st'},

            # Penalty and charges
            {'job_name': 'Apply penalty to overdue loans', 'display_name': 'Apply penalty to overdue loans',
             'cron_expression': '0 0 3 * * ?', 'active': 'Yes',
             'description': 'Apply penalty charges to overdue loans daily at 3 AM'},

            # Savings dormancy
            {'job_name': 'Update Savings Dormant Accounts', 'display_name': 'Update Savings Dormant Accounts',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Mark savings accounts as inactive/dormant based on activity'},

            # Standing instructions (recurring payments)
            {'job_name': 'Execute Standing Instruction', 'display_name': 'Execute Standing Instruction',
             'cron_expression': '0 0 0 1/1 * ? *', 'active': 'Yes',
             'description': 'Execute standing instructions (recurring payments) daily'},

            # Periodic accruals
            {'job_name': 'Add Periodic Accrual Transactions', 'display_name': 'Add Periodic Accrual Transactions',
             'cron_expression': '0 0 0 1 * ? *', 'active': 'Yes',
             'description': 'Add monthly accrual entries on the 1st of each month'},

            # NOTE: "Update Loan Paid In Advance" job does not exist in Fineract source code
            # This functionality may have been removed or renamed in newer versions
        ]
        return pd.DataFrame(data)

    def create_global_config_sheet(self):
        """Create comprehensive global configuration settings"""
        data = [
            # Maker-Checker Global Settings
            {'config_name': 'maker-checker', 'enabled': 'Yes', 'value': 'true',
             'description': 'Enable maker-checker globally'},

            # Reschedule Future Repayments
            {'config_name': 'reschedule-future-repayments', 'enabled': 'Yes', 'value': 'true',
             'description': 'Allow rescheduling of future repayments'},

            # Allow Backdated Transaction
            {'config_name': 'allow-backdated-transaction-before-interest-posting', 'enabled': 'No', 'value': 'false',
             'description': 'Prevent backdated transactions before interest posting'},

            # Allow Transactions on Non-Working Days
            {'config_name': 'allow-transactions-on-non_working_day', 'enabled': 'No', 'value': 'false',
             'description': 'Block transactions on holidays and weekends'},

            # Allow Transactions on Holidays
            {'config_name': 'allow-transactions-on-holiday', 'enabled': 'No', 'value': 'false',
             'description': 'Block transactions on public holidays'},

            # Savings Interest Posting
            {'config_name': 'savings-interest-posting-current-period-end', 'enabled': 'Yes', 'value': 'per_period',
             'description': 'Post savings interest at period end'},

            # Financial Year Beginning
            {'config_name': 'financial-year-beginning-month', 'enabled': 'Yes', 'value': '1',
             'description': 'Financial year starts in January'},

            # Min/Max days between repayments
            {'config_name': 'min-days-between-disbursal-and-first-repayment', 'enabled': 'Yes', 'value': '1',
             'description': 'Minimum days between disbursement and first repayment'},

            # Grace on arrears ageing
            {'config_name': 'grace-on-arrears-ageing', 'enabled': 'Yes', 'value': '0',
             'description': 'Grace period for arrears aging calculation (days)'},

            # Allow Dividend Calculation for Inactive Clients
            {'config_name': 'allow-dividend-calculation-for-inactive-clients', 'enabled': 'No', 'value': 'false',
             'description': 'Exclude inactive clients from dividend calculation'},

            # Enforce Minimum Required Balance
            {'config_name': 'enforce-min-required-balance', 'enabled': 'Yes', 'value': 'true',
             'description': 'Enforce minimum balance requirement for savings accounts'},

            # Allow Withdrawal Fee
            {'config_name': 'allow-withdrawals-on-savings-account-withhold-tax', 'enabled': 'No', 'value': 'false',
             'description': 'Block withdrawals if withholding tax unpaid'},

            # Interest Calculation Using Daily Balance
            {'config_name': 'interest-calculation-using-daily-balance', 'enabled': 'Yes', 'value': 'true',
             'description': 'Calculate interest using daily balance'},

            # Accounting Rule
            {'config_name': 'accounting-rule', 'enabled': 'Yes', 'value': '2',
             'description': 'Accounting rule: 1=None, 2=Cash, 3=Accrual Periodic, 4=Accrual Upfront'},

            # Days in Year
            {'config_name': 'days-in-year-type', 'enabled': 'Yes', 'value': '365',
             'description': 'Days in year for interest calculations: 360 or 365'},

            # Days in Month
            {'config_name': 'days-in-month-type', 'enabled': 'Yes', 'value': '30',
             'description': 'Days in month for interest calculations: 30 or Actual'},

            # Charge Accrual Date
            {'config_name': 'charge-accrual-date', 'enabled': 'Yes', 'value': 'due-date',
             'description': 'When to accrue charges: due-date or submitted-date'},

            # Enable Business Date
            {'config_name': 'enable_business_date', 'enabled': 'Yes', 'value': 'false',
             'description': 'Enable business date (COB) feature'},

            # Enable Auto Repayment for Down-payment
            {'config_name': 'enable-auto-generated-external-id', 'enabled': 'Yes', 'value': 'true',
             'description': 'Auto-generate external IDs if not provided'},

            # Meeting Calendar
            {'config_name': 'meetings-mandatory-for-jlg-loans', 'enabled': 'No', 'value': 'false',
             'description': 'Make meetings mandatory for JLG (Joint Liability Group) loans'},

            # Minimum Age for Clients
            {'config_name': 'minimum-age-for-client-activation', 'enabled': 'Yes', 'value': '18',
             'description': 'Minimum age required for client activation'},

            # Loan COB
            {'config_name': 'is-loan-cob-enabled', 'enabled': 'No', 'value': 'false',
             'description': 'Enable Close of Business for loans'},

            # Penalty Wait Period
            {'config_name': 'penalty-wait-period', 'enabled': 'Yes', 'value': '0',
             'description': 'Days to wait before applying penalty charges'},
        ]
        return pd.DataFrame(data)

    def create_sms_email_config_sheet(self):
        """Create SMS and Email configuration"""
        data = [
            # SMS Gateway Configuration (Twilio)
            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'account_sid',
             'config_value': 'YOUR_TWILIO_ACCOUNT_SID', 'is_active': 'Yes',
             'description': 'Twilio Account SID for SMS service'},

            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'auth_token',
             'config_value': 'YOUR_TWILIO_AUTH_TOKEN', 'is_active': 'Yes',
             'description': 'Twilio Authentication Token (keep secure!)'},

            {'config_type': 'SMS Gateway', 'provider': 'Twilio', 'config_key': 'from_phone_number',
             'config_value': '+237XXXXXXXXX', 'is_active': 'Yes',
             'description': 'Twilio phone number to send SMS from'},

            # SMS Gateway Configuration (Alternative - Infobip)
            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'base_url',
             'config_value': 'https://api.infobip.com', 'is_active': 'No',
             'description': 'Infobip API base URL'},

            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'api_key',
             'config_value': 'YOUR_INFOBIP_API_KEY', 'is_active': 'No',
             'description': 'Infobip API Key'},

            {'config_type': 'SMS Gateway', 'provider': 'Infobip', 'config_key': 'sender_id',
             'config_value': 'MicroFin', 'is_active': 'No',
             'description': 'Sender ID shown to recipients'},

            # Email SMTP Configuration
            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_host',
             'config_value': 'smtp.gmail.com', 'is_active': 'Yes',
             'description': 'SMTP server hostname'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_port',
             'config_value': '587', 'is_active': 'Yes',
             'description': 'SMTP server port (587 for TLS, 465 for SSL)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_username',
             'config_value': 'your-email@gmail.com', 'is_active': 'Yes',
             'description': 'SMTP authentication username (email address)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'smtp_password',
             'config_value': 'YOUR_APP_PASSWORD', 'is_active': 'Yes',
             'description': 'SMTP authentication password (use app password for Gmail)'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'from_email',
             'config_value': 'noreply@yourmfi.com', 'is_active': 'Yes',
             'description': 'Email address shown as sender'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'from_name',
             'config_value': 'Cameroon MicroFinance', 'is_active': 'Yes',
             'description': 'Display name shown as sender'},

            {'config_type': 'Email SMTP', 'provider': 'Gmail', 'config_key': 'use_tls',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable TLS encryption'},

            # Notification Settings
            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'enable_sms_notifications',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable SMS notifications globally'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'enable_email_notifications',
             'config_value': 'Yes', 'is_active': 'Yes',
             'description': 'Enable email notifications globally'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'max_retry_attempts',
             'config_value': '3', 'is_active': 'Yes',
             'description': 'Maximum retry attempts for failed notifications'},

            {'config_type': 'Notification Settings', 'provider': 'System', 'config_key': 'retry_delay_minutes',
             'config_value': '5', 'is_active': 'Yes',
             'description': 'Delay between retry attempts (minutes)'},
        ]
        return pd.DataFrame(data)

    def create_notification_templates_sheet(self):
        """Create notification templates for SMS and Email"""
        data = [
            # Client Notifications
            {'template_name': 'Client Activation', 'channel': 'SMS', 'event_trigger': 'Client Activated',
             'subject': '', 'message_body': 'Hello {clientName}, your account has been activated at {officeName}. Welcome to {institutionName}! Your account number is {accountNumber}.',
             'is_active': 'Yes', 'description': 'Sent when client account is activated'},

            {'template_name': 'Client Activation Email', 'channel': 'Email', 'event_trigger': 'Client Activated',
             'subject': 'Welcome to {institutionName}',
             'message_body': 'Dear {clientName},\n\nYour account has been successfully activated at {officeName}.\n\nAccount Number: {accountNumber}\nActivation Date: {activationDate}\n\nThank you for choosing {institutionName}.\n\nBest regards,\n{institutionName} Team',
             'is_active': 'Yes', 'description': 'Email sent when client account is activated'},

            # Loan Notifications
            {'template_name': 'Loan Approval', 'channel': 'SMS', 'event_trigger': 'Loan Approved',
             'subject': '', 'message_body': 'Congratulations {clientName}! Your loan of {loanAmount} XAF has been approved. Loan Account: {loanAccountNumber}. Visit {officeName} to collect.',
             'is_active': 'Yes', 'description': 'Sent when loan is approved'},

            {'template_name': 'Loan Disbursement', 'channel': 'SMS', 'event_trigger': 'Loan Disbursed',
             'subject': '', 'message_body': 'Dear {clientName}, your loan of {loanAmount} XAF has been disbursed to your account. First repayment due: {firstRepaymentDate}. Total repayments: {numberOfRepayments}.',
             'is_active': 'Yes', 'description': 'Sent when loan is disbursed'},

            {'template_name': 'Loan Repayment Due', 'channel': 'SMS', 'event_trigger': 'Repayment Due in 3 Days',
             'subject': '', 'message_body': 'Reminder: Your loan repayment of {repaymentAmount} XAF is due on {repaymentDate}. Loan Account: {loanAccountNumber}. Thank you.',
             'is_active': 'Yes', 'description': 'Reminder sent 3 days before repayment due date'},

            {'template_name': 'Loan Repayment Overdue', 'channel': 'SMS', 'event_trigger': 'Repayment Overdue',
             'subject': '', 'message_body': 'URGENT: Your loan repayment of {repaymentAmount} XAF was due on {repaymentDate}. Outstanding: {totalOutstanding} XAF. Please pay immediately to avoid penalties.',
             'is_active': 'Yes', 'description': 'Sent when repayment becomes overdue'},

            {'template_name': 'Loan Repayment Received', 'channel': 'SMS', 'event_trigger': 'Repayment Made',
             'subject': '', 'message_body': 'Payment received: {repaymentAmount} XAF for loan {loanAccountNumber}. Outstanding balance: {outstandingBalance} XAF. Next payment: {nextRepaymentDate}.',
             'is_active': 'Yes', 'description': 'Sent when repayment is received'},

            {'template_name': 'Loan Fully Repaid', 'channel': 'SMS', 'event_trigger': 'Loan Closed',
             'subject': '', 'message_body': 'Congratulations {clientName}! Your loan {loanAccountNumber} has been fully repaid. Total paid: {totalPaid} XAF. Thank you for your commitment!',
             'is_active': 'Yes', 'description': 'Sent when loan is fully repaid'},

            # Savings Notifications
            {'template_name': 'Savings Account Activation', 'channel': 'SMS', 'event_trigger': 'Savings Activated',
             'subject': '', 'message_body': 'Hello {clientName}, your savings account {savingsAccountNumber} has been activated. Current balance: {accountBalance} XAF.',
             'is_active': 'Yes', 'description': 'Sent when savings account is activated'},

            {'template_name': 'Savings Deposit', 'channel': 'SMS', 'event_trigger': 'Deposit Made',
             'subject': '', 'message_body': 'Deposit received: {depositAmount} XAF to account {savingsAccountNumber}. New balance: {accountBalance} XAF. Date: {transactionDate}.',
             'is_active': 'Yes', 'description': 'Sent when deposit is made'},

            {'template_name': 'Savings Withdrawal', 'channel': 'SMS', 'event_trigger': 'Withdrawal Made',
             'subject': '', 'message_body': 'Withdrawal: {withdrawalAmount} XAF from account {savingsAccountNumber}. New balance: {accountBalance} XAF. Date: {transactionDate}.',
             'is_active': 'Yes', 'description': 'Sent when withdrawal is made'},

            {'template_name': 'Interest Posted', 'channel': 'SMS', 'event_trigger': 'Interest Posted',
             'subject': '', 'message_body': 'Interest earned: {interestAmount} XAF posted to account {savingsAccountNumber}. New balance: {accountBalance} XAF. Period: {postingPeriod}.',
             'is_active': 'Yes', 'description': 'Sent when interest is posted'},

            {'template_name': 'Low Balance Alert', 'channel': 'SMS', 'event_trigger': 'Balance Below Minimum',
             'subject': '', 'message_body': 'Alert: Your savings account {savingsAccountNumber} balance ({accountBalance} XAF) is below minimum ({minimumBalance} XAF). Please deposit to avoid charges.',
             'is_active': 'Yes', 'description': 'Sent when balance falls below minimum'},

            {'template_name': 'Account Dormancy Warning', 'channel': 'SMS', 'event_trigger': 'Account Inactive 60 Days',
             'subject': '', 'message_body': 'Your account {savingsAccountNumber} has been inactive for 60 days. Please make a transaction within 30 days to prevent dormancy.',
             'is_active': 'Yes', 'description': 'Warning before account becomes dormant'},

            # General Notifications
            {'template_name': 'Password Changed', 'channel': 'Email', 'event_trigger': 'Password Changed',
             'subject': 'Password Changed - {institutionName}',
             'message_body': 'Dear {userName},\n\nYour password was changed on {changeDate} at {changeTime}.\n\nIf you did not make this change, please contact us immediately.\n\nBest regards,\n{institutionName} Security Team',
             'is_active': 'Yes', 'description': 'Security notification for password changes'},

            {'template_name': 'Failed Login Attempt', 'channel': 'Email', 'event_trigger': '3 Failed Login Attempts',
             'subject': 'Security Alert - Failed Login Attempts',
             'message_body': 'Dear {userName},\n\nWe detected 3 failed login attempts on your account from IP: {ipAddress} on {attemptDate}.\n\nIf this was not you, please contact us immediately.\n\nBest regards,\n{institutionName} Security Team',
             'is_active': 'Yes', 'description': 'Security alert for multiple failed logins'},
        ]
        return pd.DataFrame(data)

    def create_data_tables_sheet(self):
        """Create custom data tables (custom fields) configuration"""
        data = [
            # Client Custom Fields
            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'id_type',
             'field_type': 'Dropdown', 'mandatory': 'Yes', 'dropdown_values': 'National ID,Passport,Voter Card,Driver License',
             'description': 'Type of identification document'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'id_number',
             'field_type': 'String', 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Identification document number'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'id_expiry_date',
             'field_type': 'Date', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'ID document expiry date'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'next_of_kin_name',
             'field_type': 'String', 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Next of kin full name'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'next_of_kin_phone',
             'field_type': 'String', 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Next of kin phone number'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'next_of_kin_relationship',
             'field_type': 'Dropdown', 'mandatory': 'Yes', 'dropdown_values': 'Spouse,Parent,Sibling,Child,Friend,Other',
             'description': 'Relationship to next of kin'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'employer_name',
             'field_type': 'String', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Employer name (for salaried clients)'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'employer_phone',
             'field_type': 'String', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Employer contact number'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'years_in_business',
             'field_type': 'Number', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Years in current business (for self-employed)'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'business_location',
             'field_type': 'String', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Business physical location'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'home_ownership',
             'field_type': 'Dropdown', 'mandatory': 'No', 'dropdown_values': 'Owned,Rented,Family Home,Other',
             'description': 'Home ownership status'},

            {'entity_type': 'Client', 'table_name': 'client_additional_info', 'field_name': 'disability_status',
             'field_type': 'Dropdown', 'mandatory': 'No', 'dropdown_values': 'None,Physical,Visual,Hearing,Other',
             'description': 'Disability status (for inclusive finance)'},

            # Loan Custom Fields
            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'loan_purpose_detail',
             'field_type': 'Text', 'mandatory': 'Yes', 'dropdown_values': '',
             'description': 'Detailed description of loan purpose'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'collateral_description',
             'field_type': 'Text', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Description of collateral provided'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'collateral_value',
             'field_type': 'Decimal', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Estimated collateral value (XAF)'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'guarantor_count',
             'field_type': 'Number', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Number of guarantors'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'credit_score',
             'field_type': 'Number', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Credit score (0-100)'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'repayment_source',
             'field_type': 'Dropdown', 'mandatory': 'Yes', 'dropdown_values': 'Business Income,Salary,Remittances,Other',
             'description': 'Primary source of repayment'},

            {'entity_type': 'Loan', 'table_name': 'loan_additional_info', 'field_name': 'previous_loan_history',
             'field_type': 'Dropdown', 'mandatory': 'No', 'dropdown_values': 'First Loan,Good History,Some Delays,Defaulted Before',
             'description': 'Previous loan repayment history'},

            # Savings Custom Fields
            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'savings_goal',
             'field_type': 'String', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Client\'s savings goal'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'target_amount',
             'field_type': 'Decimal', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Target savings amount (XAF)'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'target_date',
             'field_type': 'Date', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Target date to reach goal'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'monthly_commitment',
             'field_type': 'Decimal', 'mandatory': 'No', 'dropdown_values': '',
             'description': 'Monthly savings commitment (XAF)'},

            {'entity_type': 'Savings', 'table_name': 'savings_additional_info', 'field_name': 'preferred_transaction_channel',
             'field_type': 'Dropdown', 'mandatory': 'No', 'dropdown_values': 'Branch,Mobile Money,Bank Transfer,Agent',
             'description': 'Preferred way to transact'},
        ]
        return pd.DataFrame(data)

    def format_excel(self, workbook_path):
        """Apply formatting, descriptions, and data validation to Excel workbook"""
        wb = openpyxl.load_workbook(workbook_path)

        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define column descriptions and validation rules for each sheet
        sheet_metadata = self._get_sheet_metadata()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get metadata for this sheet
            metadata = sheet_metadata.get(sheet_name, {})

            # Format headers and add descriptions
            for idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

                # Add description as comment if available
                col_name = cell.value
                if col_name and col_name in metadata:
                    col_info = metadata[col_name]
                    description = col_info.get('description', '')
                    if description:
                        # Add comment with description
                        comment = Comment(description, 'Fineract Demo Data Generator')
                        comment.width = 300
                        comment.height = 100
                        cell.comment = comment

                # Add data validation for columns with allowed values
                if col_name and col_name in metadata:
                    col_info = metadata[col_name]
                    allowed_values = col_info.get('allowed_values', [])
                    if allowed_values:
                        # Create data validation
                        dv = DataValidation(type="list", formula1=f'"{",".join(allowed_values)}"', allow_blank=True)
                        dv.error = f'Please select from: {", ".join(allowed_values)}'
                        dv.errorTitle = 'Invalid Value'
                        dv.prompt = f'Select from: {", ".join(allowed_values)}'
                        dv.promptTitle = col_name
                        ws.add_data_validation(dv)

                        # Apply validation to entire column (up to row 1000)
                        col_letter = get_column_letter(idx)
                        dv.add(f'{col_letter}2:{col_letter}1000')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze top row
            ws.freeze_panes = 'A2'

        wb.save(workbook_path)

    def _get_sheet_metadata(self):
        """Define descriptions and validation rules for all sheet columns"""
        return {
            'Charges': {
                'charge_name': {
                    'description': 'Name of the charge/fee (e.g., "Loan Processing Fee", "ATM Fee")'
                },
                'charge_type': {
                    'description': 'Entity the charge applies to',
                    'allowed_values': ['Loan', 'Savings', 'Client']
                },
                'charge_time': {
                    'description': 'When the charge is applied. IMPORTANT: Use exact values from the list.\n\nFor Loans: Disbursement, Specified Due Date, Installment Fee, Overdue Installment\n\nFor Savings: Savings Activation, Withdrawal Fee, ATM Fee, Annual Fee, Monthly Fee, Overdraft Fee, Weekly Fee, Saving No Activity Fee',
                    'allowed_values': [
                        'Disbursement', 'Specified Due Date', 'Installment Fee', 'Overdue Installment',
                        'Savings Activation', 'Withdrawal Fee', 'ATM Fee', 'Annual Fee', 'Monthly Fee',
                        'Overdraft Fee', 'Weekly Fee', 'Saving No Activity Fee'
                    ]
                },
                'calculation_type': {
                    'description': 'How the charge amount is calculated',
                    'allowed_values': ['Flat', 'Percentage of Amount', 'Percentage of Interest']
                },
                'amount': {
                    'description': 'Charge amount (flat amount in XAF or percentage value)'
                },
                'currency': {
                    'description': 'Currency code',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                },
                'active': {
                    'description': 'Whether the charge is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Loan Products': {
                'short_name': {
                    'description': 'Short identifier for the product (max 4 characters)'
                },
                'product_name': {
                    'description': 'Full product name displayed to users'
                },
                'currency': {
                    'description': 'Currency code for this product',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                }
            },
            'Savings Products': {
                'short_name': {
                    'description': 'Short identifier for the product (max 4 characters)'
                },
                'product_name': {
                    'description': 'Full product name displayed to users'
                },
                'currency': {
                    'description': 'Currency code for this product',
                    'allowed_values': ['XAF', 'USD', 'EUR']
                },
                'withdrawal_fee_for_transfers': {
                    'description': 'Charge withdrawal fee for transfers',
                    'allowed_values': ['Yes', 'No']
                },
                'overdraft_allowed': {
                    'description': 'Allow account to go negative',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Chart of Accounts': {
                'gl_code': {
                    'description': 'GL account code (OHADA compliant)'
                },
                'gl_name': {
                    'description': 'Name of the GL account'
                },
                'account_type': {
                    'description': 'Type of account in the chart',
                    'allowed_values': ['Asset', 'Liability', 'Equity', 'Income', 'Expense']
                },
                'usage': {
                    'description': 'Header accounts cannot have transactions posted to them',
                    'allowed_values': ['Detail', 'Header']
                },
                'manual_entries': {
                    'description': 'Allow manual journal entries',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Staff': {
                'role': {
                    'description': 'Staff role determines permissions',
                    'allowed_values': ['Branch Manager', 'Loan Officer', 'Cashier']
                }
            },
            'Working Days': {
                'working_day': {
                    'description': 'Is this a working day?',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Account Number Preferences': {
                'entity_type': {
                    'description': 'Entity for account number generation',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Groups', 'Centers']
                },
                'prefix_type': {
                    'description': 'Prefix to use for account numbers',
                    'allowed_values': ['Office Name', 'Client Type', 'Loan Product Short Name', 'Savings Product Short Name']
                }
            },
            'Codes and Values': {
                'is_active': {
                    'description': 'Whether this code value is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Global Configuration': {
                'enabled': {
                    'description': 'Whether this configuration is enabled',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'SMS Email Config': {
                'is_active': {
                    'description': 'Whether this configuration is active',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Notification Templates': {
                'channel': {
                    'description': 'Notification delivery channel',
                    'allowed_values': ['SMS', 'Email', 'System']
                },
                'is_active': {
                    'description': 'Whether this template is active',
                    'allowed_values': ['Yes', 'No']
                },
                'entity_type': {
                    'description': 'Entity this template applies to',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Group']
                }
            },
            'Tellers': {
                'status': {
                    'description': 'Teller operational status',
                    'allowed_values': ['Active', 'Inactive']
                }
            },
            'Scheduler Jobs': {
                'active': {
                    'description': 'Whether this job is enabled',
                    'allowed_values': ['Yes', 'No']
                }
            },
            'Data Tables': {
                'entity_type': {
                    'description': 'Entity the custom field applies to',
                    'allowed_values': ['Client', 'Loan', 'Savings', 'Group']
                },
                'field_type': {
                    'description': 'Data type for the custom field',
                    'allowed_values': ['String', 'Number', 'Decimal', 'Date', 'Datetime', 'Text', 'Dropdown']
                },
                'mandatory': {
                    'description': 'Whether this field is required',
                    'allowed_values': ['Yes', 'No']
                }
            }
        }

    def generate(self):
        """Generate complete Excel template"""
        os.makedirs(self.output_dir, exist_ok=True)

        print(f"Generating Fineract Demo Data Excel Template...")

        # Create Excel writer
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            # Generate all sheets
            print("  Creating Offices sheet...")
            self.create_offices_sheet().to_excel(writer, sheet_name='Offices', index=False)

            print("  Creating Staff sheet...")
            self.create_staff_sheet().to_excel(writer, sheet_name='Staff', index=False)

            print("  Creating Clients sheet...")
            self.create_clients_sheet().to_excel(writer, sheet_name='Clients', index=False)

            print("  Creating Loan Products sheet...")
            self.create_loan_products_sheet().to_excel(writer, sheet_name='Loan Products', index=False)

            print("  Creating Savings Products sheet...")
            self.create_savings_products_sheet().to_excel(writer, sheet_name='Savings Products', index=False)

            print("  Creating Charges sheet...")
            self.create_charges_sheet().to_excel(writer, sheet_name='Charges', index=False)

            print("  Creating Chart of Accounts sheet...")
            self.create_gl_accounts_sheet().to_excel(writer, sheet_name='Chart of Accounts', index=False)

            print("  Creating Loan Accounts sheet...")
            self.create_loan_accounts_sheet().to_excel(writer, sheet_name='Loan Accounts', index=False)

            print("  Creating Savings Accounts sheet...")
            self.create_savings_accounts_sheet().to_excel(writer, sheet_name='Savings Accounts', index=False)

            print("  Creating Fund Sources sheet...")
            self.create_fund_sources_sheet().to_excel(writer, sheet_name='Fund Sources', index=False)

            print("  Creating Payment Types sheet...")
            self.create_payment_types_sheet().to_excel(writer, sheet_name='Payment Types', index=False)

            print("  Creating Holidays sheet...")
            self.create_holidays_sheet().to_excel(writer, sheet_name='Holidays', index=False)

            print("  Creating Loan Provisioning sheet...")
            self.create_loan_provisioning_sheet().to_excel(writer, sheet_name='Loan Provisioning', index=False)

            print("  Creating Collateral Types sheet...")
            self.create_collateral_types_sheet().to_excel(writer, sheet_name='Collateral Types', index=False)

            print("  Creating Guarantor Types sheet...")
            self.create_guarantor_types_sheet().to_excel(writer, sheet_name='Guarantor Types', index=False)

            print("  Creating Configuration sheet...")
            self.create_config_sheet().to_excel(writer, sheet_name='Configuration', index=False)

            print("  Creating Loan Product Accounting Mappings sheet...")
            self.create_loan_product_accounting_sheet().to_excel(writer, sheet_name='Loan Product Accounting', index=False)

            print("  Creating Savings Product Accounting Mappings sheet...")
            self.create_savings_product_accounting_sheet().to_excel(writer, sheet_name='Savings Product Accounting', index=False)

            print("  Creating Payment Type Accounting Mappings sheet...")
            self.create_payment_type_accounting_sheet().to_excel(writer, sheet_name='Payment Type Accounting', index=False)

            print("  Creating Financial Activity Mappings sheet...")
            self.create_financial_activity_mappings_sheet().to_excel(writer, sheet_name='Financial Activity Mapping', index=False)

            print("  Creating Teller/Cashier Mappings sheet...")
            self.create_teller_cashier_mappings_sheet().to_excel(writer, sheet_name='Teller Cashier Mapping', index=False)

            print("  Creating Tellers sheet...")
            self.create_tellers_sheet().to_excel(writer, sheet_name='Tellers', index=False)

            print("  Creating Roles and Permissions sheet...")
            self.create_roles_permissions_sheet().to_excel(writer, sheet_name='Roles Permissions', index=False)

            print("  Creating Maker-Checker Configuration sheet...")
            self.create_maker_checker_config_sheet().to_excel(writer, sheet_name='Maker Checker Config', index=False)

            print("  Creating Currency Configuration sheet...")
            self.create_currency_config_sheet().to_excel(writer, sheet_name='Currency Config', index=False)

            print("  Creating Working Days sheet...")
            self.create_working_days_sheet().to_excel(writer, sheet_name='Working Days', index=False)

            print("  Creating Account Number Preferences sheet...")
            self.create_account_number_preferences_sheet().to_excel(writer, sheet_name='Account Number Preferences', index=False)

            print("  Creating Codes and Values sheet...")
            self.create_codes_and_values_sheet().to_excel(writer, sheet_name='Codes and Values', index=False)

            print("  Creating Scheduler Jobs sheet...")
            self.create_scheduler_jobs_sheet().to_excel(writer, sheet_name='Scheduler Jobs', index=False)

            print("  Creating Global Configuration sheet...")
            self.create_global_config_sheet().to_excel(writer, sheet_name='Global Configuration', index=False)

            print("  Creating SMS/Email Configuration sheet...")
            self.create_sms_email_config_sheet().to_excel(writer, sheet_name='SMS Email Config', index=False)

            print("  Creating Notification Templates sheet...")
            self.create_notification_templates_sheet().to_excel(writer, sheet_name='Notification Templates', index=False)

            print("  Creating Data Tables (Custom Fields) sheet...")
            self.create_data_tables_sheet().to_excel(writer, sheet_name='Data Tables', index=False)

        # Apply formatting
        print("  Applying formatting...")
        self.format_excel(self.filename)

        print(f"\n✓ Excel template generated successfully: {self.filename}")
        print(f"\nSheets created: 33")
        print(f"  1. Offices (4 branches)")
        print(f"  2. Staff (12 members)")
        print(f"  3. Clients (12 clients)")
        print(f"  4. Loan Products (3 products)")
        print(f"  5. Savings Products (3 products)")
        print(f"  6. Charges (11 fees)")
        print(f"  7. Chart of Accounts (30 GL accounts)")
        print(f"  8. Loan Accounts (6 sample loans)")
        print(f"  9. Savings Accounts (12 sample accounts)")
        print(f" 10. Fund Sources (6 sources)")
        print(f" 11. Payment Types (5 channels)")
        print(f" 12. Holidays (12 public holidays)")
        print(f" 13. Loan Provisioning (5 COBAC categories)")
        print(f" 14. Collateral Types (6 types)")
        print(f" 15. Guarantor Types (6 types)")
        print(f" 16. Configuration (19 system settings)")
        print(f" 17. Loan Product Accounting (39 mappings - 13 per product)")
        print(f" 18. Savings Product Accounting (33 mappings - 11 per product)")
        print(f" 19. Payment Type Accounting (5 mappings)")
        print(f" 20. Financial Activity Mapping (6 mappings - inter-branch support)")
        print(f" 21. Teller Cashier Mapping (4 GL account mappings per office)")
        print(f" 22. Tellers (4 tellers - 1 per office)")
        print(f" 23. Roles Permissions (3 roles with granular permissions)")
        print(f" 24. Maker Checker Config (11 dual-authorization rules)")
        print(f" 25. Currency Config (1 currency - XAF)")
        print(f" 26. Working Days (7 days configuration)")
        print(f" 27. Account Number Preferences (5 entity types)")
        print(f" 28. Codes and Values (48 dropdown values across 10 code categories)")
        print(f" 29. Scheduler Jobs (10 automated jobs)")
        print(f" 30. Global Configuration (23 global settings)")
        print(f" 31. SMS/Email Config (17 configuration items - Twilio, Gmail SMTP)")
        print(f" 32. Notification Templates (16 templates - SMS/Email for all events)")
        print(f" 33. Data Tables (24 custom fields - Client, Loan, Savings)")

        return self.filename

if __name__ == '__main__':
    generator = FineractDemoDataGenerator()
    generator.generate()
